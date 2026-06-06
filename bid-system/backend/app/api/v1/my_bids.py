from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import io

from ...database import get_db
from ...models import MyBidRecord, User
from ...schemas import MyBidRecordCreate, MyBidRecordUpdate, MyBidRecordOut, MyBidAnalysisResponse, DefeatAnalysisResponse, GapAnalysisResponse, WinPatternResponse
from ...services import MyBidAnalysisService, DefeatAnalysisService, WinPatternService, SekihaiService
from ...common.security import get_current_user

router = APIRouter(prefix="/my-bids", tags=["투찰이력"])


@router.get("", response_model=List[MyBidRecordOut])
def list_my_bids(
    result: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(MyBidRecord).filter(MyBidRecord.user_id == user.id)
    if result:
        q = q.filter(MyBidRecord.result == result)
    total = q.count()
    items = q.order_by(MyBidRecord.created_at.desc()).offset((page - 1) * size).limit(size).all()
    return items


@router.get("/analysis", response_model=MyBidAnalysisResponse)
def my_bid_analysis(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return MyBidAnalysisService(db).analyze(user.id)


@router.get("/defeat-analysis")
def defeat_analysis(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return DefeatAnalysisService(db).analyze(user.id)


@router.get("/stats")
def my_bid_stats(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rows = db.query(MyBidRecord).filter(MyBidRecord.user_id == user.id).all()
    total = len(rows)
    won   = sum(1 for r in rows if r.result == "won")
    lost  = sum(1 for r in rows if r.result == "lost")

    rates = [float(r.submitted_rate) for r in rows if r.submitted_rate is not None]
    rec_rates = [float(r.recommendation_rate) for r in rows if r.recommendation_rate is not None]
    winner_rates = [float(r.actual_winner_rate) for r in rows
                    if r.actual_winner_rate is not None and r.result in ("won", "lost")]

    rate_diffs = []
    for r in rows:
        if r.recommendation_rate is not None and r.submitted_rate is not None:
            rate_diffs.append(abs(float(r.submitted_rate) - float(r.recommendation_rate)))

    return {
        "total": total,
        "won": won,
        "lost": lost,
        "pending": total - won - lost,
        "win_rate": round(won / max(won + lost, 1), 4),
        "avg_submitted_rate": round(sum(rates) / len(rates), 4) if rates else None,
        "avg_recommendation_rate": round(sum(rec_rates) / len(rec_rates), 4) if rec_rates else None,
        "avg_winner_rate": round(sum(winner_rates) / len(winner_rates), 4) if winner_rates else None,
        "avg_rate_diff_from_rec": round(sum(rate_diffs) / len(rate_diffs), 4) if rate_diffs else None,
    }


@router.get("/gap-analysis", response_model=GapAnalysisResponse)
def gap_analysis(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return DefeatAnalysisService(db).get_gap_distribution(user.id)


@router.get("/win-pattern", response_model=WinPatternResponse)
def win_pattern(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return WinPatternService(db).analyze(user.id)


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    records = (
        db.query(MyBidRecord)
        .filter(MyBidRecord.user_id == user.id)
        .order_by(MyBidRecord.bid_date.desc().nullslast(), MyBidRecord.created_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "투찰이력"

    HEADERS = [
        "공고번호", "공고제목", "발주처", "입찰일", "기초금액",
        "제출투찰률", "추천투찰률", "결과", "실제낙찰률", "격차(제출-낙찰)", "비고",
    ]

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    RESULT_MAP = {"won": "낙찰", "lost": "유찰", "pending": "진행중"}

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1,  value=r.announcement_no or "")
        ws.cell(row=row_idx, column=2,  value=r.title)
        ws.cell(row=row_idx, column=3,  value=r.agency_name or "")
        ws.cell(row=row_idx, column=4,  value=str(r.bid_date) if r.bid_date else "")
        ws.cell(row=row_idx, column=5,  value=r.base_amount or 0)
        ws.cell(row=row_idx, column=6,  value=float(r.submitted_rate) if r.submitted_rate else "")
        ws.cell(row=row_idx, column=7,  value=float(r.recommendation_rate) if r.recommendation_rate else "")
        ws.cell(row=row_idx, column=8,  value=RESULT_MAP.get(r.result or "pending", r.result or ""))
        ws.cell(row=row_idx, column=9,  value=float(r.actual_winner_rate) if r.actual_winner_rate else "")
        ws.cell(row=row_idx, column=10, value=float(r.rate_diff) if r.rate_diff else "")
        ws.cell(row=row_idx, column=11, value=r.note or "")

    COL_WIDTHS = [18, 50, 25, 12, 16, 14, 14, 10, 14, 16, 30]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = quote("투찰이력.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("", response_model=MyBidRecordOut, status_code=201)
def create_my_bid(
    body: MyBidRecordCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rate_diff = None
    if body.actual_winner_rate is not None:
        rate_diff = body.submitted_rate - body.actual_winner_rate

    rec = MyBidRecord(
        user_id=user.id,
        bid_id=body.bid_id,
        title=body.title,
        agency_name=body.agency_name,
        bid_date=body.bid_date,
        base_amount=body.base_amount or 0,
        submitted_rate=body.submitted_rate,
        recommendation_rate=body.recommendation_rate,
        note=body.note,
        announcement_no=body.announcement_no,
        actual_winner_rate=body.actual_winner_rate,
        result=body.result or "pending",
        rate_diff=rate_diff,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return rec


@router.put("/{record_id}", response_model=MyBidRecordOut)
def update_my_bid(
    record_id: int,
    body: MyBidRecordUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rec = db.query(MyBidRecord).filter(
        MyBidRecord.id == record_id,
        MyBidRecord.user_id == user.id,
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="기록을 찾을 수 없습니다.")
    if body.result is not None:
        rec.result = body.result
    if body.actual_winner_rate is not None:
        rec.actual_winner_rate = body.actual_winner_rate
    if body.note is not None:
        rec.note = body.note
    if body.submitted_rate is not None:
        rec.submitted_rate = body.submitted_rate
    # rate_diff 자동계산
    submitted = float(rec.submitted_rate) if rec.submitted_rate is not None else None
    winner = float(rec.actual_winner_rate) if rec.actual_winner_rate is not None else None
    if submitted is not None and winner is not None:
        rec.rate_diff = submitted - winner
    db.commit()
    db.refresh(rec)
    return rec


@router.delete("/{record_id}", status_code=204)
def delete_my_bid(
    record_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rec = db.query(MyBidRecord).filter(
        MyBidRecord.id == record_id,
        MyBidRecord.user_id == user.id,
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="기록을 찾을 수 없습니다.")
    db.delete(rec)
    db.commit()


@router.get("/inpo-rank")
def get_inpo_rank(
    announcement_no: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """announcement_no 기준 inpo21c 실측 순위 조회 (惜敗 분석)."""
    return SekihaiService().get_rank(db, announcement_no)


@router.post("/inpo-rank-batch")
def get_inpo_rank_batch(
    body: dict,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """여러 announcement_no 일괄 inpo21c 순위 조회."""
    announcement_nos = body.get("announcement_nos", [])
    if not isinstance(announcement_nos, list) or len(announcement_nos) > 50:
        raise HTTPException(status_code=400, detail="announcement_nos는 최대 50개.")
    return SekihaiService().batch_ranks(db, announcement_nos)
