"""
비즈니스 로직 서비스 레이어.
Controller(API) -> Service -> Repository(DB) 방향 준수.
"""
import io
import re
import math
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from sqlalchemy.orm import Session
from sqlalchemy import text, func, and_, or_, desc

from .models import (
    Bid, BidResult, Competitor, Agency, Industry, Region,
    FeatureStore, PredictionLog, PredictionLogV2, CompetitorStat, User, AuditLog,
    IndustryFilter, BidBookmark, CollectionLog, MyBidRecord, Notification,
    BidExecution, DefeatAnalysis, AgencyStrategy, RateFrequencyTable, OurCompetitor,
    ActualBidOutcome, ModelPerformanceLog,
)
from .schemas import (
    BidCreate, BidResultCreate, RecommendRequest, RecommendResponse,
    RateRange, WinProbabilities, Explanation, ExplanationFactor, RiskInfo,
    SimilarCase, BidSummary, BidDetail, BidResultOut, CompetitorDetail
)
from .ml.engine import build_features, get_engine, FEATURE_LABELS
from .ml.assessment  import load_srate_stats, predict_srate, compute_market_trend
from .ml.competition import compute_competition_features, get_competitor_profiles, get_market_competitor_distributions
from .ml.simulation  import recommend_with_simulation, simulate_yejung_from_real, simulate_yejung, scan_zones_from_dist
from .ml.rank_model  import get_inpo_raw_rates
from .ml.personal    import PersonalBiasAnalyzer
from .ml.prism       import scan_prism_zones
from .ml.yega        import calc_yega_frequency, get_inpo21c_pattern_direct, load_inpo21c_yega_stats
from .ml.a_value     import calc_floor_rate

logger = logging.getLogger(__name__)


def _compute_yega_ml_features(pos_weights) -> dict:
    """pos_weights(15개 위치별 가중치) → ML 피처 3개 계산."""
    if pos_weights is None:
        return {"top3_freq": None, "entropy": None, "mode_bucket": None}
    w = np.array(pos_weights, dtype=float)
    top3_freq    = float(np.sort(w)[-3:].sum())
    entropy      = float(-np.sum(w * np.log(np.maximum(w, 1e-9))))
    mode_idx     = int(np.argmax(w))
    mode_bucket  = 1 if mode_idx < 5 else (3 if mode_idx >= 10 else 2)
    return {"top3_freq": top3_freq, "entropy": entropy, "mode_bucket": mode_bucket}


# --------------------------------------------------
# 공통 헬퍼
# --------------------------------------------------

def get_active_industry_ids(db: Session):
    """활성화된 공종 ID 목록 반환.
    industry_filters 테이블이 비어있으면 None(전체 허용) 반환.
    설정이 있으면 is_active=True인 ID 목록만 반환."""
    filters = db.query(IndustryFilter).all()
    if not filters:
        return None  # 필터 미설정 = 전체 허용
    return [f.industry_id for f in filters if f.is_active]


def _build_ind_sql(active_ids, alias: str = "b") -> str:
    """활성 공종 SQL WHERE 조건 문자열 생성."""
    if active_ids is None:
        return ""
    if not active_ids:
        return "AND 1=0"
    ids_str = ",".join(map(str, active_ids))
    return f"AND {alias}.industry_id IN ({ids_str})"


# --------------------------------------------------
# 입찰 서비스
# --------------------------------------------------

class BidService:

    def list_bids(
        self, db: Session,
        agency_id=None, industry_id=None, region_id=None,
        status=None, date_from=None, date_to=None,
        keyword=None, page=1, size=20,
        sort_by='notice_date',
    ) -> dict:
        q = db.query(Bid)
        if agency_id:   q = q.filter(Bid.agency_id == agency_id)
        if industry_id: q = q.filter(Bid.industry_id == industry_id)
        if region_id:   q = q.filter(Bid.region_id == region_id)
        if status:      q = q.filter(Bid.status == status)
        if date_from:   q = q.filter(Bid.bid_open_date >= date_from)
        if date_to:     q = q.filter(Bid.bid_open_date < datetime.combine(date_to + timedelta(days=1), datetime.min.time()))
        if keyword:
            q = q.filter(Bid.title.ilike(f"%{keyword}%"))

        # 활성 공종 필터 — 사용자가 특정 공종을 지정하지 않은 경우에만 전역 필터 적용
        if not industry_id:
            active_ids = get_active_industry_ids(db)
            if active_ids is not None:
                if not active_ids:
                    return {"items": [], "total": 0, "page": page, "size": size}
                q = q.filter(Bid.industry_id.in_(active_ids))

        total = q.count()

        if sort_by == 'bid_open_date':
            q = q.order_by(desc(Bid.bid_open_date).nullslast(), desc(Bid.created_at))
        else:
            q = q.order_by(desc(Bid.notice_date).nullslast(), desc(Bid.created_at))

        bids = q.offset((page-1)*size).limit(size).all()

        items = []
        for b in bids:
            winner = next((r for r in b.results if r.is_winner), None)
            items.append({
                "id": b.id, "announcement_no": b.announcement_no,
                "title": b.title,
                "agency_name":   b.agency.name if b.agency else "",
                "industry_name": b.industry.name if b.industry else "",
                "region_name":   b.region.name if b.region else "",
                "base_amount":   b.base_amount,
                "notice_date":   b.notice_date,
                "bid_open_date": b.bid_open_date,
                "status":        b.status,
                "source":        b.source,
                "winner_rate":   float(winner.bid_rate) if winner else None,
                "competitor_count": len(b.results),
            })
        return {"items": items, "total": total, "page": page, "size": size}

    def get_bid_detail(self, db: Session, bid_id: int) -> Optional[dict]:
        b = db.query(Bid).filter(Bid.id == bid_id).first()
        if not b:
            return None
        winner = next((r for r in b.results if r.is_winner), None)
        results = [
            {
                "id": r.id,
                "competitor_id":   r.competitor_id,
                "competitor_name": r.competitor.name if r.competitor else "",
                "bid_amount":      r.bid_amount,
                "bid_rate":        float(r.bid_rate),
                "rank":            r.rank,
                "is_winner":       r.is_winner,
                "assessment_rate": float(r.assessment_rate) if r.assessment_rate else None,
            }
            for r in sorted(b.results, key=lambda x: x.rank)
        ]
        return {
            "id": b.id, "announcement_no": b.announcement_no,
            "title": b.title,
            "agency_name":         b.agency.name if b.agency else "",
            "industry_name":       b.industry.name if b.industry else "",
            "region_name":         b.region.name if b.region else "",
            "base_amount":         b.base_amount,
            "estimated_price":     b.estimated_price,
            "a_value":             b.a_value,
            "min_bid_rate":        float(b.min_bid_rate) if b.min_bid_rate else None,
            "notice_date":         b.notice_date,
            "bid_open_date":       b.bid_open_date,
            "construction_period": b.construction_period,
            "region_restriction":  b.region_restriction,
            "status":              b.status,
            "source":              b.source,
            "ntce_url":            b.ntce_url,
            "winner_rate":         float(winner.bid_rate) if winner else None,
            "competitor_count":    len(b.results),
            "construction_site":   b.construction_site,
            "contract_method":     b.contract_method,
            "bid_method":          b.bid_method,
            "eligible_regions":    b.eligible_regions,
            "industry_limit":      b.industry_limit,
            "bid_close_date":      b.bid_close_date,
            "contact_name":        b.contact_name,
            "contact_tel":         b.contact_tel,
            "results":             results,
        }

    def create_bid(self, db: Session, data: BidCreate, results: List[BidResultCreate] = None) -> Bid:
        bid = Bid(
            announcement_no=data.announcement_no,
            title=data.title,
            agency_id=data.agency_id,
            industry_id=data.industry_id,
            region_id=data.region_id,
            base_amount=data.base_amount,
            min_bid_rate=data.min_bid_rate,
            a_value=data.a_value,
            bid_open_date=data.bid_open_date,
            construction_period=data.construction_period,
            region_restriction=data.region_restriction,
            source="manual",
            status="closed",
        )
        db.add(bid)
        db.flush()

        if results:
            for r in results:
                comp = db.query(Competitor).filter(Competitor.name == r.competitor_name).first()
                if not comp:
                    comp = Competitor(name=r.competitor_name)
                    db.add(comp)
                    db.flush()
                br = BidResult(
                    bid_id=bid.id,
                    competitor_id=comp.id,
                    bid_amount=r.bid_amount,
                    bid_rate=r.bid_rate,
                    rank=r.rank,
                    is_winner=r.is_winner,
                )
                db.add(br)
        db.commit()
        return bid

    def find_similar_bids(self, db: Session, bid_id: int, top_k: int = 8) -> List[dict]:
        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            return []
        rows = db.query(Bid).filter(
            Bid.industry_id == bid.industry_id,
            Bid.region_id == bid.region_id,
            Bid.base_amount.between(int(bid.base_amount * 0.6), int(bid.base_amount * 1.4)),
            Bid.id != bid_id,
            Bid.status == "closed",
        ).order_by(desc(Bid.bid_open_date)).limit(top_k).all()

        results = []
        for b in rows:
            winner = next((r for r in b.results if r.is_winner), None)
            amt_diff = abs(b.base_amount - bid.base_amount) / bid.base_amount
            sim_score = round(1.0 - amt_diff, 3)
            results.append({
                "bid_id": b.id, "title": b.title,
                "agency_name":     b.agency.name if b.agency else "",
                "base_amount":     b.base_amount,
                "bid_open_date":   b.bid_open_date,
                "winner_rate":     float(winner.bid_rate) if winner else None,
                "competitor_count": len(b.results),
                "similarity_score": sim_score,
            })
        return results

    def get_keyword_matches(self, db: Session) -> list:
        """활성 키워드별 매칭 공고 수 + 최근 공고 5건 반환."""
        from .models import WatchKeyword as KW
        keywords = db.query(KW).filter(KW.is_active == True).order_by(KW.created_at.desc()).all()
        result = []
        cutoff_7d = datetime.now() - timedelta(days=7)
        active_ids = get_active_industry_ids(db)
        for kw in keywords:
            if kw.kw_type == "agency":
                q = (db.query(Bid)
                     .join(Agency, Agency.id == Bid.agency_id)
                     .filter(Agency.name.ilike(f"%{kw.keyword}%")))
            else:
                q = db.query(Bid).filter(Bid.title.ilike(f"%{kw.keyword}%"))
            if active_ids is not None and active_ids:
                q = q.filter(Bid.industry_id.in_(active_ids))
            total = q.count()
            new_7d = q.filter(Bid.created_at >= cutoff_7d).count()
            recent = q.order_by(desc(Bid.notice_date)).limit(5).all()
            result.append({
                "keyword_id":  kw.id,
                "keyword":     kw.keyword,
                "kw_type":     kw.kw_type,
                "note":        kw.note,
                "match_count": total,
                "new_7d":      new_7d,
                "recent_bids": [
                    {
                        "id":          b.id,
                        "title":       b.title,
                        "agency_name": db.query(Agency).filter(Agency.id == b.agency_id).with_entities(Agency.name).scalar() or "",
                        "base_amount": b.base_amount,
                        "notice_date": b.notice_date.isoformat() if b.notice_date else None,
                        "status":      b.status,
                    }
                    for b in recent
                ],
            })
        return result

# --------------------------------------------------
# 추천 서비스
# --------------------------------------------------

class RecommendationService:

    def recommend(self, db: Session, req: RecommendRequest, user_id: int = None) -> dict:
        history_df = self._load_history(db, months=24)
        agency    = db.query(Agency).filter(Agency.id == req.agency_id).first()
        industry  = db.query(Industry).filter(Industry.id == req.industry_id).first()
        region    = db.query(Region).filter(Region.id == req.region_id).first()

        features = build_features(
            agency_id=req.agency_id,
            industry_id=req.industry_id,
            region_id=req.region_id,
            base_amount=req.base_amount,
            construction_period=req.construction_period,
            region_restriction=False,
            bid_open_date=datetime.now(),
            historical_df=history_df,
        )

        if req.known_competitor_ids:
            features["expected_competitor_count"] = max(
                features.get("expected_competitor_count", 10),
                len(req.known_competitor_ids)
            )
            features["competitor_strength_score"] = self._calc_competitor_strength(
                db, req.known_competitor_ids
            )

        engine = get_engine()
        result = engine.recommend(features)
        similar = self._find_similar(db, req, top_k=5)
        risk = self._assess_risk(features, result)
        top_factors = self._build_factors(result.get("shap_values", {}), features)
        self._save_log(db, req, result, risk, user_id, features)

        data_count = int(features.get("similar_bid_count", 0)) + int(features.get("agency_bid_count_12m", 0))
        base_rate_val = features.get("agency_avg_rate_12m") or features.get("similar_avg_rate") or 0.879

        return {
            "rate_range": result["rate_range"],
            "win_probabilities": result["win_probabilities"],
            "risk": risk,
            "explanation": {
                "top_factors": top_factors,
                "narrative_ko": result["narrative_ko"],
                "base_rate": round(float(base_rate_val), 4),
                "model_version": result["model_version"],
                "data_count": data_count,
            },
            "similar_cases": similar,
        }

    def _load_history(self, db: Session, months: int = 24) -> pd.DataFrame:
        cutoff = datetime.now() - timedelta(days=months * 30)
        rows = db.execute(text("""
            SELECT b.id, b.agency_id, b.industry_id, b.region_id,
                   b.base_amount, b.bid_open_date,
                   r.bid_rate as winner_rate,
                   (SELECT COUNT(*) FROM bid_results r2 WHERE r2.bid_id = b.id) as competitor_count
            FROM bids b
            LEFT JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
            WHERE b.bid_open_date >= :cutoff AND b.status = 'closed'
        """), {"cutoff": cutoff}).fetchall()

        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=[
            "id","agency_id","industry_id","region_id",
            "base_amount","bid_open_date","winner_rate","competitor_count"
        ])
        df["winner_rate"]      = df["winner_rate"].astype(float)
        df["base_amount"]      = df["base_amount"].astype(float)
        df["competitor_count"] = df["competitor_count"].astype(float)
        return df

    def _calc_competitor_strength(self, db: Session, ids: List[int]) -> float:
        if not ids:
            return 5.0
        rows = db.query(CompetitorStat).filter(
            CompetitorStat.competitor_id.in_(ids),
            CompetitorStat.period_month == None
        ).order_by(desc(CompetitorStat.period_year)).all()
        if not rows:
            return 5.0
        scores = []
        for r in rows:
            ag = float(r.aggression_score or 5)
            wr = float(r.win_rate or 0) * 10
            sc = ag * 0.5 + wr * 0.5
            scores.append(sc)
        return round(float(np.mean(scores)), 2)

    def _find_similar(self, db: Session, req: RecommendRequest, top_k: int) -> List[dict]:
        rows = db.query(Bid).filter(
            Bid.industry_id == req.industry_id,
            Bid.region_id   == req.region_id,
            Bid.base_amount.between(int(req.base_amount * 0.6), int(req.base_amount * 1.4)),
            Bid.status == "closed",
        ).order_by(desc(Bid.bid_open_date)).limit(top_k).all()

        results = []
        for b in rows:
            winner = next((r for r in b.results if r.is_winner), None)
            amt_diff = abs(b.base_amount - req.base_amount) / req.base_amount
            results.append({
                "bid_id": b.id, "title": b.title,
                "agency_name":     b.agency.name if b.agency else "",
                "base_amount":     b.base_amount,
                "bid_open_date":   b.bid_open_date,
                "winner_rate":     float(winner.bid_rate) if winner else None,
                "competitor_count": len(b.results),
                "similarity_score": round(1.0 - amt_diff, 3),
            })
        return results

    def _assess_risk(self, features: dict, result: dict) -> dict:
        score = 0
        factors = []
        cnt = features.get("expected_competitor_count", 10)
        if cnt > 20:
            factors.append(f"높은 경쟁강도 (예상 {cnt}개사)")
            score += 2
        elif cnt > 15:
            score += 1

        rr = result["rate_range"]
        spread = rr["safe_upper"] - rr["safe_lower"]
        if spread > 0.03:
            factors.append(f"투찰률 분산 큼 (범위 {spread:.2%})")
            score += 2
        elif spread > 0.02:
            score += 1

        wp = result["win_probabilities"].get("at_balanced")
        if wp is not None and wp < 0.1:
            factors.append(f"낮은 낙찰 기대확률 ({wp:.1%})")
            score += 2

        sim_cnt = features.get("similar_bid_count", 0) or 0
        if sim_cnt < 5:
            factors.append("유사 입찰 사례 부족 (신뢰도 낮음)")
            score += 1

        if "rule-based" in result.get("model_version",""):
            factors.append("데이터 축적 중 - 규칙 기반 추천")

        level = "LOW" if score <= 1 else "MEDIUM" if score <= 3 else "HIGH"
        return {"level": level, "factors": factors, "score": float(score)}

    def _build_factors(self, shap_vals: dict, features: dict) -> List[dict]:
        if not shap_vals:
            return []
        sorted_items = sorted(shap_vals.items(), key=lambda x: abs(x[1]), reverse=True)[:5]
        factors = []
        for feat, sv in sorted_items:
            fval = features.get(feat)
            factors.append({
                "feature": feat,
                "label":   FEATURE_LABELS.get(feat, feat),
                "value":   round(float(fval), 4) if fval is not None else None,
                "shap_value": round(sv, 6),
                "direction": "positive" if sv > 0 else "negative",
            })
        return factors

    def _save_log(self, db: Session, req, result, risk, user_id, features):
        try:
            rr = result["rate_range"]
            wp = result["win_probabilities"]
            log = PredictionLog(
                bid_id=None, user_id=user_id,
                model_version=result["model_version"],
                input_features={
                    "agency_id": req.agency_id, "industry_id": req.industry_id,
                    "region_id": req.region_id, "base_amount": req.base_amount,
                    "features": {k: float(v) if isinstance(v,(int,float)) else v
                                 for k,v in features.items() if v is not None},
                },
                rate_safe_lower=rr["safe_lower"], rate_lower=rr["lower"],
                rate_center=rr["center"],         rate_upper=rr["upper"],
                rate_safe_upper=rr["safe_upper"],
                win_prob_center=wp.get("at_balanced"),
                risk_level=risk["level"],
                shap_values=result.get("shap_values") or {},
                explanation_text=result["narrative_ko"],
            )
            db.add(log)
            db.commit()
        except Exception as e:
            logger.debug(f"추천 로그 저장 실패: {e}")
            db.rollback()

# --------------------------------------------------
# 경쟁사 서비스
# --------------------------------------------------


class CompetitorService:

    def list_competitors(self, db: Session, keyword=None, page=1, size=20, risk_level=None) -> dict:
        active_ids = get_active_industry_ids(db)
        q = db.query(Competitor)
        if keyword:
            q = q.filter(Competitor.name.ilike(f"%{keyword}%"))

        # 활성 공종 기준: 해당 공종 입찰에 참여한 경쟁사만 포함
        if active_ids is not None:
            if not active_ids:
                return {"items": [], "total": 0, "page": page, "size": size}
            ids_str = ",".join(map(str, active_ids))
            sub = db.execute(text(
                f"SELECT DISTINCT competitor_id FROM bid_results r "
                f"JOIN bids b ON b.id = r.bid_id WHERE b.industry_id IN ({ids_str})"
            )).scalars().all()
            q = q.filter(Competitor.id.in_(sub))

        if risk_level and risk_level.upper() in ("HIGH", "MEDIUM", "LOW"):
            all_comps = q.all()
            filtered = []
            for c in all_comps:
                summary = self._summarize(db, c, active_ids)
                if summary["risk_level"] == risk_level.upper():
                    filtered.append(summary)
            total = len(filtered)
            items = filtered[(page-1)*size : page*size]
        else:
            total = q.count()
            comps = q.offset((page-1)*size).limit(size).all()
            items = [self._summarize(db, c, active_ids) for c in comps]

        return {"items": items, "total": total, "page": page, "size": size}

    def get_detail(self, db: Session, competitor_id: int) -> Optional[dict]:
        c = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not c:
            return None
        active_ids = get_active_industry_ids(db)
        summary = self._summarize(db, c, active_ids)
        summary["frequent_rivals"] = self._frequent_rivals(db, competitor_id, active_ids=active_ids)
        summary["monthly_trend"]   = self._monthly_trend(db, competitor_id, active_ids=active_ids)
        return summary

    def _summarize(self, db: Session, c: Competitor, active_ids=None) -> dict:
        q = db.query(BidResult).filter(BidResult.competitor_id == c.id)
        if active_ids is not None:
            if not active_ids:
                return {
                    "id": c.id, "name": c.name, "total_bids": 0,
                    "win_count": 0, "win_rate": 0.0, "avg_bid_rate": 0.0,
                    "std_bid_rate": 0.0, "p25_rate": 0.0, "p75_rate": 0.0,
                    "aggression_score": 0.0, "consistency_score": 0.0,
                    "risk_level": "UNKNOWN",
                }
            q = q.join(Bid, Bid.id == BidResult.bid_id).filter(Bid.industry_id.in_(active_ids))
        results = q.all()
        if not results:
            return {
                "id": c.id, "name": c.name, "total_bids": 0,
                "win_count": 0, "win_rate": 0.0, "avg_bid_rate": 0.0,
                "std_bid_rate": 0.0, "p25_rate": 0.0, "p75_rate": 0.0,
                "aggression_score": 0.0, "consistency_score": 0.0,
                "risk_level": "UNKNOWN",
            }
        rates = np.array([float(r.bid_rate) for r in results])
        wins  = [r for r in results if r.is_winner]
        agg   = float(np.mean(rates < 0.88) * 10)
        cv    = (rates.std() / rates.mean() * 100) if rates.mean() > 0 else 10
        cons  = max(0.0, 10.0 - cv)

        risk_score = agg * 0.4 + (len(wins)/len(results)*10) * 0.4 + cons * 0.2
        risk = "LOW" if risk_score < 3 else "MEDIUM" if risk_score < 6 else "HIGH"

        return {
            "id": c.id, "name": c.name,
            "total_bids":        len(results),
            "win_count":         len(wins),
            "win_rate":          round(len(wins)/len(results), 4),
            "avg_bid_rate":      round(float(rates.mean()), 4),
            "std_bid_rate":      round(float(rates.std()), 4),
            "p25_rate":          round(float(np.percentile(rates,25)), 4),
            "p75_rate":          round(float(np.percentile(rates,75)), 4),
            "aggression_score":  round(agg, 2),
            "consistency_score": round(cons, 2),
            "risk_level":        risk,
        }

    def _frequent_rivals(self, db: Session, competitor_id: int, top_k: int = 8, active_ids=None) -> List[dict]:
        ind_sql = _build_ind_sql(active_ids)
        rows = db.execute(text(f"""
            SELECT c2.competitor_id, comp.name, COUNT(*) as co_count
            FROM bid_results c1
            JOIN bid_results c2 ON c1.bid_id = c2.bid_id AND c1.competitor_id != c2.competitor_id
            JOIN competitors comp ON comp.id = c2.competitor_id
            JOIN bids b ON b.id = c1.bid_id
            WHERE c1.competitor_id = :cid {ind_sql}
            GROUP BY c2.competitor_id, comp.name
            ORDER BY co_count DESC
            LIMIT :k
        """), {"cid": competitor_id, "k": top_k}).fetchall()
        return [{"competitor_id": r[0], "name": r[1], "co_occurrence": r[2]} for r in rows]

    def _monthly_trend(self, db: Session, competitor_id: int, months: int = 12, active_ids=None) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        ind_sql = _build_ind_sql(active_ids)
        rows = db.execute(text(f"""
            SELECT EXTRACT(YEAR FROM b.bid_open_date)::int  AS yr,
                   EXTRACT(MONTH FROM b.bid_open_date)::int AS mo,
                   COUNT(*)           AS bid_count,
                   SUM(r.is_winner::int) AS win_count,
                   AVG(r.bid_rate)    AS avg_rate
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE r.competitor_id = :cid
              AND b.bid_open_date >= :cutoff {ind_sql}
            GROUP BY yr, mo
            ORDER BY yr, mo
        """), {"cid": competitor_id, "cutoff": cutoff}).fetchall()
        return [
            {"year": r[0], "month": r[1], "bid_count": r[2],
             "win_count": r[3], "avg_rate": round(float(r[4]), 4) if r[4] else None}
            for r in rows
        ]

    def get_win_history(self, db: Session, competitor_id: int, limit: int = 50) -> List[dict]:
        ind_sql = _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT r.id, b.id AS bid_id, b.title, a.name AS agency_name,
                   b.base_amount, b.bid_open_date,
                   r.bid_amount, r.bid_rate, r.rank
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            LEFT JOIN agencies a ON a.id = b.agency_id
            WHERE r.competitor_id = :cid AND r.is_winner = true {ind_sql}
            ORDER BY b.bid_open_date DESC
            LIMIT :lim
        """), {"cid": competitor_id, "lim": limit}).fetchall()
        return [
            {
                "result_id":    r[0],
                "bid_id":       r[1],
                "title":        r[2],
                "agency_name":  r[3] or "",
                "base_amount":  r[4],
                "bid_open_date": r[5].isoformat() if r[5] else None,
                "bid_amount":   r[6],
                "bid_rate":     round(float(r[7]), 4) if r[7] else None,
                "rank":         r[8],
            }
            for r in rows
        ]

# --------------------------------------------------
# 통계 서비스
# --------------------------------------------------

class StatisticsService:

    def overview(self, db: Session, months: int = 12) -> dict:
        cutoff = datetime.now() - timedelta(days=months * 30)
        active_ids = get_active_industry_ids(db)
        ind_sql = _build_ind_sql(active_ids)

        bids_q = db.query(func.count(Bid.id)).filter(Bid.bid_open_date >= cutoff)
        if active_ids is not None:
            if not active_ids:
                return {"total_bids": 0, "total_competitors": 0, "avg_win_rate": 0, "avg_bid_rate": 0, "avg_competitor_count": 0, "monthly_trend": [],
                        "win_rate_change_pct": None, "bid_count_change_pct": None, "avg_competitors_change": None}
            bids_q = bids_q.filter(Bid.industry_id.in_(active_ids))
        total_bids = bids_q.scalar() or 0
        total_comps = db.query(func.count(Competitor.id)).scalar() or 0

        avg_rate_row = db.execute(text(f"""
            SELECT AVG(r.bid_rate)
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE r.is_winner = true AND b.bid_open_date >= :cutoff {ind_sql}
        """), {"cutoff": cutoff}).fetchone()
        avg_win_rate = float(avg_rate_row[0]) if avg_rate_row and avg_rate_row[0] else 0.0

        all_rate_row = db.execute(text(f"""
            SELECT AVG(r.bid_rate)
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE b.bid_open_date >= :cutoff {ind_sql}
        """), {"cutoff": cutoff}).fetchone()
        avg_bid_rate = float(all_rate_row[0]) if all_rate_row and all_rate_row[0] else 0.0

        comp_cnt_row = db.execute(text(f"""
            SELECT AVG(cnt)
            FROM (SELECT bid_id, COUNT(*) as cnt FROM bid_results
                  JOIN bids b ON b.id = bid_results.bid_id
                  WHERE b.bid_open_date >= :cutoff {ind_sql}
                  GROUP BY bid_id) sub
        """), {"cutoff": cutoff}).fetchone()
        avg_comp_count = float(comp_cnt_row[0]) if comp_cnt_row and comp_cnt_row[0] else 0.0

        trend = self._monthly_trend(db, cutoff, ind_sql)

        # ── 전월 대비 변화율 계산 ──
        now = datetime.now()
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_end = this_month_start - timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        def _get_month_stats(start, end):
            bids_m_q = db.query(func.count(Bid.id)).filter(
                Bid.bid_open_date >= start, Bid.bid_open_date <= end
            )
            if active_ids is not None and active_ids:
                bids_m_q = bids_m_q.filter(Bid.industry_id.in_(active_ids))
            bid_count = bids_m_q.scalar() or 0

            wr_row = db.execute(text(f"""
                SELECT AVG(r.bid_rate)
                FROM bid_results r JOIN bids b ON b.id = r.bid_id
                WHERE r.is_winner = true AND b.bid_open_date >= :s AND b.bid_open_date <= :e {ind_sql}
            """), {"s": start, "e": end}).fetchone()
            win_rate = float(wr_row[0]) if wr_row and wr_row[0] else None

            cc_row = db.execute(text(f"""
                SELECT AVG(cnt) FROM (
                    SELECT bid_id, COUNT(*) as cnt FROM bid_results
                    JOIN bids b ON b.id = bid_results.bid_id
                    WHERE b.bid_open_date >= :s AND b.bid_open_date <= :e {ind_sql}
                    GROUP BY bid_id
                ) sub
            """), {"s": start, "e": end}).fetchone()
            avg_comp = float(cc_row[0]) if cc_row and cc_row[0] else None

            return {"bid_count": bid_count, "win_rate": win_rate, "avg_comp": avg_comp}

        this_m = _get_month_stats(this_month_start, now)
        last_m = _get_month_stats(last_month_start, last_month_end)

        def _pct_change(curr, prev):
            if curr is None or prev is None or prev == 0:
                return None
            return round((curr - prev) / prev * 100, 2)

        win_rate_change_pct = _pct_change(this_m["win_rate"], last_m["win_rate"])
        bid_count_change_pct = _pct_change(this_m["bid_count"], last_m["bid_count"])
        avg_competitors_change = (
            round(this_m["avg_comp"] - last_m["avg_comp"], 2)
            if this_m["avg_comp"] is not None and last_m["avg_comp"] is not None
            else None
        )

        return {
            "total_bids": total_bids,
            "total_competitors": total_comps,
            "avg_win_rate": round(avg_win_rate, 4),
            "avg_bid_rate": round(avg_bid_rate, 4),
            "avg_competitor_count": round(avg_comp_count, 1),
            "monthly_trend": trend,
            "win_rate_change_pct": win_rate_change_pct,
            "bid_count_change_pct": bid_count_change_pct,
            "avg_competitors_change": avg_competitors_change,
        }

    def _monthly_trend(self, db: Session, cutoff, ind_sql: str = "") -> List[dict]:
        rows = db.execute(text(f"""
            SELECT EXTRACT(YEAR FROM b.bid_open_date)::int  AS yr,
                   EXTRACT(MONTH FROM b.bid_open_date)::int AS mo,
                   COUNT(DISTINCT b.id)          AS bid_count,
                   AVG(r.bid_rate)               AS avg_rate
            FROM bids b
            LEFT JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
            WHERE b.bid_open_date >= :cutoff {ind_sql}
            GROUP BY yr, mo
            ORDER BY yr, mo
        """), {"cutoff": cutoff}).fetchall()
        return [
            {"year": r[0], "month": r[1], "bid_count": r[2],
             "avg_rate": round(float(r[3]), 4) if r[3] else None}
            for r in rows
        ]

    def agency_stats(self, db: Session, months: int = 12) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        ind_sql = _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT a.id, a.name,
                   COUNT(DISTINCT b.id) AS bid_count,
                   AVG(r.bid_rate)      AS avg_rate,
                   AVG(cnt_sub.cnt)     AS avg_comp
            FROM agencies a
            JOIN bids b ON b.agency_id = a.id
            LEFT JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
            LEFT JOIN (
                SELECT bid_id, COUNT(*) as cnt
                FROM bid_results GROUP BY bid_id
            ) cnt_sub ON cnt_sub.bid_id = b.id
            WHERE b.bid_open_date >= :cutoff {ind_sql}
            GROUP BY a.id, a.name
            ORDER BY bid_count DESC
            LIMIT 20
        """), {"cutoff": cutoff}).fetchall()
        return [
            {"agency_id": r[0], "agency_name": r[1],
             "bid_count": r[2],
             "avg_rate":  round(float(r[3]), 4) if r[3] else None,
             "avg_competitor_count": round(float(r[4]), 1) if r[4] else None}
            for r in rows
        ]

    def rate_distribution(self, db: Session, industry_id: int = None, months: int = 12) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        q_filter = "AND b.industry_id = :iid" if industry_id else _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT FLOOR(r.bid_rate * 1000) / 10 AS rate_pct, COUNT(*) as cnt
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE b.bid_open_date >= :cutoff {q_filter}
            GROUP BY rate_pct
            ORDER BY rate_pct
        """), {"cutoff": cutoff, "iid": industry_id or 0}).fetchall()
        return [{"rate_pct": float(r[0]), "count": r[1]} for r in rows]

    def heatmap(self, db: Session, months: int = 24) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        ind_sql = _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT EXTRACT(MONTH FROM b.bid_open_date)::int AS mo,
                   ind.name,
                   AVG(r.bid_rate) AS avg_rate,
                   COUNT(*)        AS cnt
            FROM bids b
            JOIN industries ind ON ind.id = b.industry_id
            JOIN bid_results r  ON r.bid_id = b.id AND r.is_winner = true
            WHERE b.bid_open_date >= :cutoff {ind_sql}
            GROUP BY mo, ind.name
        """), {"cutoff": cutoff}).fetchall()
        return [
            {"month": r[0], "industry": r[1],
             "avg_rate": round(float(r[2]), 4) if r[2] else None,
             "count": r[3]}
            for r in rows
        ]

    def region_stats(self, db: Session, months: int = 12) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        ind_sql = _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT r.id, r.name,
                   COUNT(DISTINCT b.id)   AS bid_count,
                   AVG(res.bid_rate)      AS avg_rate,
                   SUM(b.base_amount)     AS total_amount
            FROM regions r
            JOIN bids b ON b.region_id = r.id
            LEFT JOIN bid_results res ON res.bid_id = b.id AND res.is_winner = true
            WHERE b.bid_open_date >= :cutoff AND r.parent_id IS NULL {ind_sql}
            GROUP BY r.id, r.name
            ORDER BY bid_count DESC
        """), {"cutoff": cutoff}).fetchall()
        return [
            {
                "region_id":   r[0], "region_name": r[1],
                "bid_count":   r[2],
                "avg_rate":    round(float(r[3]), 4) if r[3] else None,
                "total_amount": int(r[4]) if r[4] else 0,
            }
            for r in rows
        ]

    def industry_stats(self, db: Session, months: int = 12) -> List[dict]:
        cutoff = datetime.now() - timedelta(days=months * 30)
        ind_sql = _build_ind_sql(get_active_industry_ids(db), alias="b")
        rows = db.execute(text(f"""
            SELECT ind.id, ind.name,
                   COUNT(DISTINCT b.id)   AS bid_count,
                   AVG(res.bid_rate)      AS avg_rate,
                   AVG(cnt_sub.cnt)       AS avg_comp,
                   SUM(b.base_amount)     AS total_amount
            FROM industries ind
            JOIN bids b ON b.industry_id = ind.id
            LEFT JOIN bid_results res ON res.bid_id = b.id AND res.is_winner = true
            LEFT JOIN (
                SELECT bid_id, COUNT(*) AS cnt FROM bid_results GROUP BY bid_id
            ) cnt_sub ON cnt_sub.bid_id = b.id
            WHERE b.bid_open_date >= :cutoff {ind_sql}
            GROUP BY ind.id, ind.name
            ORDER BY bid_count DESC
            LIMIT 30
        """), {"cutoff": cutoff}).fetchall()
        return [
            {
                "industry_id":   r[0], "industry_name": r[1],
                "bid_count":     r[2],
                "avg_rate":      round(float(r[3]), 4) if r[3] else None,
                "avg_competitor_count": round(float(r[4]), 1) if r[4] else None,
                "total_amount":  int(r[5]) if r[5] else 0,
            }
            for r in rows
        ]

    def cluster_analysis(self, db: Session, industry_id=None, months: int = 24, k: int = 4) -> dict:
        """유사 입찰 클러스터링 (K-Means)."""
        try:
            from sklearn.cluster import KMeans
            from sklearn.preprocessing import StandardScaler
            from sklearn.impute import SimpleImputer
        except ImportError:
            return {"error": "scikit-learn not available", "clusters": []}

        cutoff = datetime.now() - timedelta(days=months * 30)
        q_filter = "AND b.industry_id = :iid" if industry_id else _build_ind_sql(get_active_industry_ids(db))
        rows = db.execute(text(f"""
            SELECT b.id, b.base_amount, b.industry_id,
                   ind.name AS ind_name,
                   AVG(res.bid_rate)  AS avg_rate,
                   COUNT(res.id)      AS comp_cnt
            FROM bids b
            LEFT JOIN industries ind ON ind.id = b.industry_id
            LEFT JOIN bid_results res ON res.bid_id = b.id
            WHERE b.bid_open_date >= :cutoff AND b.base_amount > 0 {q_filter}
            GROUP BY b.id, b.base_amount, b.industry_id, ind.name
            HAVING COUNT(res.id) > 0
            LIMIT 500
        """), {"cutoff": cutoff, "iid": industry_id or 0}).fetchall()

        if len(rows) < k * 5:
            return {"error": "분석 데이터 부족", "clusters": [], "count": len(rows)}

        X = np.array([
            [math.log10(max(r[1], 1)), float(r[4] or 0.88), float(r[5])]
            for r in rows
        ])
        imp = SimpleImputer(strategy="median")
        X_imp = imp.fit_transform(X)
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_imp)

        km = KMeans(n_clusters=k, random_state=42, n_init=10)
        labels = km.fit_predict(X_scaled)

        clusters = []
        for cid in range(k):
            mask = labels == cid
            cluster_rows = [rows[i] for i in range(len(rows)) if mask[i]]
            amounts = [r[1] for r in cluster_rows]
            rates   = [float(r[4]) for r in cluster_rows if r[4]]
            comps   = [float(r[5]) for r in cluster_rows]
            inds    = {}
            for r in cluster_rows:
                name = r[3] or "기타"
                inds[name] = inds.get(name, 0) + 1
            top_ind = max(inds, key=inds.get) if inds else "기타"
            clusters.append({
                "cluster_id":   cid,
                "count":        int(mask.sum()),
                "avg_amount":   int(np.mean(amounts)),
                "avg_rate":     round(float(np.mean(rates)), 4) if rates else None,
                "avg_comp":     round(float(np.mean(comps)), 1),
                "top_industry": top_ind,
                "amount_range": [int(min(amounts)), int(max(amounts))],
            })
        clusters.sort(key=lambda x: x["avg_amount"])
        if len(clusters) == 4 and all(c["avg_rate"] is not None for c in clusters):
            sorted_by_rate = sorted(range(len(clusters)), key=lambda i: clusters[i]["avg_rate"])
            cluster_labels = [""] * len(clusters)
            cluster_labels[sorted_by_rate[0]]  = "공격형"
            cluster_labels[sorted_by_rate[-1]] = "보수형"
            mid_a, mid_b = sorted_by_rate[1], sorted_by_rate[2]
            if (clusters[mid_a]["avg_comp"] or 0) >= (clusters[mid_b]["avg_comp"] or 0):
                cluster_labels[mid_a] = "중앙집중형"
                cluster_labels[mid_b] = "랜덤형"
            else:
                cluster_labels[mid_b] = "중앙집중형"
                cluster_labels[mid_a] = "랜덤형"
            for i, c in enumerate(clusters):
                c["label"] = cluster_labels[i]
        return {"clusters": clusters, "total_count": len(rows)}

    def srate_distribution_detail(self, db: Session, agency_id=None, industry_id=None, months=24) -> dict:
        from sqlalchemy import func as sqlfunc, cast, Float
        cutoff = datetime.now() - timedelta(days=30 * months)

        # ── 1. assessment_rate_stats 기반 발주처/공종별 통계 (우선)
        agency_stats = industry_stats = global_stats = None
        if agency_id:
            row = db.execute(text("""
                SELECT srate_mean::float, srate_std::float, sample_count,
                       srate_p10::float, srate_p25::float, srate_p50::float,
                       srate_p75::float, srate_p90::float
                FROM assessment_rate_stats
                WHERE group_type='agency' AND group_id=:aid
                ORDER BY updated_at DESC LIMIT 1
            """), {"aid": agency_id}).fetchone()
            if row:
                agency_stats = {
                    "mean": round(row[0], 4), "std": round(row[1], 4),
                    "sample_count": row[2],
                    "p10": round(row[3], 4), "p25": round(row[4], 4),
                    "p50": round(row[5], 4), "p75": round(row[6], 4),
                    "p90": round(row[7], 4),
                }
        if industry_id:
            row = db.execute(text("""
                SELECT srate_mean::float, srate_std::float, sample_count,
                       srate_p25::float, srate_p50::float, srate_p75::float
                FROM assessment_rate_stats
                WHERE group_type='industry' AND group_id=:iid
                ORDER BY updated_at DESC LIMIT 1
            """), {"iid": industry_id}).fetchone()
            if row:
                industry_stats = {
                    "mean": round(row[0], 4), "std": round(row[1], 4),
                    "sample_count": row[2],
                    "p25": round(row[3], 4), "p50": round(row[4], 4), "p75": round(row[5], 4),
                }
        row = db.execute(text("""
            SELECT srate_mean::float, srate_std::float, sample_count,
                   srate_p25::float, srate_p50::float, srate_p75::float
            FROM assessment_rate_stats WHERE group_type='global'
            ORDER BY updated_at DESC LIMIT 1
        """)).fetchone()
        if row:
            global_stats = {
                "mean": round(row[0], 4), "std": round(row[1], 4),
                "sample_count": row[2],
                "p25": round(row[3], 4), "p50": round(row[4], 4), "p75": round(row[5], 4),
            }

        # ── 2. bids.estimated_price 기반 히스토그램 (상세 분포)
        query = db.execute(text("""
            SELECT b.estimated_price::float / b.base_amount AS srate
            FROM bids b
            WHERE b.estimated_price IS NOT NULL
              AND b.base_amount > 0
              AND b.bid_open_date >= :cutoff
              -- 부가세 제외 고정비율(base×10/11≈0.9091) 공고 제거
              AND ABS(b.estimated_price::numeric / NULLIF(b.base_amount,0) - (10.0/11.0)) > 0.002
              AND (:agency_id  IS NULL OR b.agency_id   = :agency_id)
              AND (:industry_id IS NULL OR b.industry_id = :industry_id)
        """), {"cutoff": cutoff, "agency_id": agency_id, "industry_id": industry_id})
        rows = query.fetchall()
        values = [float(r[0]) for r in rows if r[0] and 0.80 <= float(r[0]) <= 1.05]

        # 히스토그램 데이터 생성 (bids 데이터 또는 stats 기반 합성)
        if values:
            arr = np.array(values)
            src_mean = float(np.mean(arr))
            src_std  = float(np.std(arr))
            src_n    = len(values)
        elif agency_stats:
            # bids.estimated_price 없으면 assessment_rate_stats 기반 정규 분포 합성
            src_mean = agency_stats["mean"]
            src_std  = agency_stats["std"]
            src_n    = min(agency_stats["sample_count"], 200)
            rng = np.random.default_rng(42)
            arr = rng.normal(src_mean, max(src_std, 0.005), src_n)
            arr = arr[(arr >= 0.80) & (arr <= 1.10)]
            values = arr.tolist()
        elif global_stats:
            src_mean = global_stats["mean"]
            src_std  = global_stats["std"]
            src_n    = 0
            arr = np.array([src_mean])
            values = [src_mean]
        else:
            return {
                "bins": [], "mode": None, "mean": None, "std": None, "sample_count": 0,
                "p25": None, "p50": None, "p75": None,
                "agency_stats": None, "industry_stats": None, "global_stats": global_stats,
            }

        arr = np.array(values)
        bins_range = np.arange(0.800, 1.101, 0.001)
        counts, edges = np.histogram(arr, bins=bins_range)
        bins = [{"rate_pct": round(float(edges[i]), 4), "count": int(counts[i])} for i in range(len(counts))]
        mode_idx = int(np.argmax(counts))
        return {
            "bins": bins,
            "mode": round(float(edges[mode_idx]), 4) if counts[mode_idx] > 0 else None,
            "mean": round(float(np.mean(arr)), 4),
            "std":  round(float(np.std(arr)),  4),
            "p25":  round(float(np.percentile(arr, 25)), 4),
            "p50":  round(float(np.percentile(arr, 50)), 4),
            "p75":  round(float(np.percentile(arr, 75)), 4),
            "sample_count": len(values),
            # 발주처별 세분화 통계
            "agency_stats":   agency_stats,
            "industry_stats": industry_stats,
            "global_stats":   global_stats,
        }

    def model_info(self, db: Session, months: int = 12) -> dict:
        import json, os
        from pathlib import Path
        MODEL_DIR = Path(os.getenv("ML_MODELS_PATH", "/app/ml_models"))
        meta_path = MODEL_DIR / "meta.json"
        if meta_path.exists():
            with open(meta_path) as f:
                meta = json.load(f)
        else:
            meta = None

        total_results = db.execute(text("SELECT COUNT(*) FROM bid_results")).scalar() or 0
        winner_results = db.execute(text("SELECT COUNT(*) FROM bid_results WHERE is_winner=true")).scalar() or 0
        recent_preds = db.execute(text(
            "SELECT COUNT(*) FROM prediction_logs WHERE created_at >= NOW() - INTERVAL '30 days'"
        )).scalar() or 0

        cutoff = datetime.now() - timedelta(days=months * 30)
        period_results = db.execute(text(
            "SELECT COUNT(*) FROM bid_results r JOIN bids b ON b.id = r.bid_id WHERE b.bid_open_date >= :cutoff"
        ), {"cutoff": cutoff}).scalar() or 0
        period_winners = db.execute(text(
            "SELECT COUNT(*) FROM bid_results r JOIN bids b ON b.id = r.bid_id WHERE r.is_winner=true AND b.bid_open_date >= :cutoff"
        ), {"cutoff": cutoff}).scalar() or 0
        period_preds = db.execute(text(
            "SELECT COUNT(*) FROM prediction_logs WHERE created_at >= :cutoff"
        ), {"cutoff": cutoff}).scalar() or 0

        return {
            "model": meta or {"version": "rule-based-v1", "train_size": 0, "winner_size": 0},
            "data_availability": {
                "total_results":  total_results,
                "winner_results": winner_results,
                "ready_for_ml":   winner_results >= 20,
            },
            "period_data": {
                "results": period_results,
                "winners": period_winners,
                "months":  months,
            },
            "usage": {
                "predictions_30d":    recent_preds,
                "predictions_period": period_preds,
            },
        }

# ==================================================
# 사정율 트렌드 서비스
# ==================================================

class SrateTrendService:
    """발주처×공종 최근 3개월 vs 이전 3개월 사정율 트렌드 분석."""

    THRESHOLD = 0.002  # ±0.2%p 이상이면 상승/하락

    def get_trend(self, db: Session, agency_id: Optional[int], industry_id: Optional[int]) -> dict:
        rows = self._from_assessment_stats(db, agency_id, industry_id)
        if not rows:
            rows = self._from_bid_results(db, agency_id, industry_id)
        return self._build_result(rows, datetime.now())

    def get_top_trends(self, db: Session, limit: int = 3) -> list:
        sql_rows = db.execute(text("""
            SELECT ars.group_id, a.name,
                   ars.period_year, ars.period_month,
                   ars.srate_mean::float, ars.sample_count
            FROM assessment_rate_stats ars
            JOIN agencies a ON a.id = ars.group_id
            WHERE ars.group_type = 'agency' AND ars.period_month IS NOT NULL
              AND make_date(ars.period_year::int, ars.period_month::int, 1) >= NOW() - INTERVAL '12 months'
            ORDER BY ars.group_id, ars.period_year, ars.period_month
        """)).fetchall()

        agency_data: dict = {}
        for r in sql_rows:
            gid = r[0]
            if gid not in agency_data:
                agency_data[gid] = {"agency_id": gid, "agency_name": r[1], "rows": []}
            agency_data[gid]["rows"].append({
                "period_year": r[2], "period_month": r[3],
                "srate_mean": r[4], "sample_count": r[5] or 0,
            })

        now = datetime.now()
        results = []
        for gid, info in agency_data.items():
            trend = self._build_result(info["rows"], now)
            if trend["direction"] != "stable":
                results.append({"agency_id": gid, "agency_name": info["agency_name"], **trend})

        results.sort(key=lambda x: abs(x["delta"]), reverse=True)
        return results[:limit]

    def _from_assessment_stats(self, db: Session, agency_id: Optional[int], industry_id: Optional[int]) -> list:
        if agency_id:
            group_type, gid = "agency", agency_id
        elif industry_id:
            group_type, gid = "industry", industry_id
        else:
            group_type, gid = "global", None

        if gid is not None:
            rows = db.execute(text("""
                SELECT period_year, period_month, srate_mean::float, sample_count
                FROM assessment_rate_stats
                WHERE group_type = :gt AND group_id = :gid AND period_month IS NOT NULL
                  AND make_date(period_year::int, period_month::int, 1) >= NOW() - INTERVAL '12 months'
                ORDER BY period_year, period_month
            """), {"gt": group_type, "gid": gid}).fetchall()
        else:
            rows = db.execute(text("""
                SELECT period_year, period_month, srate_mean::float, sample_count
                FROM assessment_rate_stats
                WHERE group_type = 'global' AND period_month IS NOT NULL
                  AND make_date(period_year::int, period_month::int, 1) >= NOW() - INTERVAL '12 months'
                ORDER BY period_year, period_month
            """)).fetchall()

        return [
            {"period_year": r[0], "period_month": r[1], "srate_mean": r[2], "sample_count": r[3] or 0}
            for r in rows
        ]

    def _from_bid_results(self, db: Session, agency_id: Optional[int], industry_id: Optional[int]) -> list:
        rows = db.execute(text("""
            SELECT
                EXTRACT(YEAR FROM b.bid_open_date)::int,
                EXTRACT(MONTH FROM b.bid_open_date)::int,
                AVG(r.assessment_rate)::float,
                COUNT(*)::int
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE r.assessment_rate IS NOT NULL
              AND b.bid_open_date >= NOW() - INTERVAL '12 months'
              AND (:agency_id IS NULL OR b.agency_id = :agency_id)
              AND (:industry_id IS NULL OR b.industry_id = :industry_id)
            GROUP BY 1, 2
            ORDER BY 1, 2
        """), {"agency_id": agency_id, "industry_id": industry_id}).fetchall()
        return [
            {"period_year": r[0], "period_month": r[1], "srate_mean": r[2], "sample_count": r[3]}
            for r in rows
        ]

    def _build_result(self, rows: list, now: datetime) -> dict:
        if not rows:
            return {
                "direction": "stable", "delta": 0.0,
                "recent_mean": 0.0, "prev_mean": None,
                "sample_count": 0, "signal": "데이터 부족으로 트렌드 분석 불가",
            }

        recent, prev = [], []
        for row in rows:
            months_ago = (now.year - row["period_year"]) * 12 + (now.month - row["period_month"])
            if months_ago < 6:
                recent.append(row)
            elif months_ago < 12:
                prev.append(row)

        if not recent:
            # fallback: 데이터 분포 기반 전반/후반 분할
            all_rows = sorted(rows, key=lambda r: (r["period_year"], r["period_month"]))
            if len(all_rows) >= 2:
                mid = len(all_rows) // 2
                prev = all_rows[:mid]
                recent = all_rows[mid:]
            else:
                return {
                    "direction": "stable", "delta": 0.0,
                    "recent_mean": 0.0, "prev_mean": None,
                    "sample_count": 0, "signal": "최근 데이터 없음",
                }

        total_recent = sum(r["sample_count"] for r in recent)
        if total_recent > 0:
            recent_mean = sum(r["srate_mean"] * r["sample_count"] for r in recent) / total_recent
        else:
            recent_mean = sum(r["srate_mean"] for r in recent) / len(recent)
        sample_count = total_recent or len(recent)

        if not prev:
            return {
                "direction": "stable", "delta": 0.0,
                "recent_mean": round(recent_mean, 4), "prev_mean": None,
                "sample_count": sample_count, "signal": "이전 기간 데이터 부족",
            }

        total_prev = sum(r["sample_count"] for r in prev)
        if total_prev > 0:
            prev_mean = sum(r["srate_mean"] * r["sample_count"] for r in prev) / total_prev
        else:
            prev_mean = sum(r["srate_mean"] for r in prev) / len(prev)

        delta = recent_mean - prev_mean

        if delta > self.THRESHOLD:
            direction = "up"
            signal = f"최근 3개월 사정율 +{delta*100:.2f}%p 상승 중 → 균형형 이상 추천"
        elif delta < -self.THRESHOLD:
            direction = "down"
            signal = f"최근 3개월 사정율 {delta*100:.2f}%p 하락 중 → 안정형 또는 낮게 입찰 추천"
        else:
            direction = "stable"
            signal = f"사정율 변화 미미 ({delta*100:+.2f}%p) → 균형형 전략 유지"

        return {
            "direction": direction,
            "delta": round(delta, 6),
            "recent_mean": round(recent_mean, 4),
            "prev_mean": round(prev_mean, 4),
            "sample_count": sample_count,
            "signal": signal,
        }


# ==================================================
# 하이브리드 추천 서비스 (v2)
# ==================================================

class HybridRecommendService:
    """
    4개 엔진 앙상블:
      A: 사정율 예측 (LightGBM Quantile)
      B: 역사 패턴 낙찰률 예측 (기존 XGBoost)
      C: 동적 경쟁강도 분석
      D: 시장 변동성/추세 분석
    """

    def recommend_v2(self, db: Session, req, user_id: int = None) -> dict:

        bid_date = getattr(req, "bid_open_date", None) or datetime.now()
        min_bid_rate = getattr(req, "min_bid_rate", 0.87745)

        # bid_id 제공 시 실제 공고 낙찰하한율 자동 적용 [2순위 개선]
        _bid_id = getattr(req, "bid_id", None)
        if _bid_id:
            try:
                _bid_row = db.execute(text(
                    "SELECT min_bid_rate, bid_open_date FROM bids WHERE id = :bid_id"
                ), {"bid_id": _bid_id}).fetchone()
                if _bid_row:
                    if _bid_row[0] and float(_bid_row[0]) > 0.5:
                        min_bid_rate = max(min_bid_rate, float(_bid_row[0]))
                    if _bid_row[1] and not getattr(req, "bid_open_date", None):
                        bid_date = _bid_row[1]
            except Exception as _e:
                logger.debug("bid_id floor_rate 조회 실패 (무시): %s", _e)

        # ── Step 1: 역사 데이터 로드 (Engine B용)
        history_df = self._load_history(db, months=24)

        # ── Step 2: Engine A+D 피처 생성
        features_a = load_srate_stats(
            db, req.agency_id, req.industry_id, req.region_id,
            req.base_amount, bid_date
        )

        # ── Step 3: Engine B 피처 생성 (기존)
        features_b = build_features(
            agency_id=req.agency_id,
            industry_id=req.industry_id,
            region_id=req.region_id,
            base_amount=req.base_amount,
            construction_period=getattr(req,"construction_period",None),
            region_restriction=False,
            bid_open_date=bid_date,
            historical_df=history_df,
        )

        # ── Step 4: Engine C 동적 경쟁강도
        features_c = compute_competition_features(
            db, req.agency_id, req.industry_id, req.base_amount,
            bid_date, req.known_competitor_ids or []
        )
        # Engine C 결과를 B 피처에 주입 (기존 고정값 대체)
        features_b["expected_competitor_count"] = features_c["expected_competitor_count"]
        features_b["competitor_strength_score"] = features_c["competitor_strength_score"]

        # ── Step 5: Engine D 시장 변동성
        features_d = compute_market_trend(db, req.agency_id, req.industry_id, bid_date)

        # ── Step 6: Engine A 사정율 예측
        ep_result = predict_srate(features_a, req.base_amount)
        srate_c   = ep_result["srate_range"]["center"]
        ep_conf   = ep_result["confidence"]

        # 사정율 예측 결과를 B 피처에 주입
        features_b["srate_pred_center"] = srate_c

        # ── Step 7: Engine B 역사 패턴 예측
        engine = get_engine()
        b_result = engine.recommend(features_b)
        b_center = b_result["rate_range"]["center"]
        b_lower  = b_result["rate_range"]["lower"]
        b_upper  = b_result["rate_range"]["upper"]

        # ── Step 8: 앙상블 결합
        w_a = min(0.40, ep_conf * 0.45)
        w_b = 1.0 - w_a

        pressure  = features_c.get("market_pressure_index", 0.4)
        comp_adj  = -(pressure - 0.5) * 0.003
        trend_adj = features_d.get("trend_adjustment", 0.0)

        srate_lo = ep_result["srate_range"]["lower"]
        srate_hi = ep_result["srate_range"]["upper"]

        ens_center = (srate_c * w_a + b_center * w_b) + comp_adj + trend_adj
        ens_lower  = (srate_lo * w_a + b_lower  * w_b) + comp_adj + trend_adj
        ens_upper  = (srate_hi * w_a + b_upper  * w_b) + comp_adj + trend_adj

        # 낙찰하한율 이상 강제
        hard_floor = max(features_c["expected_floor_rate"], min_bid_rate)
        ens_lower  = max(ens_lower,  hard_floor + 0.0005)
        ens_center = max(ens_center, ens_lower  + 0.0005)
        ens_upper  = max(ens_upper,  ens_center + 0.0005)

        # ── Step 8.5: 개인화 편향 보정 (Monte Carlo 전 적용)
        personal_info = {"correction": 0.0, "agency_correction": None,
                         "confidence": 0.0, "direction": "balanced",
                         "avg_bias_pct": 0.0, "sample_count": 0, "narrative": ""}
        if user_id:
            try:
                agency_name_for_bias = None
                _ag = db.query(Agency).filter(Agency.id == req.agency_id).first()
                if _ag:
                    agency_name_for_bias = _ag.name
                personal_info = PersonalBiasAnalyzer().compute(
                    db, user_id,
                    agency_name=agency_name_for_bias,
                )
                # 신뢰도 비례 보정값 (낮은 신뢰도면 보정 축소)
                effective_corr = personal_info["correction"] * personal_info["confidence"]
                # 발주처 특화 보정 병합
                if personal_info.get("agency_correction") is not None:
                    agency_conf = min(1.0, len([1]) * personal_info["confidence"])
                    effective_corr = (effective_corr * 0.6 +
                                      personal_info["agency_correction"] * personal_info["confidence"] * 0.4)
                ens_center += effective_corr
                ens_lower  += effective_corr
                ens_upper  += effective_corr
                # 낙찰하한율 재확인
                ens_lower  = max(ens_lower,  hard_floor + 0.0005)
                ens_center = max(ens_center, ens_lower  + 0.0005)
                ens_upper  = max(ens_upper,  ens_center + 0.0005)
            except Exception as _e:
                logger.debug(f"개인화 보정 실패 (무시): {_e}")

        # ── Step 9-pre: 경쟁사 프로파일 (시뮬레이션 입력용)
        comp_profiles = get_competitor_profiles(
            db, req.known_competitor_ids or [],
            req.agency_id, req.industry_id, bid_date
        )

        # ── Step 9: 복수예가 Monte Carlo 시뮬레이션 기반 4전략
        industry_name = self._get_industry_name(db, req.industry_id)
        srate_std_val = features_a.get("agency_srate_std") or 0.012
        expected_n = max(3, min(features_c.get("expected_competitor_count", 8), 15))
        if comp_profiles:
            comp_means = [p["avg_rate"] for p in comp_profiles]
            comp_stds  = [max(p["std_rate"], 0.003) for p in comp_profiles]
        else:
            comp_means, comp_stds = get_market_competitor_distributions(
                db, req.agency_id, req.industry_id, bid_date
            )
        # 기대 경쟁업체 수에 맞게 상위 N개만 사용
        comp_means = comp_means[:expected_n]
        comp_stds  = comp_stds[:expected_n]
        # inpo21c 실증 분포 (데이터 있으면 합성 정규분포 대체)
        inpo_rates = None
        try:
            inpo_rates = get_inpo_raw_rates(db, expected_n)
        except Exception as _e:
            logger.debug("inpo21c 실증 분포 조회 실패 (무시): %s", _e)

        _yega_stats  = load_inpo21c_yega_stats(db, req.agency_id or 0)
        _pos_weights = _yega_stats.get("pos_weights")

        # 복수예가 패턴 ML 피처 주입 (Engine B)
        _yega_ml = _compute_yega_ml_features(_pos_weights)
        features_b["yega_top3_freq"]   = _yega_ml["top3_freq"]
        features_b["yega_entropy"]     = _yega_ml["entropy"]
        features_b["yega_mode_bucket"] = _yega_ml["mode_bucket"]

        sim_result = recommend_with_simulation(
            base_amount=req.base_amount,
            industry_name=industry_name,
            srate_center=srate_c,
            srate_std=srate_std_val,
            competitor_means=comp_means,
            competitor_stds=comp_stds,
            hard_floor=hard_floor,
            ens_center=ens_center,
            ens_upper=ens_upper,
            empirical_comp_rates=inpo_rates,
            expected_n_comp=expected_n if inpo_rates is not None else 0,
            pos_weights=_pos_weights,
        )
        strategies = sim_result["strategies"]

        # ── Step 10: 리스크 평가
        hhi       = features_c.get("hhi_score", 0.1)
        agg_ratio = features_c.get("aggressive_comp_ratio", 0.2)
        risk      = self._compute_risk_v2(ep_conf, pressure, hhi, agg_ratio, b_result)

        # ── Step 11: 낙찰 확률 (Monte Carlo)
        win_probs = sim_result["win_probabilities"]

        # ── Step 12: 유사 사례
        similar = self._find_similar(db, req)

        # ── Step 14: SHAP 통합 설명
        explanation = self._build_explanation_v2(
            features_a, features_b, features_c, features_d,
            b_result, ep_result, w_a, w_b
        )

        # ── Step 15: 로그 저장
        self._save_log_v2(db, req, ens_center, ens_lower, ens_upper,
                          ep_result, features_c, risk, explanation, user_id, w_a, w_b)

        return {
            "rate_range": {
                "safe_lower": round(ens_lower  - 0.003, 4),
                "lower":      round(ens_lower,   4),
                "center":     round(ens_center,  4),
                "upper":      round(ens_upper,   4),
                "safe_upper": round(ens_upper  + 0.003, 4),
            },
            "strategies":    strategies,
            "estimated_price": ep_result,
            "win_probabilities": win_probs,
            "risk":          risk,
            "competition": {
                "score":                round(features_c["competitor_strength_score"], 2),
                "pressure":             round(pressure, 4),
                "hhi":                  round(hhi, 4),
                "expected_competitors": features_c["expected_competitor_count"],
                "floor_rate":           round(hard_floor, 4),
                "aggressive_ratio":     round(agg_ratio, 4),
                "recent_winner_min":    round(features_c.get("recent_winner_min_rate", hard_floor), 4),
                "profiles":             comp_profiles,
            },
            "ensemble_weights": {"engine_a": round(w_a, 3), "engine_b": round(w_b, 3)},
            "explanation":   explanation,
            "similar_cases": similar,
            "market_trend":  features_d,
            "simulation":    sim_result["simulation"],
            "personal_correction": personal_info,
        }

    # ──────────────────────────────────────────
    # 내부 메서드
    # ──────────────────────────────────────────

    def _get_industry_name(self, db: Session, industry_id: int) -> str:
        from .models import Industry
        ind = db.query(Industry).filter(Industry.id == industry_id).first()
        return ind.name if ind else ""

    def _load_history(self, db: Session, months: int = 24) -> "pd.DataFrame":
        import pandas as pd
        cutoff = datetime.now() - timedelta(days=months * 30)
        rows = db.execute(text("""
            SELECT b.id, b.agency_id, b.industry_id, b.region_id,
                   b.base_amount, b.bid_open_date,
                   r.bid_rate AS winner_rate,
                   (SELECT COUNT(*) FROM bid_results r2 WHERE r2.bid_id = b.id) AS competitor_count
            FROM bids b
            LEFT JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
            WHERE b.bid_open_date >= :cutoff AND b.status = 'closed'
        """), {"cutoff": cutoff}).fetchall()
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=[
            "id","agency_id","industry_id","region_id",
            "base_amount","bid_open_date","winner_rate","competitor_count"
        ])
        df["winner_rate"]      = pd.to_numeric(df["winner_rate"],      errors="coerce")
        df["base_amount"]      = pd.to_numeric(df["base_amount"],      errors="coerce")
        df["competitor_count"] = pd.to_numeric(df["competitor_count"], errors="coerce")
        return df

    def _compute_risk_v2(self, ep_conf, pressure, hhi, agg_ratio, b_result) -> dict:
        score = (
            (1.0 - ep_conf)  * 2.5 +
            pressure         * 2.5 +
            hhi              * 2.5 +
            agg_ratio        * 2.5
        )
        level = "LOW" if score < 3 else "MEDIUM" if score < 6 else "HIGH"
        factors = []
        if ep_conf < 0.4:   factors.append(f"사정율 예측 신뢰도 낮음({ep_conf:.0%}) — 기관 데이터 부족")
        if pressure > 0.6:  factors.append(f"시장 압박 지수 높음({pressure:.2f}) — 공격적 경쟁사 다수")
        if hhi > 0.30:      factors.append(f"낙찰 집중도 높음(HHI={hhi:.3f}) — 특정 업체 우세")
        if agg_ratio > 0.4: factors.append(f"공격적 업체 비율 {agg_ratio:.0%} — 저가 투찰 경쟁 심화")
        # 기존 Engine B 리스크도 참고
        spread = b_result["rate_range"]["safe_upper"] - b_result["rate_range"]["safe_lower"]
        if spread > 0.03:   factors.append(f"낙찰률 분산 큼(범위 {spread:.2%}) — 예측 불확실성 높음")
        return {"level": level, "score": round(score, 2), "factors": factors}

    def _find_similar(self, db: Session, req) -> List[dict]:
        rows = db.query(Bid).filter(
            Bid.industry_id == req.industry_id,
            Bid.region_id   == req.region_id,
            Bid.base_amount.between(int(req.base_amount * 0.6), int(req.base_amount * 1.4)),
            Bid.status == "closed",
        ).order_by(desc(Bid.bid_open_date)).limit(5).all()
        result = []
        for b in rows:
            winner = next((r for r in b.results if r.is_winner), None)
            amt_diff = abs(b.base_amount - req.base_amount) / max(req.base_amount, 1)
            result.append({
                "bid_id": b.id, "title": b.title,
                "agency_name":     b.agency.name if b.agency else "",
                "base_amount":     b.base_amount,
                "bid_open_date":   b.bid_open_date,
                "winner_rate":     float(winner.bid_rate) if winner else None,
                "competitor_count": len(b.results),
                "similarity_score": round(max(0, 1.0 - amt_diff), 3),
            })
        return result

    def _build_explanation_v2(self, fa, fb, fc, fd, b_result, ep_result, w_a, w_b) -> dict:
        from .ml.engine import FEATURE_LABELS
        shap_b = b_result.get("shap_values", {})
        top_factors = []
        for feat, val in sorted(shap_b.items(), key=lambda x: abs(x[1]), reverse=True)[:5]:
            top_factors.append({
                "feature":    feat,
                "label":      FEATURE_LABELS.get(feat, feat),
                "shap_value": round(val, 6),
                "direction":  "positive" if val > 0 else "negative",
                "value":      round(float(fb.get(feat)), 4) if fb.get(feat) is not None else None,
            })

        parts = []
        srate_c = ep_result["srate_range"]["center"]
        ep_conf = ep_result["confidence"]
        if fa.get("agency_srate_mean"):
            parts.append(
                f"이 기관의 과거 사정율 평균은 {fa['agency_srate_mean']:.4f}이며, "
                f"예정가격은 기초금액의 약 {srate_c*100:.1f}% 수준으로 예상됩니다."
            )
        if ep_conf < 0.4:
            parts.append("사정율 예측 데이터가 부족하여 역사 패턴 비중을 높였습니다.")
        elif ep_conf > 0.7:
            parts.append(f"사정율 신뢰도 {ep_conf:.0%} — Engine A 가중치를 높여 적용했습니다.")

        pressure = fc.get("market_pressure_index", 0.4)
        if pressure > 0.6:
            parts.append(f"경쟁 압박 지수 {pressure:.2f} — 최근 공격적 투찰 업체 비율이 높습니다.")
        elif pressure < 0.3:
            parts.append(f"경쟁 압박 지수 낮음({pressure:.2f}) — 비교적 여유 있는 경쟁 환경입니다.")

        if fd.get("has_recent_data") and abs(fd.get("srate_4w_change", 0)) > 0.001:
            chg = fd["srate_4w_change"]
            direction = "상승" if chg > 0 else "하락"
            parts.append(f"최근 4주 사정율 {direction} 추세({chg:+.4f}) 반영.")

        for f in top_factors[:3]:
            direction = "높이는" if f["shap_value"] > 0 else "낮추는"
            parts.append(f"'{f['label']}'이(가) 추천 투찰률을 {direction} 주요 요인입니다.")

        base_rate_val = fb.get("agency_avg_rate_12m") or fb.get("similar_avg_rate") or 0.879
        data_count = int(fb.get("agency_bid_count_12m") or 0) + int(fb.get("similar_bid_count") or 0)

        return {
            "top_factors":   top_factors,
            "narrative_ko":  " ".join(parts) if parts else "유사 입찰 이력 기반 앙상블 추천.",
            "model_version": b_result.get("model_version", "rule-based"),
            "data_count":    data_count,
            "base_rate":     round(float(base_rate_val), 4),
        }

    def _save_log_v2(self, db, req, center, lower, upper,
                     ep_result, fc, risk, explanation, user_id, w_a, w_b):
        try:
            from .models import PredictionLogV2
            log = PredictionLogV2(
                user_id=user_id,
                engine_weights={"engine_a": w_a, "engine_b": w_b},
                input_features={
                    "agency_id":   req.agency_id,
                    "industry_id": req.industry_id,
                    "region_id":   req.region_id,
                    "base_amount": req.base_amount,
                },
                srate_pred_center = ep_result["srate_range"]["center"],
                ep_confidence     = ep_result["confidence"],
                rate_aggressive   = lower,
                rate_balanced     = center,
                rate_conservative = upper,
                rate_center       = center,
                risk_level        = risk["level"],
                risk_score        = risk["score"],
                competition_score = fc.get("competitor_strength_score"),
                hhi_score         = fc.get("hhi_score"),
                explanation_text  = explanation.get("narrative_ko",""),
            )
            db.add(log)
            db.commit()
        except Exception as e:
            logger.debug(f"v2 로그 저장 실패: {e}")
            db.rollback()


# ==================================================
# 기관 분석 서비스
# ==================================================

class AgencyAnalysisService:
    def __init__(self, db: Session):
        self.db = db

    def list_agencies(self, q: str = None, page: int = 1, size: int = 20) -> dict:
        from sqlalchemy import func as sqlfunc
        query = self.db.query(
            Agency,
            sqlfunc.count(Bid.id).label("bid_count")
        ).outerjoin(Bid, Bid.agency_id == Agency.id).group_by(Agency.id)
        if q:
            query = query.filter(Agency.name.ilike(f"%{q}%"))
        total = query.count()
        rows = query.order_by(sqlfunc.count(Bid.id).desc()).offset((page - 1) * size).limit(size).all()
        items = []
        for agency, bid_count in rows:
            region_name = None
            if agency.region_id:
                r = self.db.query(Region).filter(Region.id == agency.region_id).first()
                if r:
                    region_name = r.name
            items.append({
                "id": agency.id,
                "name": agency.name,
                "type": agency.type,
                "region_name": region_name,
                "bid_count": bid_count or 0
            })
        return {"items": items, "total": total}

    def analyze(self, agency_id: int) -> dict:
        from fastapi import HTTPException
        agency = self.db.query(Agency).filter(Agency.id == agency_id).first()
        if not agency:
            raise HTTPException(status_code=404, detail="기관을 찾을 수 없습니다")

        cutoff = datetime.utcnow() - timedelta(days=730)  # 24개월
        bids = self.db.query(Bid).filter(
            Bid.agency_id == agency_id,
            Bid.bid_open_date >= cutoff
        ).all()

        total_bids = len(bids)
        bid_ids = [b.id for b in bids]

        results = self.db.query(BidResult).filter(
            BidResult.bid_id.in_(bid_ids),
            BidResult.is_winner == True
        ).all() if bid_ids else []
        avg_win_rate = float(sum(float(r.bid_rate) for r in results) / len(results)) if results else None

        srates = [float(r.assessment_rate) for r in results if r.assessment_rate is not None]
        avg_srate = float(sum(srates) / len(srates)) if srates else None

        from collections import Counter
        industry_ids = [b.industry_id for b in bids if b.industry_id]
        dominant_industry = None
        if industry_ids:
            most_common_id = Counter(industry_ids).most_common(1)[0][0]
            ind = self.db.query(Industry).filter(Industry.id == most_common_id).first()
            dominant_industry = ind.name if ind else None

        summary = {
            "name": agency.name,
            "total_bids": total_bids,
            "avg_win_rate": avg_win_rate,
            "avg_srate": avg_srate,
            "dominant_industry": dominant_industry
        }

        # monthly_trend (24개월)
        monthly_trend = []
        for i in range(24):
            target = datetime.utcnow() - timedelta(days=30 * i)
            y, m = target.year, target.month
            month_bids = [b for b in bids if b.bid_open_date and b.bid_open_date.year == y and b.bid_open_date.month == m]
            month_bid_ids = [b.id for b in month_bids]
            month_winners = self.db.query(BidResult).filter(
                BidResult.bid_id.in_(month_bid_ids), BidResult.is_winner == True
            ).all() if month_bid_ids else []
            win_rates = [float(r.bid_rate) for r in month_winners if r.bid_rate]
            srates_m = [float(r.assessment_rate) for r in month_winners if r.assessment_rate]
            monthly_trend.append({
                "year_month": f"{y}-{m:02d}",
                "bid_count": len(month_bids),
                "win_rate": sum(win_rates) / len(win_rates) if win_rates else None,
                "avg_srate": sum(srates_m) / len(srates_m) if srates_m else None
            })
        monthly_trend.reverse()

        # srate_distribution
        all_results = self.db.query(BidResult).filter(
            BidResult.bid_id.in_(bid_ids),
            BidResult.assessment_rate.isnot(None)
        ).all() if bid_ids else []
        srate_vals = [float(r.assessment_rate) for r in all_results]
        srate_dist = self._make_distribution(srate_vals, 80, 110, 0.5)

        # top_winners
        from sqlalchemy import func as sqlfunc
        if bid_ids:
            winner_query = self.db.query(
                Competitor.name,
                sqlfunc.count(BidResult.id).label("win_count"),
                sqlfunc.avg(BidResult.bid_rate).label("avg_rate")
            ).join(BidResult, BidResult.competitor_id == Competitor.id).filter(
                BidResult.bid_id.in_(bid_ids),
                BidResult.is_winner == True
            ).group_by(Competitor.name).order_by(sqlfunc.count(BidResult.id).desc()).limit(10).all()
            top_winners = [{"competitor_name": r.name, "win_count": r.win_count, "avg_bid_rate": float(r.avg_rate) if r.avg_rate else None} for r in winner_query]
        else:
            top_winners = []

        # amount_distribution
        buckets = [
            ("1억 미만", 0, 100_000_000),
            ("1~3억", 100_000_000, 300_000_000),
            ("3~10억", 300_000_000, 1_000_000_000),
            ("10억 이상", 1_000_000_000, float("inf")),
        ]
        amount_dist = []
        for label, low, high in buckets:
            bucket_bids = [b for b in bids if b.base_amount and low <= b.base_amount < high]
            bucket_ids = [b.id for b in bucket_bids]
            bucket_winners = self.db.query(BidResult).filter(
                BidResult.bid_id.in_(bucket_ids), BidResult.is_winner == True
            ).all() if bucket_ids else []
            wrs = [float(r.bid_rate) for r in bucket_winners if r.bid_rate]
            amount_dist.append({
                "bucket_label": label,
                "count": len(bucket_bids),
                "avg_win_rate": sum(wrs) / len(wrs) if wrs else None
            })

        return {
            "summary": summary,
            "monthly_trend": monthly_trend,
            "srate_distribution": srate_dist,
            "top_winners": top_winners,
            "amount_distribution": amount_dist
        }

    def srate_histogram(self, agency_id: int, months: int = 12) -> dict:
        from fastapi import HTTPException
        agency = self.db.query(Agency).filter(Agency.id == agency_id).first()
        if not agency:
            raise HTTPException(status_code=404, detail="기관을 찾을 수 없습니다")

        cutoff = datetime.utcnow() - timedelta(days=30 * months)
        bid_ids = [b.id for b in self.db.query(Bid).filter(
            Bid.agency_id == agency_id,
            Bid.bid_open_date >= cutoff,
        ).all()]

        results = self.db.query(BidResult).filter(
            BidResult.bid_id.in_(bid_ids),
            BidResult.is_winner == True,
            BidResult.assessment_rate.isnot(None),
        ).all() if bid_ids else []

        srate_vals = sorted([float(r.assessment_rate) for r in results])
        n = len(srate_vals)

        mean = float(np.mean(srate_vals)) if n > 0 else None
        std = float(np.std(srate_vals)) if n > 0 else None

        lo, hi, width = 0.860, 0.960, 0.005
        bins = []
        b = lo
        while b < hi - 1e-9:
            cnt = sum(1 for v in srate_vals if b <= v < b + width)
            pct = round(cnt / n * 100, 1) if n > 0 else 0.0
            bins.append({
                "range_lo": round(b, 3),
                "range_hi": round(b + width, 3),
                "count": cnt,
                "pct": pct,
            })
            b = round(b + width, 3)

        def _pct(p):
            return float(np.percentile(srate_vals, p)) if srate_vals else None

        percentiles = {
            "p10": _pct(10), "p25": _pct(25), "p50": _pct(50),
            "p75": _pct(75), "p90": _pct(90),
        }

        return {
            "agency_id": agency_id,
            "agency_name": agency.name,
            "months": months,
            "sample_count": n,
            "mean": mean,
            "std": std,
            "bins": bins,
            "percentiles": percentiles,
        }

    def recent_results(self, agency_id: int, limit: int = 20) -> dict:
        from fastapi import HTTPException
        from sqlalchemy import func as sqlfunc
        agency = self.db.query(Agency).filter(Agency.id == agency_id).first()
        if not agency:
            raise HTTPException(status_code=404, detail="기관을 찾을 수 없습니다")

        bids = self.db.query(Bid).filter(
            Bid.agency_id == agency_id,
            Bid.bid_open_date.isnot(None),
        ).order_by(Bid.bid_open_date.desc()).limit(limit * 3).all()

        items = []
        for bid in bids:
            if len(items) >= limit:
                break
            winner = self.db.query(BidResult).filter(
                BidResult.bid_id == bid.id,
                BidResult.is_winner == True,
            ).first()
            comp_count = self.db.query(sqlfunc.count(BidResult.id)).filter(
                BidResult.bid_id == bid.id,
            ).scalar() or 0
            items.append({
                "bid_id": bid.id,
                "title": bid.title,
                "base_amount": float(bid.base_amount) if bid.base_amount else 0.0,
                "bid_open_date": bid.bid_open_date.isoformat() if bid.bid_open_date else None,
                "assessment_rate": float(winner.assessment_rate) if winner and winner.assessment_rate else None,
                "competitor_count": comp_count,
            })

        return {"items": items, "total": len(items)}

    def _make_distribution(self, values: list, low: float, high: float, step: float) -> dict:
        if not values:
            return {"bins": [], "mode": None, "p25": None, "p50": None, "p75": None, "mean": None, "std": None}
        arr = np.array(values)
        bins_range = np.arange(low, high + step, step)
        counts, edges = np.histogram(arr, bins=bins_range)
        bins = [{"rate_pct": float(edges[i]), "count": int(counts[i])} for i in range(len(counts))]
        mode_idx = int(np.argmax(counts))
        mode_val = float(edges[mode_idx]) if counts[mode_idx] > 0 else None
        return {
            "bins": bins,
            "mode": mode_val,
            "p25": float(np.percentile(arr, 25)),
            "p50": float(np.percentile(arr, 50)),
            "p75": float(np.percentile(arr, 75)),
            "mean": float(np.mean(arr)),
            "std": float(np.std(arr)),
            "sample_count": len(values)
        }


# ============================================================
# 수주율 최적화 서비스 레이어
# ============================================================

class CompanyProfileService:
    """회사 프로파일 CRUD + 역량 조회"""

    def get_profile(self, db: Session, user_id: int):
        from .models import CompanyProfile
        return db.query(CompanyProfile).first()

    def upsert_profile(self, db: Session, data: dict) -> "CompanyProfile":
        from .models import CompanyProfile
        profile = db.query(CompanyProfile).first()
        if profile is None:
            profile = CompanyProfile(**data)
            db.add(profile)
        else:
            for k, v in data.items():
                setattr(profile, k, v)
        db.commit()
        db.refresh(profile)
        return profile

    def get_remaining_bond(self, db: Session) -> dict:
        from .models import CompanyProfile, PortfolioState
        profile = db.query(CompanyProfile).first()
        if not profile:
            return {"total": 0, "used": 0, "remaining": 0, "usage_rate": 0.0}

        # 현재 ACTIVE 상태 포트폴리오 보증 소요액 합산
        active_bond = db.query(func.sum(PortfolioState.bond_exposure)).filter(
            PortfolioState.status == "ACTIVE"
        ).scalar() or 0

        total   = profile.bond_limit_total or 0
        used    = int(active_bond)
        remaining = max(0, total - used)
        rate    = used / total if total > 0 else 0.0
        return {"total": total, "used": used, "remaining": remaining, "usage_rate": round(rate, 4)}


class QualificationService:
    """E2: 적격심사 엔진 서비스"""

    def check(
        self,
        db: Session,
        bid_id: int,
        user_id: int,
        our_share_rate: float = 1.0,
        our_experience: int = 0,
        reputation_score: float = 0.0,
        contract_law: str = "local",
    ) -> dict:
        from .models import Bid, QualificationCheck, CompanyProfile
        from .ml.qualification import check_qualification, QualificationResult
        from .ml.assessment import predict_srate, load_srate_stats

        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            from fastapi import HTTPException
            raise HTTPException(404, "공고를 찾을 수 없습니다")

        profile = db.query(CompanyProfile).first()

        # 사정율 예측 (중앙값 + 표준편차)
        features_a   = load_srate_stats(db, bid.agency_id, bid.industry_id, bid.region_id, bid.base_amount)
        srate_result = predict_srate(features_a, bid.base_amount)
        _rng         = srate_result["srate_range"]
        srate_center = _rng["center"]
        srate_std    = (_rng["upper"] - _rng["lower"]) / 2

        result: QualificationResult = check_qualification(
            base_amount=bid.base_amount,
            estimated_price_center=srate_center,
            estimated_price_std=srate_std,
            our_experience=our_experience or (profile.performance_records.get("total", 0) if profile else 0),
            annual_revenue=profile.annual_revenue if profile else 0,
            workforce_count=profile.workforce_count if profile else 0,
            share_rate=our_share_rate,
            reputation_score=reputation_score,
            contract_law=contract_law,
        )

        # DB 저장
        check = QualificationCheck(
            bid_id=bid_id,
            user_id=user_id,
            our_share_rate=our_share_rate,
            our_experience=our_experience,
            pass_prob=result.pass_prob,
            min_pass_amount=result.min_pass_amount,
            max_pass_amount=result.max_pass_amount,
            score_breakdown=result.score_breakdown,
            verdict=result.verdict,
            fail_reason=result.fail_reason,
        )
        db.add(check)
        db.commit()

        return {
            "bid_id":          bid_id,
            "verdict":         result.verdict,
            "pass_prob":       result.pass_prob,
            "min_pass_amount": result.min_pass_amount,
            "max_pass_amount": result.max_pass_amount,
            "score_breakdown": result.score_breakdown,
            "fail_reason":     result.fail_reason,
            "criteria_type":   result.criteria_type,
        }


class BidSelectionService:
    """E1: 공고 선별 엔진 서비스"""

    def evaluate_bid(self, db: Session, bid_id: int, user_id: int) -> dict:
        from .models import Bid, CompanyProfile, BidDecision, PortfolioState
        from .ml.selection import SelectionInput, evaluate
        from .ml.assessment import predict_srate, load_srate_stats
        from .ml.competition import compute_competition_features

        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            from fastapi import HTTPException
            raise HTTPException(404, "공고를 찾을 수 없습니다")

        profile   = db.query(CompanyProfile).first()
        qual_svc  = QualificationService()

        # 적격심사 사전 체크 (profile 없으면 기본값)
        try:
            qual = qual_svc.check(db, bid_id, user_id)
            qualify_prob = qual["pass_prob"]
        except Exception:
            qualify_prob = 0.8

        # 면허 / 지역 매칭
        license_match = True
        region_ok     = True
        if profile and profile.license_codes and bid.license_codes:
            license_match = bool(set(profile.license_codes) & set(bid.license_codes or []))
        if bid.region_restriction and profile and profile.region_codes:
            region_ok = bool(set(profile.region_codes) & set([str(bid.region_id or "")]))

        # 경쟁 강도
        comp = compute_competition_features(
            db=db,
            agency_id=bid.agency_id,
            industry_id=bid.industry_id,
            base_amount=bid.base_amount,
        )
        comp_score   = comp.get("competitor_strength_score", 5.0)
        strong_count = sum(1 for s in comp.get("competitor_scores", []) if s >= 7.0)

        # 사정율 예측
        _fa        = load_srate_stats(db, bid.agency_id, bid.industry_id, bid.region_id, bid.base_amount)
        _sr        = predict_srate(_fa, bid.base_amount)
        _rng2      = _sr["srate_range"]
        srate_info = {"center": _rng2["center"], "std": (_rng2["upper"] - _rng2["lower"]) / 2}
        floor_rate = float(bid.min_bid_rate or 0.87745)
        best_wp    = max(0.0, min(1.0, 1.0 - comp_score / 12.0))  # 근사값

        # 전략 부합도
        in_target_region   = bool(profile and bid.region_id and bid.region_id in (profile.target_industries or []))
        in_target_industry = bool(profile and bid.industry_id and bid.industry_id in (profile.target_industries or []))

        # 과거 승률
        hist = db.execute(
            text("""
                SELECT COUNT(*) FILTER (WHERE abo.result='WON') as wins,
                       COUNT(*) as total
                FROM actual_bid_outcomes abo
                JOIN bids b ON b.id = abo.bid_id
                WHERE b.agency_id = :aid AND abo.user_id = :uid
            """),
            {"aid": bid.agency_id, "uid": user_id},
        ).fetchone()
        hist_win_rate = (hist.wins / hist.total) if hist and hist.total > 0 else 0.20

        # 보증한도 현황
        bond_svc = CompanyProfileService()
        bond = bond_svc.get_remaining_bond(db)

        # 현재 활성 투찰 건수
        active_count = db.query(func.count(PortfolioState.id)).filter(
            PortfolioState.status == "ACTIVE",
        ).scalar() or 0

        inp = SelectionInput(
            bid_id=bid_id,
            base_amount=bid.base_amount,
            agency_id=bid.agency_id,
            industry_id=bid.industry_id,
            region_id=bid.region_id,
            license_match=license_match,
            region_restriction_ok=region_ok,
            qualify_prob=qualify_prob,
            expected_competitor_count=comp.get("expected_competitor_count", 5),
            competitor_strength_score=comp_score,
            strong_competitor_count=strong_count,
            best_win_prob=best_wp,
            estimated_margin=profile.target_min_margin if profile else 0.05,
            in_target_region=in_target_region,
            in_target_industry=in_target_industry,
            bond_limit_total=bond["total"],
            bond_limit_used=bond["used"],
            max_concurrent_bids=profile.max_concurrent_bids if profile else 5,
            current_active_bids=active_count,
            historical_win_rate=hist_win_rate,
        )

        from .ml.selection import evaluate
        result = evaluate(inp)

        # 결과 저장
        decision = BidDecision(
            bid_id=bid_id,
            user_id=user_id,
            selection_score=result.score,
            ev_score=result.ev_score,
            qualify_prob=result.qualify_prob,
            win_prob_best=result.win_prob_best,
            expected_margin=result.expected_margin,
            competitor_risk=result.competitor_risk,
            verdict=result.verdict,
            no_go_reasons=result.no_go_reasons,
            recommended_strategy=result.recommended_strategy,
        )
        db.add(decision)
        db.commit()
        db.refresh(decision)

        return {
            "bid_id":               bid_id,
            "verdict":              result.verdict,
            "score":                result.score,
            "ev_score":             result.ev_score,
            "qualify_prob":         result.qualify_prob,
            "win_prob_best":        result.win_prob_best,
            "expected_margin":      result.expected_margin,
            "competitor_risk":      result.competitor_risk,
            "no_go_reasons":        result.no_go_reasons,
            "score_detail":         result.score_detail,
            "recommended_strategy": result.recommended_strategy,
            "decision_id":          decision.id,
        }

    def get_go_list(self, db: Session, user_id: int, days: int = 7) -> dict:
        """최근 n일 GO 목록 반환"""
        from .models import BidDecision, Bid
        cutoff = datetime.utcnow() - timedelta(days=days)
        decisions = db.query(BidDecision).filter(
            BidDecision.user_id == user_id,
            BidDecision.created_at >= cutoff,
        ).order_by(BidDecision.selection_score.desc()).all()

        go    = [d for d in decisions if d.verdict == "GO"]
        watch = [d for d in decisions if d.verdict == "WATCH"]
        no_go = [d for d in decisions if d.verdict == "NO_GO"]

        def _fmt(d: BidDecision) -> dict:
            bid = db.query(Bid).filter(Bid.id == d.bid_id).first()
            # 근거 데이터 건수 (신뢰도 지표)
            agency_id = bid.agency_id if bid else None
            data_count = 0
            if agency_id:
                row = db.execute(text(
                    "SELECT COUNT(*) FROM bid_results r JOIN bids b ON b.id=r.bid_id WHERE b.agency_id=:aid"
                ), {"aid": agency_id}).scalar()
                data_count = int(row or 0)
            confidence = "high" if data_count >= 100 else ("medium" if data_count >= 20 else "low")
            return {
                "bid_id":               d.bid_id,
                "title":                bid.title if bid else "",
                "base_amount":          bid.base_amount if bid else 0,
                "bid_open_date":        bid.bid_open_date.isoformat() if bid and bid.bid_open_date else None,
                "verdict":              d.verdict,
                "score":                float(d.selection_score or 0),
                "ev_score":             d.ev_score or 0,
                "qualify_prob":         float(d.qualify_prob or 0),
                "win_prob_best":        float(d.win_prob_best or 0),
                "competitor_risk":      d.competitor_risk,
                "no_go_reasons":        d.no_go_reasons or [],
                "recommended_strategy": d.recommended_strategy,
                "recommended_rate":     float(d.recommended_rate) if d.recommended_rate else None,
                "actual_action":        d.actual_action,
                "data_count":           data_count,
                "confidence":           confidence,
            }

        return {
            "go":         [_fmt(d) for d in go],
            "watch":      [_fmt(d) for d in watch],
            "no_go":      [_fmt(d) for d in no_go],
            "total":      len(decisions),
            "go_count":   len(go),
            "watch_count": len(watch),
            "no_go_count": len(no_go),
        }


class SingleRecommendService:
    """E5: 단일 최적 전략 추천 서비스"""

    def recommend(self, db: Session, user_id: int, req: dict) -> dict:
        from .models import Bid, CompanyProfile, BidDecision, ActualBidOutcome
        from .ml.assessment  import predict_srate, load_srate_stats
        from .ml.competition import compute_competition_features, get_market_competitor_distributions
        from .ml.simulation  import simulate_yejung
        from .ml.personal    import PersonalBiasAnalyzer
        from .ml.qualification import check_qualification, get_valid_bid_range
        from .ml.strategy    import StrategyInput, recommend as strategy_recommend
        from .ml.a_value     import calc_floor_rate

        bid_id   = req.get("bid_id")
        base_amt = req["base_amount"]

        bid = db.query(Bid).filter(Bid.id == bid_id).first() if bid_id else None
        profile = db.query(CompanyProfile).first()

        # 사정율 예측
        _fa2         = load_srate_stats(db, req["agency_id"], req.get("industry_id"), req.get("region_id"), base_amt)
        _sr2         = predict_srate(_fa2, base_amt)
        _rng3        = _sr2["srate_range"]
        srate_center = _rng3["center"]
        srate_std    = (_rng3["upper"] - _rng3["lower"]) / 2

        # Monte Carlo 사정율 분포 생성
        import numpy as np
        rng = np.random.default_rng(42)
        srate_dist = simulate_yejung(base_amt, srate_center, srate_std, n_sim=30_000, rng=rng)

        # 경쟁사 최소 투찰률 분포
        comp = compute_competition_features(
            db, req["agency_id"], req.get("industry_id"), base_amt
        )
        comp_means, comp_stds = get_market_competitor_distributions(
            db, req["agency_id"], req.get("industry_id")
        )

        # 경쟁사 min 분포 시뮬레이션
        if comp_means:
            n_sim = 30_000
            n_comp = len(comp_means)
            comp_matrix = np.column_stack([
                rng.normal(m, max(s, 0.002), n_sim)
                for m, s in zip(comp_means, comp_stds)
            ])
            comp_min_dist = comp_matrix.min(axis=1)
        else:
            comp_min_dist = None

        # 낙찰하한율 — industry_id → 공종명 조회 후 calc_floor_rate 호출
        _industry_id = req.get("industry_id")
        _industry_name = ""
        if _industry_id:
            _ind_row = db.execute(
                text("SELECT name FROM industries WHERE id = :id"),
                {"id": _industry_id},
            ).fetchone()
            if _ind_row:
                _industry_name = _ind_row[0]
        floor_rate = calc_floor_rate(_industry_name)

        # 적격심사 유효 범위
        valid_low = valid_high = None
        qual_result_dict = None
        if bid_id and profile:
            try:
                qual_svc = QualificationService()
                qual_result_dict = qual_svc.check(db, bid_id, user_id)
                from .ml.qualification import QualificationResult, get_valid_bid_range
                from .ml.qualification import QualificationResult as QR
                qr = QR(
                    verdict=qual_result_dict["verdict"],
                    pass_prob=qual_result_dict["pass_prob"],
                    min_pass_amount=qual_result_dict["min_pass_amount"],
                    max_pass_amount=qual_result_dict["max_pass_amount"],
                    score_breakdown=qual_result_dict["score_breakdown"],
                    fail_reason=qual_result_dict["fail_reason"],
                    criteria_type=qual_result_dict["criteria_type"],
                )
                lo_amt, hi_amt = get_valid_bid_range(qr, floor_rate, base_amt)
                if lo_amt and hi_amt:
                    valid_low  = lo_amt / base_amt
                    valid_high = hi_amt / base_amt
            except Exception:
                pass

        # 개인 편향 보정
        bias_correction = 0.0
        try:
            analyzer = PersonalBiasAnalyzer()
            bias_info = analyzer.analyze(db, user_id)
            bias_correction = bias_info.get("correction", 0.0)
        except Exception:
            pass

        # 월 수주 목표 현황
        monthly_target     = profile.monthly_win_target if profile else 3
        from .models import ActualBidOutcome
        current_month_wins = db.execute(
            text("SELECT COUNT(*) FROM actual_bid_outcomes WHERE result='WON' AND DATE_TRUNC('month', created_at) = DATE_TRUNC('month', NOW())")
        ).scalar() or 0

        inp = StrategyInput(
            base_amount=base_amt,
            floor_rate=floor_rate,
            srate_center=srate_center,
            srate_std=srate_std,
            srate_dist=srate_dist,
            competitor_means=comp_means,
            competitor_stds=comp_stds,
            competitor_min_dist=comp_min_dist,
            valid_low=valid_low,
            valid_high=valid_high,
            bias_correction=bias_correction,
            monthly_target=monthly_target,
            current_month_wins=current_month_wins,
            historical_win_rate=0.20,
        )

        from .ml.strategy import recommend as do_recommend
        rec = do_recommend(inp)

        return {
            "rate":              rec.rate,
            "bid_amount":        rec.bid_amount,
            "win_prob":          rec.win_prob,
            "expected_value":    rec.expected_value,
            "confidence":        rec.confidence,
            "strategy_type":     rec.strategy_type,
            "rationale":         rec.rationale,
            "rationale_details": rec.rationale_details,
            "valid_range":       list(rec.valid_range),
            "prism_top5":        rec.prism_top5,
            "qualification":     qual_result_dict,
        }


class ActualOutcomeService:
    """E6: 실제 투찰 결과 수집 및 피드백 처리"""

    def record_outcome(self, db: Session, user_id: int, data: dict) -> dict:
        from .models import ActualBidOutcome, PredictionLogV2, BidDecision

        bid_id = data["bid_id"]

        # 직전 예측값 조회 (캘리브레이션용)
        pred_log = db.query(PredictionLogV2).filter(
            PredictionLogV2.bid_id == bid_id
        ).order_by(PredictionLogV2.created_at.desc()).first()

        predicted_wp    = float(pred_log.win_prob_center) if pred_log and pred_log.win_prob_center else None
        predicted_srate = float(pred_log.srate_pred_center) if pred_log and pred_log.srate_pred_center else None

        # 사정율 오차
        srate_error = None
        if predicted_srate and data.get("actual_srate"):
            srate_error = abs(predicted_srate - data["actual_srate"])

        outcome = ActualBidOutcome(
            bid_id=bid_id,
            user_id=user_id,
            bid_decision_id=data.get("bid_decision_id"),
            submitted_rate=data["submitted_rate"],
            result=data["result"],
            disqualify_reason=data.get("disqualify_reason"),
            actual_srate=data.get("actual_srate"),
            winner_rate=data.get("winner_rate"),
            winner_biz_no=data.get("winner_biz_no"),
            our_rank=data.get("our_rank"),
            total_bidders=data.get("total_bidders"),
            predicted_win_prob=predicted_wp,
            predicted_srate=predicted_srate,
            srate_error=srate_error,
            collected_at=datetime.utcnow(),
        )
        db.add(outcome)

        # 포트폴리오 상태 갱신
        from .models import PortfolioState
        ps = db.query(PortfolioState).filter(
            PortfolioState.bid_id == bid_id,
            PortfolioState.user_id == user_id,
        ).first()
        if ps:
            ps.status     = "WON" if data["result"] == "WON" else "LOST"
            ps.result_date = datetime.utcnow().date()

        # bid_decision actual_action 갱신
        if data.get("bid_decision_id"):
            dec = db.query(BidDecision).filter(
                BidDecision.id == data["bid_decision_id"]
            ).first()
            if dec:
                dec.actual_action = "BID"
                dec.actual_rate   = data["submitted_rate"]

        # my_bid_records도 같이 갱신 (기존 테이블 연동)
        rec = db.query(MyBidRecord).filter(
            MyBidRecord.bid_id == bid_id,
            MyBidRecord.user_id == user_id,
        ).first()
        if rec:
            rec.result            = data["result"].lower()
            rec.actual_winner_rate = data.get("winner_rate")

        db.commit()
        db.refresh(outcome)

        # 재학습 트리거 위임
        MyBidFeedbackService(db)._check_retrain_trigger()

        return {"id": outcome.id, "result": outcome.result}


class KpiService:
    """E8: KPI 집계 + 경영진 대시보드"""

    def get_dashboard(self, db: Session, user_id: int, period_type: str = "MONTHLY") -> dict:
        from .models import CompanyProfile, ActualBidOutcome, BidDecision, KpiSnapshot
        from .ml.feedback import build_kpi_snapshot, should_alert

        today = datetime.utcnow().date()

        # 캐시된 스냅샷 조회
        cached = db.query(KpiSnapshot).filter(
            KpiSnapshot.snapshot_date == today,
            KpiSnapshot.user_id == user_id,
            KpiSnapshot.period_type == period_type,
        ).first()

        # 스냅샷 재계산
        kpi = build_kpi_snapshot(db, user_id, today, period_type)
        if not kpi:
            kpi = {
                "total_bids": 0, "total_wins": 0, "win_rate": 0.0,
                "qualify_pass_rate": None, "avg_rank_at_loss": None,
                "srate_mae": None, "win_prob_calibration": None,
                "go_rate": None, "no_go_saved": 0,
            }

        profile    = db.query(CompanyProfile).first()
        monthly_target = profile.monthly_win_target if profile else 3

        alerts = should_alert(kpi)

        # 최근 6개월 트렌드
        trend = []
        for i in range(5, -1, -1):
            d = today.replace(day=1)
            if i > 0:
                m = d.month - i
                y = d.year
                while m <= 0:
                    m += 12
                    y -= 1
                d = d.replace(year=y, month=m)
            snap = db.query(KpiSnapshot).filter(
                KpiSnapshot.snapshot_date == d,
                KpiSnapshot.user_id == user_id,
                KpiSnapshot.period_type == "MONTHLY",
            ).first()
            trend.append({
                "month":     d.strftime("%Y-%m"),
                "win_rate":  float(snap.win_rate) if snap and snap.win_rate else 0.0,
                "total_bids": snap.total_bids if snap else 0,
                "total_wins": snap.total_wins if snap else 0,
            })

        wins = kpi.get("total_wins", 0)
        return {
            "period_type":          period_type,
            "snapshot_date":        today.isoformat(),
            "total_bids":           kpi.get("total_bids", 0),
            "total_wins":           wins,
            "win_rate":             kpi.get("win_rate", 0.0),
            "monthly_target":       monthly_target,
            "target_achievement":   round(wins / monthly_target, 4) if monthly_target else 0.0,
            "qualify_pass_rate":    kpi.get("qualify_pass_rate"),
            "avg_rank_at_loss":     kpi.get("avg_rank_at_loss"),
            "srate_mae":            kpi.get("srate_mae"),
            "win_prob_calibration": kpi.get("win_prob_calibration"),
            "go_rate":              kpi.get("go_rate"),
            "no_go_saved":          kpi.get("no_go_saved", 0),
            "alerts":               alerts,
            "monthly_trend":        trend,
        }

    def upsert_snapshot(self, db: Session, kpi: dict):
        from .models import KpiSnapshot
        from datetime import date as date_type
        snap_date = kpi.get("snapshot_date", datetime.utcnow().date())
        if isinstance(snap_date, str):
            snap_date = date_type.fromisoformat(snap_date)

        existing = db.query(KpiSnapshot).filter(
            KpiSnapshot.snapshot_date == snap_date,
            KpiSnapshot.user_id == kpi.get("user_id"),
            KpiSnapshot.period_type == kpi.get("period_type", "MONTHLY"),
        ).first()

        fields = {
            "total_bids":           kpi.get("total_bids", 0),
            "total_wins":           kpi.get("total_wins", 0),
            "win_rate":             kpi.get("win_rate"),
            "qualify_pass_rate":    kpi.get("qualify_pass_rate"),
            "avg_rank_at_loss":     kpi.get("avg_rank_at_loss"),
            "srate_mae":            kpi.get("srate_mae"),
            "win_prob_calibration": kpi.get("win_prob_calibration"),
            "go_rate":              kpi.get("go_rate"),
            "no_go_saved":          kpi.get("no_go_saved", 0),
        }

        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            snap = KpiSnapshot(
                snapshot_date=snap_date,
                user_id=kpi.get("user_id"),
                period_type=kpi.get("period_type", "MONTHLY"),
                **fields,
            )
            db.add(snap)
        db.commit()


class PortfolioService:
    """E7: 포트폴리오 최적화 서비스"""

    def optimize(self, db: Session, user_id: int, bid_ids: list) -> dict:
        from .models import Bid, BidDecision, CompanyProfile, PortfolioState
        from .ml.portfolio import PortfolioBidItem, PortfolioConstraints, optimize, compute_portfolio_stats

        profile    = db.query(CompanyProfile).first()
        bond_svc   = CompanyProfileService()
        bond       = bond_svc.get_remaining_bond(db)
        active_cnt = db.query(func.count(PortfolioState.id)).filter(
            PortfolioState.status == "ACTIVE",
        ).scalar() or 0

        constraints = PortfolioConstraints(
            remaining_bond=bond["remaining"],
            max_concurrent_bids=profile.max_concurrent_bids if profile else 5,
            active_bid_count=active_cnt,
            weekly_prep_hours=40.0,
            monthly_target=profile.monthly_win_target if profile else 3,
            current_month_wins=0,
        )

        items: list[PortfolioBidItem] = []
        sel_svc = BidSelectionService()

        for bid_id in bid_ids:
            bid = db.query(Bid).filter(Bid.id == bid_id).first()
            if not bid:
                continue

            # 최신 선별 결과 조회 (없으면 즉시 평가)
            dec = db.query(BidDecision).filter(
                BidDecision.bid_id == bid_id,
                BidDecision.user_id == user_id,
            ).order_by(BidDecision.created_at.desc()).first()

            if not dec:
                try:
                    sel_svc.evaluate_bid(db, bid_id, user_id)
                    dec = db.query(BidDecision).filter(
                        BidDecision.bid_id == bid_id,
                        BidDecision.user_id == user_id,
                    ).order_by(BidDecision.created_at.desc()).first()
                except Exception:
                    pass

            verdict       = dec.verdict if dec else "WATCH"
            sel_score     = float(dec.selection_score or 5.0) if dec else 5.0
            ev_score      = dec.ev_score or 0 if dec else 0
            qualify_prob  = float(dec.qualify_prob or 0.8) if dec else 0.8
            win_prob      = float(dec.win_prob_best or 0.3) if dec else 0.3
            rec_rate      = float(dec.recommended_rate or 0.0) if dec else 0.0
            bond_exposure = int(bid.base_amount * 0.1)  # 기초금액의 10% 보증 근사

            items.append(PortfolioBidItem(
                bid_id=bid_id,
                title=bid.title,
                base_amount=bid.base_amount,
                bid_date=bid.bid_open_date.date().isoformat() if bid.bid_open_date else "unknown",
                verdict=verdict,
                selection_score=sel_score,
                ev_score=ev_score,
                qualify_prob=qualify_prob,
                win_prob=win_prob,
                bond_exposure=bond_exposure,
                recommended_rate=rec_rate,
            ))

        plan = optimize(items, constraints)
        stats = compute_portfolio_stats(plan)

        def _fmt(item: PortfolioBidItem) -> dict:
            return {
                "bid_id":           item.bid_id,
                "title":            item.title,
                "base_amount":      item.base_amount,
                "bid_date":         item.bid_date,
                "verdict":          item.verdict,
                "selection_score":  item.selection_score,
                "ev_score":         item.ev_score,
                "qualify_prob":     item.qualify_prob,
                "win_prob":         item.win_prob,
                "recommended_rate": item.recommended_rate,
            }

        return {
            "selected":             [_fmt(i) for i in plan.selected],
            "not_selected":         [_fmt(i) for i in plan.not_selected],
            "no_go_list":           [_fmt(i) for i in plan.no_go_list],
            "expected_wins":        plan.expected_wins,
            "expected_win_amount":  plan.expected_win_amount,
            "total_ev":             plan.total_ev,
            "bond_usage":           plan.bond_usage,
            "remaining_bond_after": plan.remaining_bond_after,
            "alerts":               plan.alerts,
            "schedule":             plan.schedule,
            "stats":                stats,
        }


# ==================================================
# 경쟁사 투찰성향 분석 서비스
# ==================================================

class CompetitorPatternService:
    def __init__(self, db: Session):
        self.db = db

    def get_pattern(self, competitor_id: int) -> dict:
        from fastapi import HTTPException
        from collections import Counter
        competitor = self.db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not competitor:
            raise HTTPException(status_code=404, detail="경쟁사를 찾을 수 없습니다")

        cutoff = datetime.utcnow() - timedelta(days=365)
        results = self.db.query(BidResult).join(Bid, Bid.id == BidResult.bid_id).filter(
            BidResult.competitor_id == competitor_id,
            Bid.bid_open_date >= cutoff
        ).all()

        if not results:
            radar = {"aggression": 0.0, "consistency": 0.0, "concentration": 0.0, "risk": 0.0, "activity": 0.0}
            return {"radar": radar, "amount_pattern": [], "recent_trend": {"direction": "stable", "change_pct": None}}

        rates = [float(r.bid_rate) for r in results if r.bid_rate]
        aggression = len([r for r in rates if r < 87.0]) / len(rates) * 10 if rates else 0
        std_val = float(np.std(rates)) if len(rates) > 1 else 1.0
        consistency = min(10.0, 2.0 / (std_val + 0.01))

        bid_ids = [r.bid_id for r in results]
        bids_data = self.db.query(Bid).filter(Bid.id.in_(bid_ids)).all()
        agency_ids = [b.agency_id for b in bids_data]
        if agency_ids:
            top_count = Counter(agency_ids).most_common(1)[0][1]
            concentration = min(10.0, top_count / len(agency_ids) * 10)
        else:
            concentration = 0.0

        risk = len([r for r in rates if r < 87.745]) / len(rates) * 10 if rates else 0

        monthly_counts = Counter([
            f"{b.bid_open_date.year}-{b.bid_open_date.month:02d}"
            for b in bids_data if b.bid_open_date
        ])
        activity = min(10.0, sum(monthly_counts.values()) / max(len(monthly_counts), 1) * 2)

        radar = {
            "aggression": round(aggression, 2),
            "consistency": round(consistency, 2),
            "concentration": round(concentration, 2),
            "risk": round(risk, 2),
            "activity": round(activity, 2)
        }

        buckets = [
            ("1억 미만", 0, 100_000_000),
            ("1~3억", 100_000_000, 300_000_000),
            ("3~10억", 300_000_000, 1_000_000_000),
            ("10억 이상", 1_000_000_000, float("inf")),
        ]
        bid_map = {b.id: b for b in bids_data}
        amount_pattern = []
        for label, low, high in buckets:
            bucket_results = [r for r in results if r.bid_id in bid_map and bid_map[r.bid_id].base_amount and low <= bid_map[r.bid_id].base_amount < high]
            b_rates = [float(r.bid_rate) for r in bucket_results if r.bid_rate]
            wins = [r for r in bucket_results if r.is_winner]
            amount_pattern.append({
                "bucket": label,
                "bid_count": len(bucket_results),
                "win_count": len(wins),
                "avg_rate": round(sum(b_rates) / len(b_rates), 4) if b_rates else None,
                "win_rate": round(len(wins) / len(bucket_results), 4) if bucket_results else None
            })

        from datetime import timezone as _tz
        six_months_ago = datetime.now(_tz.utc) - timedelta(days=180)
        def _naive(dt): return dt.replace(tzinfo=None) if dt.tzinfo else dt
        recent_bids = [b for b in bids_data if b.bid_open_date and _naive(b.bid_open_date) >= six_months_ago.replace(tzinfo=None)]
        older_bids = [b for b in bids_data if b.bid_open_date and _naive(b.bid_open_date) < six_months_ago.replace(tzinfo=None)]
        recent_ids = {b.id for b in recent_bids}
        older_ids = {b.id for b in older_bids}
        recent_rates = [float(r.bid_rate) for r in results if r.bid_id in recent_ids and r.bid_rate]
        older_rates = [float(r.bid_rate) for r in results if r.bid_id in older_ids and r.bid_rate]
        direction = "stable"
        change_pct = None
        if recent_rates and older_rates:
            r_mean = sum(recent_rates) / len(recent_rates)
            o_mean = sum(older_rates) / len(older_rates)
            change_pct = round((r_mean - o_mean) / o_mean * 100, 2)
            if change_pct < -0.5:
                direction = "aggressive"
            elif change_pct > 0.5:
                direction = "defensive"

        return {
            "radar": radar,
            "amount_pattern": amount_pattern,
            "recent_trend": {"direction": direction, "change_pct": change_pct}
        }

    def compare(self, ids: list) -> dict:
        from collections import defaultdict
        competitors_data = []
        for cid in ids[:2]:
            competitor = self.db.query(Competitor).filter(Competitor.id == cid).first()
            if not competitor:
                continue
            pattern = self.get_pattern(cid)

            cutoff = datetime.utcnow() - timedelta(days=365)
            results = self.db.query(BidResult).join(Bid).filter(
                BidResult.competitor_id == cid, Bid.bid_open_date >= cutoff
            ).all()
            bid_ids = [r.bid_id for r in results]
            bids_data = self.db.query(Bid).filter(Bid.id.in_(bid_ids)).all()
            monthly = defaultdict(lambda: {"bid_count": 0, "win_count": 0, "rates": []})
            for r in results:
                b = next((b for b in bids_data if b.id == r.bid_id), None)
                if b and b.bid_open_date:
                    key = f"{b.bid_open_date.year}-{b.bid_open_date.month:02d}"
                    monthly[key]["bid_count"] += 1
                    if r.is_winner:
                        monthly[key]["win_count"] += 1
                    if r.bid_rate:
                        monthly[key]["rates"].append(float(r.bid_rate))
            trend = [
                {
                    "year_month": k,
                    "bid_count": v["bid_count"],
                    "win_count": v["win_count"],
                    "avg_rate": round(sum(v["rates"]) / len(v["rates"]), 4) if v["rates"] else None
                }
                for k, v in sorted(monthly.items())
            ]
            competitors_data.append({
                "id": competitor.id,
                "name": competitor.name,
                "radar": pattern["radar"],
                "monthly_trend": trend
            })
        return {"competitors": competitors_data}


# ==================================================
# 북마크 서비스
# ==================================================

class BookmarkService:
    def __init__(self, db: Session):
        self.db = db

    def add(self, bid_id: int, user_id: int, note: str = None) -> BidBookmark:
        existing = self.db.query(BidBookmark).filter(
            BidBookmark.user_id == user_id, BidBookmark.bid_id == bid_id
        ).first()
        if existing:
            return existing
        bookmark = BidBookmark(bid_id=bid_id, user_id=user_id, note=note)
        self.db.add(bookmark)
        self.db.commit()
        self.db.refresh(bookmark)
        return bookmark

    def remove(self, bid_id: int, user_id: int):
        bookmark = self.db.query(BidBookmark).filter(
            BidBookmark.user_id == user_id, BidBookmark.bid_id == bid_id
        ).first()
        if bookmark:
            self.db.delete(bookmark)
            self.db.commit()

    def list_bookmarks(self, user_id: int, page: int = 1, size: int = 20) -> dict:
        query = self.db.query(BidBookmark).filter(BidBookmark.user_id == user_id)
        total = query.count()
        items = query.order_by(BidBookmark.created_at.desc()).offset((page - 1) * size).limit(size).all()
        return {"items": items, "total": total}

    def get_bookmarked_ids(self, user_id: int, bid_ids: list) -> set:
        rows = self.db.query(BidBookmark.bid_id).filter(
            BidBookmark.user_id == user_id,
            BidBookmark.bid_id.in_(bid_ids)
        ).all()
        return {r.bid_id for r in rows}


# ==================================================
# 투찰이력 엑셀 업로드 서비스
# ==================================================

class MyBidFeedbackService:
    """MyBidRecord → ActualBidOutcome 동기화 + 임계치 도달 시 자동 재학습."""

    RETRAIN_LOCK = False  # 동시 재학습 방지 (프로세스 내 단순 플래그)

    def __init__(self, db: Session):
        self.db = db

    def sync_outcome(self, rec: "MyBidRecord") -> bool:
        """낙찰/패찰 결과를 ActualBidOutcome에 반영. pending이면 skip."""
        if rec.result not in ("won", "lost"):
            return False

        result_val = "WON" if rec.result == "won" else "LOST"

        # 기존 레코드 조회 — bid_id 우선, 없으면 announcement_no
        existing = None
        if rec.bid_id:
            existing = (
                self.db.query(ActualBidOutcome)
                .filter(ActualBidOutcome.user_id == rec.user_id,
                        ActualBidOutcome.bid_id == rec.bid_id)
                .first()
            )
        if existing is None and rec.announcement_no:
            existing = (
                self.db.query(ActualBidOutcome)
                .filter(ActualBidOutcome.user_id == rec.user_id,
                        ActualBidOutcome.announcement_no == rec.announcement_no)
                .first()
            )

        if existing:
            existing.result       = result_val
            existing.submitted_rate = rec.submitted_rate
            existing.winner_rate  = rec.actual_winner_rate
            existing.collected_at = datetime.utcnow()
        else:
            if rec.bid_id is None and not rec.announcement_no:
                return False  # 연결 키 없음 — 동기화 불가
            self.db.add(ActualBidOutcome(
                bid_id          = rec.bid_id,
                user_id         = rec.user_id,
                announcement_no = rec.announcement_no,
                submitted_rate  = rec.submitted_rate,
                result          = result_val,
                winner_rate     = rec.actual_winner_rate,
                collected_at    = datetime.utcnow(),
            ))

        self.db.flush()
        self._check_retrain_trigger()
        return True

    def _check_retrain_trigger(self):
        from .ml.feedback import RETRAIN_THRESHOLD
        last_log = (
            self.db.query(ModelPerformanceLog)
            .order_by(ModelPerformanceLog.created_at.desc())
            .first()
        )
        last_date = last_log.eval_date if last_log else None
        cutoff = datetime.combine(last_date, datetime.min.time()) if last_date else datetime.min
        new_count = (
            self.db.query(func.count(ActualBidOutcome.id))
            .filter(
                ActualBidOutcome.created_at >= cutoff,
                ActualBidOutcome.result.in_(["WON", "LOST"]),
            )
            .scalar() or 0
        )
        if new_count >= RETRAIN_THRESHOLD and not MyBidFeedbackService.RETRAIN_LOCK:
            logger.info("자동 재학습 트리거: 신규 결과 %d건 누적", new_count)
            import threading
            threading.Thread(target=self._run_retrain, daemon=True).start()

    @classmethod
    def _run_retrain(cls):
        """백그라운드 재학습 — Engine A(사정율) + Engine B(낙찰률)."""
        if cls.RETRAIN_LOCK:
            return
        cls.RETRAIN_LOCK = True
        from .database import SessionLocal
        from .ml.engine import train_models, train_models_temporal, build_features, FEATURE_COLS, get_engine
        from .ml.assessment import compute_and_store_stats, train_srate_model
        import pandas as pd
        from sqlalchemy import text as sa_text

        db = SessionLocal()
        try:
            # Engine A
            compute_and_store_stats(db)
            train_srate_model(db)

            # Engine B — 낙찰률 회귀 모델
            rows = db.execute(sa_text("""
                SELECT b.id, b.agency_id, b.industry_id,
                       COALESCE(b.region_id, 0) AS region_id,
                       b.base_amount, b.bid_open_date,
                       COALESCE(b.region_restriction, false),
                       b.construction_period, r.bid_rate
                FROM bids b
                JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
                WHERE b.base_amount > 0
                  AND r.bid_rate BETWEEN 0.80 AND 1.00
                ORDER BY b.bid_open_date
            """)).fetchall()

            from datetime import timedelta
            cutoff = datetime.now() - timedelta(days=24 * 30)
            hist_rows = db.execute(sa_text("""
                SELECT b.id, b.agency_id, b.industry_id,
                       COALESCE(b.region_id, 0), b.base_amount, b.bid_open_date,
                       r.bid_rate,
                       (SELECT COUNT(*) FROM bid_results r2 WHERE r2.bid_id = b.id)
                FROM bids b
                LEFT JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
                WHERE b.bid_open_date >= :cutoff AND b.base_amount > 0
            """), {"cutoff": cutoff}).fetchall()

            hist_df = pd.DataFrame(hist_rows, columns=[
                "id", "agency_id", "industry_id", "region_id",
                "base_amount", "bid_open_date", "winner_rate", "competitor_count",
            ])
            for col in ["winner_rate", "base_amount", "competitor_count"]:
                hist_df[col] = pd.to_numeric(hist_df[col], errors="coerce")

            from .ml.yega import load_inpo21c_yega_stats as _load_yega
            _yega_cache: dict = {}

            records = []
            for row in rows:
                bid_id, agency_id, industry_id, region_id, base_amount, bid_open_date, region_restriction, construction_period, winner_rate = row
                if winner_rate is None:
                    continue
                hist_before = hist_df[hist_df["bid_open_date"] < bid_open_date].copy() if bid_open_date else hist_df.copy()
                _aid = int(agency_id) if agency_id else 0
                if _aid not in _yega_cache:
                    try:
                        _yega_cache[_aid] = _load_yega(db, _aid)
                    except Exception:
                        _yega_cache[_aid] = {}
                _yega_ml = _compute_yega_ml_features(_yega_cache[_aid].get("pos_weights"))
                try:
                    feats = build_features(
                        agency_id=_aid,
                        industry_id=int(industry_id) if industry_id else 0,
                        region_id=int(region_id),
                        base_amount=int(base_amount),
                        construction_period=int(construction_period) if construction_period else None,
                        region_restriction=bool(region_restriction),
                        bid_open_date=bid_open_date,
                        historical_df=hist_before,
                        yega_features=_yega_ml,
                    )
                    feats["target_rate"] = float(winner_rate)
                    feats["is_winner"]   = True
                    records.append(feats)
                except Exception:
                    pass

            if len(records) >= 20:
                train_df = pd.DataFrame(records)
                for col in FEATURE_COLS:
                    if col not in train_df.columns:
                        train_df[col] = None
                result = train_models_temporal(train_df, val_weeks=4, date_col="bid_open_date")
                if result:
                    get_engine().reload()
                    db.add(ModelPerformanceLog(
                        model_name    = "auto_retrain",
                        model_version = result.get("version", ""),
                        eval_date     = datetime.utcnow().date(),
                        sample_count  = result.get("train_size", 0),
                    ))
                    db.commit()
                    logger.info("자동 재학습 완료: %s", result)
            else:
                logger.warning("자동 재학습 스킵 — 피처 빌드 성공 %d건 (최소 20건 필요)", len(records))
        except Exception as exc:
            logger.error("자동 재학습 실패: %s", exc)
        finally:
            cls.RETRAIN_LOCK = False
            db.close()


class MyBidImportService:
    """투찰이력 엑셀 파일 → MyBidRecord 일괄 등록."""

    RESULT_MAP = {
        "낙찰": "won", "수주": "won", "won": "won",
        "유찰": "lost", "패찰": "lost", "미낙찰": "lost", "lost": "lost",
    }

    def __init__(self, db: Session):
        self.db = db

    # ── 숫자 파싱 헬퍼 ──────────────────────────────────────
    @staticmethod
    def _to_float(val) -> float | None:
        if val is None or str(val).strip() in ("", "-"):
            return None
        try:
            return float(str(val).replace("%", "").replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_int(val) -> int:
        if val is None:
            return 0
        try:
            return int(float(str(val).replace(",", "").strip()))
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def _to_date(val):
        """문자열 또는 datetime/date → date 객체."""
        if val is None:
            return None
        from datetime import date, datetime
        if isinstance(val, date):
            return val if isinstance(val, date) and not isinstance(val, datetime) else val.date()
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    # ── 헤더 정규화 (다양한 표기 허용) ─────────────────────
    _HEADER_ALIASES: dict[str, str] = {
        # 공고번호
        "공고번호": "공고번호", "announcement_no": "공고번호",
        # 공고제목
        "공고제목": "공고제목", "공고명": "공고제목", "제목": "공고제목", "title": "공고제목",
        # 발주처
        "발주처": "발주처", "발주기관": "발주처", "agency": "발주처", "agency_name": "발주처",
        # 입찰일
        "입찰일": "입찰일", "투찰일": "입찰일", "bid_date": "입찰일",
        # 기초금액
        "기초금액": "기초금액", "예정가격": "기초금액", "base_amount": "기초금액",
        # 제출투찰률
        "제출투찰률": "제출투찰률", "투찰률": "제출투찰률", "투찰율": "제출투찰률",
        "submitted_rate": "제출투찰률",
        # 추천투찰률
        "추천투찰률": "추천투찰률", "ai추천률": "추천투찰률", "ai추천율": "추천투찰률",
        "recommendation_rate": "추천투찰률",
        # 결과
        "결과": "결과", "result": "결과", "낙패": "결과",
        # 실제낙찰률
        "실제낙찰률": "실제낙찰률", "낙찰률": "실제낙찰률", "actual_winner_rate": "실제낙찰률",
        # 비고
        "비고": "비고", "메모": "비고", "note": "비고",
    }

    def _normalize_headers(self, raw: list[str]) -> list[str]:
        return [self._HEADER_ALIASES.get(h.strip(), h.strip()) for h in raw]

    # ── 메인 임포트 ─────────────────────────────────────────
    def import_excel(self, file_bytes: bytes, user_id: int) -> dict:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        headers: list[str] = []
        imported = skipped = 0
        errors: list[str] = []
        details: list[str] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not headers:
                headers = self._normalize_headers([str(c or "") for c in row])
                continue

            if all(v is None or str(v).strip() == "" for v in row):
                continue

            rd = dict(zip(headers, row))
            title = str(rd.get("공고제목") or "").strip()
            if not title:
                skipped += 1
                continue

            announcement_no = str(rd.get("공고번호") or "").strip() or None

            # 중복 체크: 같은 사용자의 동일 공고번호
            if announcement_no:
                exists = self.db.query(MyBidRecord).filter(
                    MyBidRecord.user_id == user_id,
                    MyBidRecord.announcement_no == announcement_no,
                ).first()
                if exists:
                    skipped += 1
                    details.append(f"중복 건너뜀: {title[:40]} ({announcement_no})")
                    continue

            submitted_rate_raw = self._to_float(rd.get("제출투찰률"))
            if submitted_rate_raw is None:
                errors.append(f"행 {row_idx}: 제출투찰률 없음 — {title[:40]}")
                skipped += 1
                continue

            # 투찰률이 퍼센트 표기(예: 87.123)이면 /100
            submitted_rate = submitted_rate_raw / 100 if submitted_rate_raw > 1 else submitted_rate_raw

            result_raw = str(rd.get("결과") or "").strip()
            result = self.RESULT_MAP.get(result_raw, "pending")

            actual_winner_raw = self._to_float(rd.get("실제낙찰률"))
            actual_winner_rate = None
            if actual_winner_raw is not None:
                actual_winner_rate = actual_winner_raw / 100 if actual_winner_raw > 1 else actual_winner_raw

            rec_raw = self._to_float(rd.get("추천투찰률"))
            recommendation_rate = None
            if rec_raw is not None:
                recommendation_rate = rec_raw / 100 if rec_raw > 1 else rec_raw

            rate_diff = None
            if submitted_rate is not None and actual_winner_rate is not None:
                rate_diff = round(submitted_rate - actual_winner_rate, 6)

            try:
                rec = MyBidRecord(
                    user_id=user_id,
                    announcement_no=announcement_no,
                    title=title,
                    agency_name=str(rd.get("발주처") or "").strip() or None,
                    bid_date=self._to_date(rd.get("입찰일")),
                    base_amount=self._to_int(rd.get("기초금액")),
                    submitted_rate=submitted_rate,
                    recommendation_rate=recommendation_rate,
                    result=result,
                    actual_winner_rate=actual_winner_rate,
                    rate_diff=rate_diff,
                    note=str(rd.get("비고") or "").strip() or None,
                )
                self.db.add(rec)
                self.db.flush()
                MyBidFeedbackService(self.db).sync_outcome(rec)
                imported += 1
                details.append(f"등록: {title[:40]}")
            except Exception as exc:
                self.db.rollback()
                errors.append(f"행 {row_idx} 저장 실패: {exc}")
                skipped += 1

        self.db.commit()
        return {
            "imported": imported,
            "skipped": skipped,
            "competitors_added": 0,
            "errors": errors[:30],
            "details": details[:50],
        }


# ==================================================
# 투찰 정확도 분석 서비스
# ==================================================

class MyBidAnalysisService:
    def __init__(self, db: Session):
        self.db = db

    def analyze(self, user_id: int) -> dict:
        from collections import defaultdict
        records = self.db.query(MyBidRecord).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.result.in_(["won", "lost"])
        ).order_by(MyBidRecord.bid_date).all()

        if not records:
            return {
                "accuracy_stats": {"avg_error": None, "median_error": None, "accuracy_1pct": None, "accuracy_3pct": None, "total_records": 0},
                "rate_scatter": [],
                "monthly_accuracy": []
            }

        errors = []
        scatter = []
        for r in records:
            if r.submitted_rate and r.recommendation_rate:
                err = abs(float(r.submitted_rate) - float(r.recommendation_rate))
                errors.append(err)
            scatter.append({
                "submitted_rate": float(r.submitted_rate) if r.submitted_rate else 0.0,
                "recommendation_rate": float(r.recommendation_rate) if r.recommendation_rate else None,
                "result": r.result,
                "bid_date": r.bid_date.strftime("%Y-%m-%d") if r.bid_date else ""
            })

        arr = np.array(errors) if errors else np.array([])
        accuracy_stats = {
            "avg_error": float(np.mean(arr)) if len(arr) > 0 else None,
            "median_error": float(np.median(arr)) if len(arr) > 0 else None,
            "accuracy_1pct": float(np.mean(arr <= 1.0)) if len(arr) > 0 else None,
            "accuracy_3pct": float(np.mean(arr <= 3.0)) if len(arr) > 0 else None,
            "total_records": len(records)
        }

        monthly = defaultdict(lambda: {"errors": [], "win_count": 0, "total": 0})
        for r in records:
            if r.bid_date:
                key = r.bid_date.strftime("%Y-%m")
                monthly[key]["total"] += 1
                if r.result == "won":
                    monthly[key]["win_count"] += 1
                if r.submitted_rate and r.recommendation_rate:
                    monthly[key]["errors"].append(abs(float(r.submitted_rate) - float(r.recommendation_rate)))

        monthly_accuracy = [
            {
                "year_month": k,
                "mae": round(float(np.mean(v["errors"])), 4) if v["errors"] else None,
                "win_count": v["win_count"],
                "total": v["total"]
            }
            for k, v in sorted(monthly.items())
        ]

        return {
            "accuracy_stats": accuracy_stats,
            "rate_scatter": scatter,
            "monthly_accuracy": monthly_accuracy
        }




# ==================================================
# ② 패찰 원인 분석 (MyBidAnalysisService 확장)
# ==================================================

class DefeatAnalysisService:
    """
    패찰 이력에서 원인 분석:
    - 낙찰자 대비 얼마나 높게/낮게 입찰했는지
    - 발주처별 패턴
    - 시간 흐름에 따른 개선 추이
    """

    def __init__(self, db: Session):
        self.db = db

    def analyze(self, user_id: int) -> dict:
        from collections import defaultdict

        records = self.db.query(MyBidRecord).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.result == "lost",
            MyBidRecord.actual_winner_rate.isnot(None),
            MyBidRecord.submitted_rate.isnot(None),
        ).order_by(MyBidRecord.bid_date.desc()).limit(500).all()

        if not records:
            return self._empty()

        # rate_diff (percentage point): winner_rate - submitted_rate
        diffs = []
        for r in records:
            if r.rate_diff is not None:
                d = float(r.rate_diff)
            else:
                d = (float(r.actual_winner_rate) - float(r.submitted_rate)) * 100
            diffs.append(d)

        arr = np.array(diffs)
        # 아웃라이어 제거 (|diff| > 5%)
        clean = arr[np.abs(arr) <= 5.0]

        miss_stats = self._compute_miss_stats(clean if len(clean) > 0 else arr)

        # 히스토그램 bins: -5 ~ +5, 0.2 단위
        bins = np.arange(-5.0, 5.2, 0.2)
        hist, edges = np.histogram(clean if len(clean) > 0 else arr, bins=bins)
        distribution = [
            {"from": round(float(edges[i]), 2), "to": round(float(edges[i+1]), 2), "count": int(hist[i])}
            for i in range(len(hist))
        ]

        # 발주처별 분석
        agency_map = defaultdict(list)
        for r, d in zip(records, diffs):
            if r.agency_name and abs(d) <= 5.0:
                agency_map[r.agency_name].append(d)

        agency_breakdown = []
        for name, ds in sorted(agency_map.items(), key=lambda x: len(x[1]), reverse=True)[:15]:
            a = np.array(ds)
            agency_breakdown.append({
                "agency_name": name,
                "count": len(ds),
                "avg_diff": round(float(np.mean(a)), 4),
                "direction": "too_low" if np.mean(a) > 0.15 else "too_high" if np.mean(a) < -0.15 else "balanced",
            })

        # 월별 추이
        monthly_map = defaultdict(list)
        for r, d in zip(records, diffs):
            if r.bid_date and abs(d) <= 5.0:
                key = r.bid_date.strftime("%Y-%m")
                monthly_map[key].append(d)

        trend = [
            {
                "year_month": k,
                "avg_diff": round(float(np.mean(monthly_map[k])), 4),
                "count": len(monthly_map[k]),
            }
            for k in sorted(monthly_map.keys())
        ]

        # 낙찰 구간 분석 (won 레코드)
        won_records = self.db.query(MyBidRecord).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.result == "won",
            MyBidRecord.actual_winner_rate.isnot(None),
        ).all()
        win_zone = None
        if won_records:
            won_diffs = [float(r.rate_diff or 0) for r in won_records]
            win_zone = {
                "avg_diff": round(float(np.mean(won_diffs)), 4),
                "sample_count": len(won_diffs),
                "note": "낙찰 시 낙찰자 요율 대비 본인 요율 평균 차이",
            }

        return {
            "miss_stats": miss_stats,
            "distribution": distribution,
            "agency_breakdown": agency_breakdown,
            "trend": trend,
            "win_zone": win_zone,
            "total_analyzed": len(records),
        }

    def _compute_miss_stats(self, arr: np.ndarray) -> dict:
        if len(arr) == 0:
            return {}
        avg = float(np.mean(arr))
        return {
            "avg_diff_pct":      round(avg, 4),
            "median_diff_pct":   round(float(np.median(arr)), 4),
            "std_diff_pct":      round(float(np.std(arr)), 4),
            "pct_too_low":       round(float(np.mean(arr > 0.15)), 4),
            "pct_too_high":      round(float(np.mean(arr < -0.15)), 4),
            "pct_balanced":      round(float(np.mean(np.abs(arr) <= 0.15)), 4),
            "direction":         "too_low" if avg > 0.15 else "too_high" if avg < -0.15 else "balanced",
            "within_0_5pct":     round(float(np.mean(np.abs(arr) <= 0.5)), 4),
            "within_1pct":       round(float(np.mean(np.abs(arr) <= 1.0)), 4),
        }

    def _empty(self) -> dict:
        return {
            "miss_stats": {}, "distribution": [],
            "agency_breakdown": [], "trend": [],
            "win_zone": None, "total_analyzed": 0,
        }

    def get_gap_distribution(self, user_id: int) -> dict:
        """
        rate_diff 분포 분석 (0.005 버킷).
        diff = submitted_rate - actual_winner_rate (소수점 기준)
        양수 → 낙찰자보다 높게 투찰 (too_high), 음수 → 낙찰자보다 낮게 투찰 (too_low).
        """
        personal_bias = PersonalBiasAnalyzer().compute(self.db, user_id)

        records = (
            self.db.query(MyBidRecord)
            .filter(
                MyBidRecord.user_id == user_id,
                MyBidRecord.actual_winner_rate.isnot(None),
                MyBidRecord.submitted_rate.isnot(None),
            )
            .order_by(MyBidRecord.bid_date.desc())
            .limit(500)
            .all()
        )

        if not records:
            return self._gap_empty(personal_bias)

        OUTLIER = 0.05  # 5%p 이상 이상치 제거
        BUCKET = 0.005  # 소수점 기준 버킷 크기

        diffs = []
        for r in records:
            d = float(r.submitted_rate) - float(r.actual_winner_rate)
            if abs(d) <= OUTLIER:
                diffs.append(d)

        if not diffs:
            return self._gap_empty(personal_bias)

        arr = np.array(diffs)
        mean_diff = round(float(np.mean(arr)), 6)
        median_diff = round(float(np.median(arr)), 6)

        THRESHOLD = 0.0015  # 0.15%p
        if mean_diff > THRESHOLD:
            consistent_direction = "too_high"
        elif mean_diff < -THRESHOLD:
            consistent_direction = "too_low"
        else:
            consistent_direction = "mixed"

        win_if_lower_by = round(abs(mean_diff), 6) if consistent_direction == "too_high" else None

        bucket_counts: dict[float, int] = {}
        for d in diffs:
            key = round(math.floor(d / BUCKET) * BUCKET, 3)
            bucket_counts[key] = bucket_counts.get(key, 0) + 1

        buckets = [
            {
                "range_lo": lo,
                "range_hi": round(lo + BUCKET, 3),
                "count": cnt,
            }
            for lo, cnt in sorted(bucket_counts.items())
        ]

        return {
            "buckets": buckets,
            "mean_diff": mean_diff,
            "median_diff": median_diff,
            "win_if_lower_by": win_if_lower_by,
            "consistent_direction": consistent_direction,
            "personal_bias": personal_bias,
            "total_analyzed": len(diffs),
        }

    def _gap_empty(self, personal_bias: dict) -> dict:
        return {
            "buckets": [],
            "mean_diff": None,
            "median_diff": None,
            "win_if_lower_by": None,
            "consistent_direction": "mixed",
            "personal_bias": personal_bias,
            "total_analyzed": 0,
        }


# ==================================================
# 경쟁사 투찰 구간 분포 서비스
# ==================================================

class CompetitorZoneService:
    BUCKET_SIZE = 0.005
    ZONE_MIN    = 0.800
    ZONE_MAX    = 0.980

    def get_recent_zones(self, db: Session, competitor_id: int, days: int = 90) -> dict:
        from sqlalchemy import text as sa_text
        competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not competitor or not competitor.biz_reg_no:
            return self._empty()

        # inpo21c_participants에 date 컬럼 없으므로 전체 조회 (days 파라미터는 API 호환용)
        rows = db.execute(sa_text("""
            SELECT base_ratio::float
            FROM inpo21c_participants
            WHERE biz_reg_no = :biz_reg_no
              AND base_ratio IS NOT NULL
              AND base_ratio BETWEEN :lo AND :hi
        """), {
            "biz_reg_no": competitor.biz_reg_no,
            "lo": self.ZONE_MIN,
            "hi": self.ZONE_MAX,
        }).fetchall()

        if not rows:
            return self._empty()

        total = len(rows)
        bucket_counts: dict[float, int] = {}
        for (ratio,) in rows:
            key = round(math.floor(ratio / self.BUCKET_SIZE) * self.BUCKET_SIZE, 3)
            bucket_counts[key] = bucket_counts.get(key, 0) + 1

        zones = [
            {
                "range_lo": lo,
                "range_hi": round(lo + self.BUCKET_SIZE, 3),
                "count":    cnt,
                "pct":      round(cnt / total * 100, 1),
            }
            for lo, cnt in sorted(bucket_counts.items())
        ]
        peak_zone = max(zones, key=lambda z: z["count"]) if zones else None

        return {
            "zones":       zones,
            "peak_zone":   peak_zone,
            "total_count": total,
            "last_updated": None,
        }

    def _empty(self) -> dict:
        return {"zones": [], "peak_zone": None, "total_count": 0, "last_updated": None}


# ==================================================
# ⑦-2 경쟁사 행동 예측 서비스
# ==================================================

class CompetitorPredictService:
    """inpo21c 데이터 기반 경쟁사 참여 확률 + 투찰 구간 예측."""

    def predict(self, db: Session, competitor_id: int, bid_id: int) -> dict:
        from fastapi import HTTPException
        from .ml.competitor_predict import predict_participation, predict_bid_zone

        competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not competitor:
            raise HTTPException(status_code=404, detail="경쟁사를 찾을 수 없습니다.")

        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            raise HTTPException(status_code=404, detail="공고를 찾을 수 없습니다.")

        bid_dict = {
            "agency_id":   bid.agency_id,
            "industry_id": bid.industry_id,
            "base_amount": bid.base_amount,
        }

        participation = predict_participation(competitor_id, bid_dict, db)
        bid_zone      = predict_bid_zone(competitor_id, bid.base_amount, db)

        return {
            "competitor_id":   competitor_id,
            "competitor_name": competitor.name,
            "bid_id":          bid_id,
            "participation":   participation,
            "bid_zone":        bid_zone,
        }


# ==================================================
# ⑧ 공고 자동 평가 점수 서비스
# ==================================================

class OpportunityScoreService:
    """
    신규 공고에 대해 내게 유리한 정도를 0~100점으로 산출.
    점수 구성:
      - 경쟁 약함 (40점): HHI 낮고 경쟁자 수 적을수록
      - 내 이력 (30점): 이 발주처에서 본인 낙찰 이력 있을수록
      - 시장 추세 (15점): 최근 낙찰율 상승 추세
      - 금액 적합 (15점): 내가 자주 참여하는 금액 구간
    """

    def __init__(self, db: Session):
        self.db = db

    def score(self, bid_id: int, user_id: int) -> dict:
        bid = self.db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            return {"score": None, "error": "공고를 찾을 수 없습니다."}

        agency_score  = self._agency_track_score(user_id, bid.agency_id, bid.agency)
        comp_score    = self._competition_score(bid)
        trend_score   = self._trend_score(bid)
        amount_score  = self._amount_fit_score(user_id, bid.base_amount)

        total = (comp_score["pts"] + agency_score["pts"] +
                 trend_score["pts"] + amount_score["pts"])
        total = round(min(100, max(0, total)), 1)

        grade = "A" if total >= 75 else "B" if total >= 55 else "C" if total >= 35 else "D"

        return {
            "bid_id": bid_id,
            "score": total,
            "grade": grade,
            "breakdown": {
                "competition":    comp_score,
                "personal_track": agency_score,
                "market_trend":   trend_score,
                "amount_fit":     amount_score,
            },
            "recommendation": self._grade_message(grade, comp_score, agency_score),
        }

    def get_top_recommended(self, user_id: int, limit: int = 5) -> list:
        """7일 이내 개찰 예정 open 공고 중 점수 상위 limit개 반환."""
        now = datetime.now(timezone.utc)
        cutoff = now + timedelta(days=7)

        q = self.db.query(Bid).filter(
            Bid.status == "open",
            Bid.bid_open_date >= now,
            Bid.bid_open_date <= cutoff,
        )

        active_ids = get_active_industry_ids(self.db)
        if active_ids is not None:
            if not active_ids:
                return []
            q = q.filter(Bid.industry_id.in_(active_ids))

        bids = q.all()

        results = []
        for bid in bids:
            scored = self.score(bid.id, user_id)
            if scored.get("error"):
                continue
            results.append({
                "bid_id": bid.id,
                "title": bid.title,
                "agency_name": bid.agency.name if bid.agency else "",
                "score": scored["score"],
                "grade": scored["grade"],
                "open_date": bid.bid_open_date.isoformat() if bid.bid_open_date else None,
                "base_amount": bid.base_amount,
                "score_breakdown": scored["breakdown"],
            })

        results.sort(key=lambda x: (x["score"] or 0), reverse=True)
        return results[:limit]

    def _competition_score(self, bid: "Bid") -> dict:
        # 최근 해당 발주처 + 업종 낙찰 데이터로 경쟁강도 추정
        cutoff = datetime.now(timezone.utc) - timedelta(days=180)
        rows = self.db.execute(text("""
            SELECT COUNT(r.id) as comp_count
            FROM bids b
            JOIN bid_results r ON r.bid_id = b.id
            WHERE b.agency_id = :aid AND b.industry_id = :iid
              AND b.bid_open_date >= :cutoff AND b.status = 'closed'
            GROUP BY b.id
        """), {"aid": bid.agency_id, "iid": bid.industry_id, "cutoff": cutoff}).fetchall()

        if not rows:
            pts = 25.0
            note = "경쟁 데이터 없음 (중간값 적용)"
        else:
            avg_comp = float(np.mean([r.comp_count for r in rows]))
            # 경쟁자 수 적을수록 좋음: 5명 이하 = 40pt, 15명 이상 = 5pt
            pts = max(5.0, 40.0 - (avg_comp - 1) * 2.5)
            pts = min(40.0, pts)
            note = f"최근 6개월 평균 경쟁사 {avg_comp:.1f}명"

        return {"pts": round(pts, 1), "max": 40, "note": note}

    def _agency_track_score(self, user_id: int, agency_id: int, agency) -> dict:
        if not agency:
            return {"pts": 0.0, "max": 30, "note": "발주처 정보 없음"}

        agency_name = agency.name if hasattr(agency, "name") else ""
        total = self.db.query(func.count(MyBidRecord.id)).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.agency_name.ilike(f"%{agency_name[:10]}%"),
        ).scalar() or 0

        won = self.db.query(func.count(MyBidRecord.id)).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.agency_name.ilike(f"%{agency_name[:10]}%"),
            MyBidRecord.result == "won",
        ).scalar() or 0

        if total == 0:
            pts, note = 15.0, "이 발주처 참여 이력 없음 (중간값)"
        elif won == 0:
            pts = max(5.0, 15.0 - total * 0.5)
            note = f"참여 {total}건 / 낙찰 0건"
        else:
            win_rate = won / total
            pts = 15.0 + win_rate * 15.0
            note = f"참여 {total}건 / 낙찰 {won}건 ({win_rate:.0%})"

        return {"pts": round(min(30.0, pts), 1), "max": 30, "note": note}

    def _trend_score(self, bid: "Bid") -> dict:
        cutoff = datetime.now(timezone.utc) - timedelta(days=60)
        prev   = datetime.now(timezone.utc) - timedelta(days=120)
        rows = self.db.execute(text("""
            SELECT b.bid_open_date, r.bid_rate
            FROM bids b
            JOIN bid_results r ON r.bid_id = b.id AND r.is_winner = true
            WHERE b.agency_id = :aid AND b.industry_id = :iid
              AND b.bid_open_date >= :prev AND b.status = 'closed'
        """), {"aid": bid.agency_id, "iid": bid.industry_id, "prev": prev}).fetchall()

        if len(rows) < 4:
            return {"pts": 8.0, "max": 15, "note": "추세 데이터 부족"}

        recent = [float(r.bid_rate) for r in rows if r.bid_open_date >= cutoff]
        older  = [float(r.bid_rate) for r in rows if r.bid_open_date < cutoff]

        if not recent or not older:
            return {"pts": 8.0, "max": 15, "note": "추세 계산 불가"}

        diff = np.mean(recent) - np.mean(older)
        # 낙찰률 상승은 예가 가까워짐 → 진입 기회
        if diff > 0.002:
            pts, note = 15.0, f"최근 낙찰률 상승 추세 (+{diff*100:.2f}%)"
        elif diff < -0.002:
            pts, note = 5.0,  f"최근 낙찰률 하락 추세 ({diff*100:.2f}%)"
        else:
            pts, note = 10.0, "낙찰률 안정적"

        return {"pts": pts, "max": 15, "note": note}

    def _amount_fit_score(self, user_id: int, base_amount: int) -> dict:
        rows = self.db.query(MyBidRecord.base_amount).filter(
            MyBidRecord.user_id == user_id,
            MyBidRecord.base_amount.isnot(None),
            MyBidRecord.base_amount > 0,
        ).all()

        if not rows:
            return {"pts": 8.0, "max": 15, "note": "금액 이력 없음"}

        amounts = [r.base_amount for r in rows]
        p10, p90 = np.percentile(amounts, 10), np.percentile(amounts, 90)

        if p10 <= base_amount <= p90:
            pts, note = 15.0, f"자주 참여하는 금액 구간 ({base_amount/1e6:.0f}백만원)"
        elif base_amount < p10:
            ratio = base_amount / max(p10, 1)
            pts = max(5.0, 15.0 * ratio)
            note = f"평소보다 소액 ({base_amount/1e6:.0f}백만원)"
        else:
            ratio = p90 / max(base_amount, 1)
            pts = max(5.0, 15.0 * ratio)
            note = f"평소보다 고액 ({base_amount/1e6:.0f}백만원)"

        return {"pts": round(pts, 1), "max": 15, "note": note}

    def _grade_message(self, grade: str, comp: dict, track: dict) -> str:
        if grade == "A":
            return "낙찰 가능성이 높은 유망 공고입니다. 적극 참여를 추천합니다."
        elif grade == "B":
            return "참여 가능한 공고입니다. 요율 전략 수립 후 참여하세요."
        elif grade == "C":
            if comp["pts"] < 20:
                return "경쟁이 치열한 공고입니다. 신중하게 참여 여부를 검토하세요."
            return "발주처 이력이 부족합니다. 정보 수집 후 참여 여부를 결정하세요."
        else:
            return "불리한 조건의 공고입니다. 다른 공고를 우선 검토하세요."





# ==================================================
# G2B 개찰 결과 자동 연계 서비스
# ==================================================

class G2BSyncService:
    """
    G2B 개찰 결과를 투찰이력(my_bid_records)에 자동 반영.

    개찰 완료된 공고의 낙찰자 정보를 기반으로:
    - result: 'won' / 'lost' 자동 설정
    - actual_winner_rate, winner_name, winner_biz_no, rate_diff 채우기

    낙찰 판정 기준:
      abs(submitted_rate - winner_bid_rate) < 0.0003 → won
      그 외 → lost
    """

    def sync(self, db: Session) -> dict:
        pending = (
            db.query(MyBidRecord)
            .filter(
                MyBidRecord.result.is_(None),
                MyBidRecord.bid_id.isnot(None),
            )
            .all()
        )

        won = lost = skipped = 0
        for rec in pending:
            bid = db.query(Bid).filter(Bid.id == rec.bid_id).first()
            if not bid or bid.status != "closed":
                skipped += 1
                continue

            winner_row = (
                db.query(BidResult, Competitor)
                .join(Competitor, Competitor.id == BidResult.competitor_id)
                .filter(BidResult.bid_id == rec.bid_id, BidResult.is_winner == True)
                .first()
            )
            if not winner_row:
                skipped += 1
                continue

            winner_result, winner_comp = winner_row
            winner_rate     = float(winner_result.bid_rate)
            submitted_rate  = float(rec.submitted_rate)

            rec.actual_winner_rate = winner_rate
            rec.winner_name        = winner_comp.name
            rec.winner_biz_no      = winner_comp.biz_reg_no
            rec.rate_diff          = round(submitted_rate - winner_rate, 4)

            if abs(submitted_rate - winner_rate) < 0.0003:
                rec.result = "won"
                won += 1
            else:
                rec.result = "lost"
                lost += 1
                # 惜敗 알림: 낙찰자와 근소한 차이(1%p 이내) — 2위 가능성
                rate_diff_pct = abs(submitted_rate - winner_rate) * 100
                if rate_diff_pct < 1.0:
                    self._notify_seikihai(db, rec, winner_rate, rate_diff_pct)

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            logger.error("G2B 연계 커밋 실패: %s", e)
            return {"won": 0, "lost": 0, "skipped": skipped, "error": str(e)}

        logger.info("G2B 자동 연계: won=%d, lost=%d, skipped=%d", won, lost, skipped)
        return {"won": won, "lost": lost, "skipped": skipped}

    def _notify_seikihai(self, db: Session, rec, winner_rate: float, diff_pct: float) -> None:
        try:
            title = f"[惜敗 알림] {rec.title[:40]}"
            body = (
                f"낙찰자와 {diff_pct:.2f}%p 차이로 아깝게 패찰했습니다. "
                f"낙찰율: {winner_rate*100:.3f}%, 귀사 투찰율: {float(rec.submitted_rate)*100:.3f}%"
            )
            link = f"/my-bids"
            users = db.query(User).filter(User.is_active == True).all()
            for u in users:
                n = Notification(user_id=u.id, ntype="seikihai", title=title, body=body, link=link)
                db.add(n)
        except Exception as exc:
            logger.warning("惜敗 알림 생성 실패: %s", exc)


class AgencyYegaService:
    """발주처 특화 예가 번호 빈도 패턴 분석."""

    def __init__(self, db: Session):
        self.db = db

    def get_pattern(self, agency_id: int, industry_id: Optional[int] = None, months: int = 12) -> dict:
        from .ml.yega import get_agency_yega_pattern

        # inpo21c 실측 데이터 우선 (역산 방식보다 정확)
        try:
            direct = get_inpo21c_pattern_direct(self.db, agency_id)
            if direct.get("sample_count", 0) >= 3:
                return {**direct, "source": "inpo21c_direct"}
        except Exception:
            pass

        cutoff = datetime.utcnow() - timedelta(days=months * 30)

        query = (
            self.db.query(
                BidResult.assessment_rate,
                Bid.base_amount,
                Bid.a_value,
            )
            .join(Bid, BidResult.bid_id == Bid.id)
            .filter(
                Bid.agency_id == agency_id,
                BidResult.assessment_rate.isnot(None),
                Bid.bid_open_date >= cutoff,
            )
        )

        if industry_id:
            query = query.filter(Bid.industry_id == industry_id)

        rows = query.limit(500).all()

        bid_data = [
            {
                "assessment_rate": float(r.assessment_rate),
                "base_amount":     int(r.base_amount),
                "a_value":         int(r.a_value) if r.a_value else None,
            }
            for r in rows
        ]

        result = get_agency_yega_pattern(bid_data)
        return {**result, "source": "reverse_calc"}


# ==================================================
# 공동도급 적격심사 매칭 서비스
# ==================================================

class JointQualService:
    """공동도급 파트너 적격심사 AI 매칭.
    경쟁사 DB 기반으로 협정 가능 업체를 탐색하고 적격 여부를 추정한다.
    실제 실적 데이터 없이 낙찰 이력을 프록시로 사용하는 간소화 모델.
    """

    _MIN_PARTNER_RATE = 0.30  # 부계약자 최소 참여지분율 (건설산업기본법)

    def __init__(self, db: Session):
        self.db = db

    def find_matching_partners(
        self,
        bid_id: int,
        user_track_amount: float,
        participation_rate: float = 0.6,
    ) -> dict:
        from fastapi import HTTPException
        from sqlalchemy import func as sqlfunc

        bid = self.db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            raise HTTPException(status_code=404, detail="공고를 찾을 수 없습니다")

        base_amount = bid.base_amount or 0
        min_bid_rate = float(bid.min_bid_rate or 0.87745)
        partner_rate = max(1.0 - participation_rate, self._MIN_PARTNER_RATE)
        min_bid_amount = int(base_amount * min_bid_rate * participation_rate)

        threshold_note = (
            f"기초금액 {base_amount:,}원 기준 — "
            f"파트너 최소 지분 {partner_rate:.0%}, "
            f"귀사 최소 투찰금액 약 {min_bid_amount:,}원"
        )

        # 최근 24개월 CompetitorStat 집계
        cutoff_year = (datetime.now() - timedelta(days=730)).year
        stats_subq = (
            self.db.query(
                CompetitorStat.competitor_id,
                sqlfunc.sum(CompetitorStat.total_bid_count).label("total_bids"),
                sqlfunc.sum(CompetitorStat.win_count).label("win_count"),
                sqlfunc.avg(CompetitorStat.avg_bid_rate).label("avg_rate"),
            )
            .filter(CompetitorStat.period_year >= cutoff_year)
            .group_by(CompetitorStat.competitor_id)
            .subquery()
        )

        rows = (
            self.db.query(Competitor, stats_subq.c.total_bids, stats_subq.c.win_count, stats_subq.c.avg_rate)
            .outerjoin(stats_subq, Competitor.id == stats_subq.c.competitor_id)
            .limit(500)
            .all()
        )

        partners = []
        for competitor, _tb, _wc, _ar in rows:
            total_bids = int(_tb or 0) if _tb is not None else 0
            win_count  = int(_wc or 0) if _wc is not None else 0
            avg_rate   = float(_ar or 0) if _ar is not None else 0.0
            win_rate = win_count / total_bids if total_bids > 0 else 0.0

            # 적격심사 통과 예상: 낙찰 이력 존재 + 최소 입찰 건수 기준
            qualification_ok = win_count >= 1 and total_bids >= 3

            # 궁합 점수 (안정성·실적·활동성)
            stability  = 40.0 if qualification_ok else 8.0
            perf       = min(win_rate / 0.3, 1.0) * 25
            activity   = min(total_bids / 50, 1.0) * 15
            compat     = round(stability + perf + activity, 1)

            partners.append({
                "competitor_id":    competitor.id,
                "name":             competitor.name,
                "biz_reg_no":       competitor.biz_reg_no,
                "joint_min_rate":   round(partner_rate, 2),
                "qualification_ok": qualification_ok,
                "win_rate":         round(win_rate, 4),
                "total_bids":       total_bids,
                "avg_bid_rate":     round(avg_rate, 4) if avg_rate else None,
                "compat_score":     compat,
            })

        # 적격 업체 우선, 궁합 점수 내림차순 정렬
        partners.sort(key=lambda x: (-int(x["qualification_ok"]), -x["compat_score"]))
        partners = partners[:50]

        return {
            "partners":       partners,
            "bid_title":      bid.title,
            "base_amount":    base_amount,
            "threshold_note": threshold_note,
        }


# ==================================================
# 공동도급 적격심사 시뮬레이터
# ==================================================

class JointSimulateService:
    """공동도급 파트너 구성 + 지분율 조합으로 심사통과 여부 및 최저 투찰가 계산."""

    # 기초금액 구간별 적격심사 기준점수 (간소화 모델)
    _THRESHOLD_TABLE = [
        (300_000_000,    12.0),   # 3억 미만
        (3_000_000_000,  14.0),   # 30억 미만
        (10_000_000_000, 16.0),   # 100억 미만
        (float("inf"),   18.0),   # 100억 이상
    ]

    def __init__(self, db: Session):
        self.db = db

    def simulate(self, bid_id: int, partners: list) -> dict:
        from fastapi import HTTPException

        bid = self.db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            raise HTTPException(status_code=404, detail="공고를 찾을 수 없습니다")

        base_amount  = int(bid.base_amount or 0)
        min_bid_rate = float(bid.min_bid_rate or 0.87745)

        # 기초금액에 맞는 기준점수 결정
        threshold = 18.0
        for limit, t in self._THRESHOLD_TABLE:
            if base_amount < limit:
                threshold = t
                break

        cutoff_year = (datetime.now() - timedelta(days=730)).year

        partner_results = []
        for p in partners:
            competitor_id      = p.get("competitor_id")
            user_track         = p.get("user_track")
            participation_rate = float(p.get("participation_rate", 0.0))

            if competitor_id:
                # 경쟁사 DB 조회
                stats_row = (
                    self.db.query(
                        func.sum(CompetitorStat.total_bid_count).label("total_bids"),
                        func.sum(CompetitorStat.win_count).label("win_count"),
                        func.avg(CompetitorStat.avg_bid_rate).label("avg_rate"),
                    )
                    .filter(
                        CompetitorStat.competitor_id == competitor_id,
                        CompetitorStat.period_year >= cutoff_year,
                    )
                    .first()
                )
                competitor = self.db.query(Competitor).filter(Competitor.id == competitor_id).first()
                name       = competitor.name if competitor else f"업체 #{competitor_id}"

                total_bids = int(stats_row.total_bids or 0) if stats_row else 0
                win_count  = int(stats_row.win_count  or 0) if stats_row else 0
                avg_rate   = float(stats_row.avg_rate or 0.87) if stats_row else 0.87

                # 낙찰 이력으로 시공실적 추정
                track_amount = int(base_amount * avg_rate * win_count)

                if win_count >= 1 and total_bids >= 3:
                    win_rate   = win_count / total_bids
                    perf_score = 8.0 + min(win_rate / 0.3, 1.0) * 6.0 + min(total_bids / 30, 1.0) * 6.0
                else:
                    perf_score = 4.0

            else:
                # 귀사
                name         = "귀사"
                track_amount = int(user_track or 0)
                ratio        = track_amount / base_amount if base_amount > 0 else 0
                perf_score   = min(20.0, 5.0 + ratio * 15.0)

            qual_score     = round(perf_score * participation_rate, 2)
            partner_passes = perf_score >= 8.0  # 개별 최소 충족 기준

            partner_results.append({
                "name":               name,
                "participation_rate": participation_rate,
                "track_amount":       track_amount,
                "qual_score":         qual_score,
                "passes":             partner_passes,
            })

        total_qual_score    = round(sum(r["qual_score"] for r in partner_results), 2)
        joint_passes        = total_qual_score >= threshold
        bid_amount_required = int(base_amount * min_bid_rate)
        margin              = int(bid_amount_required * 0.003)

        return {
            "bid_id":              bid_id,
            "bid_amount_required": bid_amount_required,
            "partners":            partner_results,
            "joint_result": {
                "passes":           joint_passes,
                "total_qual_score": total_qual_score,
                "threshold":        threshold,
                "min_bid_amount":   bid_amount_required,
                "min_bid_rate":     round(min_bid_rate, 4),
                "margin":           margin,
            },
        }


# ==================================================
# 최종 투찰 추천 종합 서비스
# ==================================================

class FinalRecommendService:
    """사정율통계 + 프리즘 + 예가 + 트렌드 + 개인화를 합산해 최종 투찰 사정율 1개 산출."""

    def __init__(self, db: Session):
        self.db = db

    def get(self, bid_id: int, user_id: int) -> dict:
        from fastapi import HTTPException

        bid = self.db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            raise HTTPException(status_code=404, detail="공고를 찾을 수 없습니다.")

        agency   = self.db.query(Agency).filter(Agency.id == bid.agency_id).first()
        industry = (self.db.query(Industry).filter(Industry.id == bid.industry_id).first()
                    if bid.industry_id else None)
        industry_name = industry.name if industry else ""
        base_amount   = int(bid.base_amount)

        # 1. 사정율 통계 (기준 mean + 표본수)
        features = load_srate_stats(
            self.db, bid.agency_id, bid.industry_id or 0, bid.region_id or 0, base_amount
        )
        srate_mean = (
            features.get("agency_srate_mean")
            or features.get("industry_srate_mean")
            or features.get("global_srate_mean")
            or 0.8876
        )
        sample_count = int(features.get("agency_srate_n") or 0)

        # 2. 트렌드 방향 (SrateTrendService 재사용)
        trend           = SrateTrendService().get_trend(self.db, bid.agency_id, bid.industry_id)
        trend_direction = trend.get("direction", "stable")

        # 3. 프리즘 스캔 (전 구간 win_prob + top1 근거)
        try:
            all_zones, top10 = scan_prism_zones(
                base_amount=base_amount,
                industry_name=industry_name,
                agency_id=bid.agency_id,
                industry_id=bid.industry_id or 0,
                db=self.db,
                n_sim=15_000,
            )
        except Exception:
            all_zones, top10 = [], []

        prism_top = top10[0] if top10 else None

        # 4. 예가 빈도 top1 (크로스체크 근거) — yega_range를 inpo21c_bid_notices에서 동적 조회
        try:
            a_val = int(bid.a_value) if bid.a_value else None
            spread_half = 0.028  # fallback
            if bid.announcement_no:
                ibn_row = self.db.execute(text(
                    "SELECT yega_range_max FROM inpo21c_bid_notices "
                    "WHERE announcement_no LIKE :ano ORDER BY announcement_no LIMIT 1"
                ), {"ano": bid.announcement_no + "%"}).fetchone()
                if ibn_row and ibn_row[0] is not None:
                    spread_half = abs(int(ibn_row[0])) / 100.0
            yega   = calc_yega_frequency(base_amount, a_value=a_val, srate_center=srate_mean, spread_half=spread_half)
            yega_rows     = yega.get("frequency", [])
            yega_top_rate = float(yega_rows[0]["rate"])        if yega_rows else None
            yega_top_prob = float(yega_rows[0]["probability"]) if yega_rows else None
        except Exception:
            yega_top_rate = yega_top_prob = None

        # 5. 개인화 편향 보정
        try:
            bias = PersonalBiasAnalyzer().compute(
                self.db, user_id,
                agency_name=agency.name if agency else None,
            )
        except Exception:
            bias = PersonalBiasAnalyzer()._empty_result()

        # 6. 권장 사정율 합산
        trend_adj    = 0.001 if trend_direction == "up" else (-0.001 if trend_direction == "down" else 0.0)
        correction   = float(bias.get("correction", 0.0))
        bias_applied = abs(correction) > 0.0001

        recommended_rate = round(float(srate_mean) + trend_adj + correction, 4)

        # 7. 낙찰하한율 (A값 기준 floor → base_amount 기준 환산)
        floor_pct = calc_floor_rate(industry_name)
        if bid.min_bid_rate:
            floor_rate = round(float(bid.min_bid_rate), 5)
        else:
            floor_rate = round(float(srate_mean) * floor_pct, 5)

        # recommended_rate가 floor 미달이면 floor로 올림
        if recommended_rate < floor_rate:
            recommended_rate = floor_rate

        # 8. 전략별 근접 zone win_prob 탐색
        def _win_prob(rate: float) -> float:
            if rate < floor_rate:          # floor 미달 → 실격이므로 확률 0
                return 0.0
            if not all_zones:
                return 0.0
            nearest = min(all_zones, key=lambda z: abs(z["rate"] - rate))
            return round(float(nearest.get("win_prob", 0.0)), 4)

        def _strat(rate: float) -> dict:
            r = round(max(rate, floor_rate), 4)   # 항상 floor 이상으로 클램핑
            return {"rate": r, "amount": round(base_amount * r), "win_prob": _win_prob(r)}

        # recommended_rate가 floor에 클램핑됐을 때 aggressive/balanced가 모두
        # floor로 수렴하는 문제 수정.
        #
        # _gap = rec - floor:
        #   정상 케이스(gap 충분): 기존 ±0.5% 간격 그대로 유지
        #   클램핑 케이스(gap=0):  floor 기준 최소 간격 0.1%/0.3%/0.6% 보장
        #   → 세 전략이 항상 aggressive < balanced < conservative 순서 유지
        _gap = max(recommended_rate - floor_rate, 0.0)
        strategies = {
            "aggressive":   _strat(floor_rate + max(_gap - 0.005, 0.001)),
            "balanced":     _strat(floor_rate + max(_gap,         0.003)),
            "conservative": _strat(floor_rate + max(_gap + 0.005, 0.006)),
            "floor_safe":   _strat(floor_rate + 0.001),
        }

        # 9. 신뢰도 (표본 수 기준)
        confidence = "high" if sample_count >= 50 else ("medium" if sample_count >= 10 else "low")

        # 10. 시그널 메시지
        dir_msg = {
            "up":     "발주처 최근 사정율 상승 추세 → 균형형 이상 추천",
            "down":   "발주처 최근 사정율 하락 추세 → 공격형 이하 고려",
            "stable": "발주처 사정율 안정 추세 → 균형형 추천",
        }
        signal = dir_msg.get(trend_direction, "균형형 추천")
        if bias_applied:
            if bias.get("direction") == "too_low":
                signal += " · 과거 낮게 투찰 경향 → 상향 보정 적용"
            elif bias.get("direction") == "too_high":
                signal += " · 과거 높게 투찰 경향 → 하향 보정 적용"

        # 11. 근거 패널
        evidence = {
            "srate_stats": {
                "mean": round(float(srate_mean), 4),
                "sample_count": sample_count,
                "trend_direction": trend_direction,
            },
            "prism_top": {
                "rate":        round(float(prism_top["rate"]), 4),
                "probability": round(float(prism_top["win_prob"]), 4),
            } if prism_top else None,
            "yega_top": {
                "rate":        round(yega_top_rate, 4),
                "probability": round(yega_top_prob, 2),
            } if (yega_top_rate is not None and yega_top_prob is not None) else None,
            "personal_bias": {
                "rate_diff_mean": round(correction, 4),
                "applied":        bias_applied,
            },
        }

        return {
            "bid_id":             bid_id,
            "base_amount":        base_amount,
            "recommended_rate":   recommended_rate,
            "recommended_amount": round(base_amount * recommended_rate),
            "confidence":         confidence,
            "floor_rate":         floor_rate,
            "strategies":         strategies,
            "evidence":           evidence,
            "signal":             signal,
        }


# ==================================================
# 자사 승률 패턴 진단 서비스
# ==================================================

class WinPatternService:
    """
    my_bid_records를 분석해 편향 방향·패배 원인·발주처별 승률을 반환.
    rate_diff = submitted_rate - actual_winner_rate (소수점 기준)
    양수 = 낙찰자보다 높게 투찰(above), 음수 = 낮게 투찰(below).
    """

    BIAS_THRESHOLD = 0.003  # 0.3%p 이상이면 편향 있음

    def __init__(self, db: Session):
        self.db = db

    def analyze(self, user_id: int) -> dict:
        from collections import defaultdict

        total = self.db.query(MyBidRecord).filter(MyBidRecord.user_id == user_id).count()

        records = (
            self.db.query(MyBidRecord)
            .filter(
                MyBidRecord.user_id == user_id,
                MyBidRecord.result.in_(["won", "lost"]),
                MyBidRecord.actual_winner_rate.isnot(None),
                MyBidRecord.submitted_rate.isnot(None),
            )
            .all()
        )

        won_count = sum(1 for r in records if r.result == "won")
        lost_count = sum(1 for r in records if r.result == "lost")
        overall_win_rate = round(won_count / max(won_count + lost_count, 1) * 100, 2)

        diffs = []
        for r in records:
            if r.rate_diff is not None:
                diffs.append(float(r.rate_diff))
            elif r.submitted_rate and r.actual_winner_rate:
                diffs.append(float(r.submitted_rate) - float(r.actual_winner_rate))

        mean_diff = round(float(np.mean(diffs)), 6) if diffs else None
        bias = self._compute_bias(mean_diff)

        return {
            "total": total,
            "won": won_count,
            "lost": lost_count,
            "overall_win_rate": overall_win_rate,
            "bias": bias,
            "by_agency": self._by_agency(records),
            "by_industry": [],
            "by_year": self._by_year(records),
            "loss_reasons": self._loss_reasons([r for r in records if r.result == "lost"]),
        }

    def _compute_bias(self, mean_diff: Optional[float]) -> dict:
        if mean_diff is None:
            return {"rate_diff_mean": None, "direction": "balanced", "signal": "분석 데이터 없음"}
        direction = "above" if mean_diff > 0 else "below" if mean_diff < 0 else "balanced"
        has_bias = abs(mean_diff) > self.BIAS_THRESHOLD
        pct_str = f"{abs(mean_diff * 100):.2f}%p"
        if not has_bias:
            signal = "편향 없음 — 균형적으로 투찰하는 경향"
        elif direction == "above":
            signal = f"평균 {pct_str} 높게 투찰하는 경향 — 낮게 조정 권장"
        else:
            signal = f"평균 {pct_str} 낮게 투찰하는 경향 — 높게 조정 권장"
        return {"rate_diff_mean": mean_diff, "direction": direction, "signal": signal}

    def _by_agency(self, records: list) -> list:
        from collections import defaultdict
        agency_map: dict = defaultdict(lambda: {"total": 0, "won": 0, "diffs": []})
        for r in records:
            if not r.agency_name:
                continue
            agency_map[r.agency_name]["total"] += 1
            if r.result == "won":
                agency_map[r.agency_name]["won"] += 1
            diff = float(r.rate_diff) if r.rate_diff is not None else (
                float(r.submitted_rate) - float(r.actual_winner_rate)
                if r.submitted_rate and r.actual_winner_rate else None
            )
            if diff is not None:
                agency_map[r.agency_name]["diffs"].append(diff)
        return [
            {
                "agency_name": name,
                "total": v["total"],
                "won": v["won"],
                "win_rate": round(v["won"] / max(v["total"], 1) * 100, 1),
                "avg_rate_diff": round(float(np.mean(v["diffs"])), 4) if v["diffs"] else None,
            }
            for name, v in sorted(agency_map.items(), key=lambda x: -x[1]["total"])
        ]

    def _by_year(self, records: list) -> list:
        from collections import defaultdict
        year_map: dict = defaultdict(lambda: {"total": 0, "won": 0})
        for r in records:
            if r.bid_date:
                y = r.bid_date.year
                year_map[y]["total"] += 1
                if r.result == "won":
                    year_map[y]["won"] += 1
        return [
            {
                "year": year,
                "total": v["total"],
                "won": v["won"],
                "win_rate": round(v["won"] / max(v["total"], 1) * 100, 1),
            }
            for year, v in sorted(year_map.items())
        ]

    def _loss_reasons(self, lost_records: list) -> dict:
        above_winner = 0
        below_floor = 0
        below_winner = 0
        bid_ids = [r.bid_id for r in lost_records if r.bid_id is not None]
        floor_map: dict[int, Optional[float]] = {}
        if bid_ids:
            bids = self.db.query(Bid).filter(Bid.id.in_(bid_ids)).all()
            floor_map = {b.id: float(b.min_bid_rate) if b.min_bid_rate else None for b in bids}
        for r in lost_records:
            diff = float(r.rate_diff) if r.rate_diff is not None else (
                float(r.submitted_rate) - float(r.actual_winner_rate)
                if r.submitted_rate and r.actual_winner_rate else None
            )
            if diff is None:
                continue
            if diff > 0:
                above_winner += 1
            else:
                floor_rate = floor_map.get(r.bid_id) if r.bid_id else None
                if floor_rate and r.submitted_rate and float(r.submitted_rate) < floor_rate:
                    below_floor += 1
                else:
                    below_winner += 1
        return {"above_winner": above_winner, "below_floor": below_floor, "below_winner": below_winner}


class NotificationService:
    def __init__(self, db: Session):
        self.db = db

    # ── 내부: dedup_key 기반 중복 체크 후 INSERT ────────────────────────
    def _create_deduped(
        self,
        user_id: Optional[int],
        ntype: str,
        title: str,
        body: Optional[str] = None,
        link: Optional[str] = None,
        dedup_key: Optional[str] = None,
    ) -> tuple["Notification", bool]:
        """(notification, is_new) 반환. dedup_key가 오늘 이미 존재하면 기존 반환."""
        if dedup_key:
            q = self.db.query(Notification).filter(
                Notification.dedup_key == dedup_key,
            )
            if user_id is not None:
                q = q.filter(Notification.user_id == user_id)
            else:
                q = q.filter(Notification.user_id.is_(None))
            existing = q.first()
            if existing:
                return existing, False

        n = Notification(
            user_id=user_id, ntype=ntype, title=title,
            body=body, link=link, dedup_key=dedup_key,
        )
        self.db.add(n)
        self.db.flush()   # id 채번 (commit은 호출자가)
        return n, True

    def create(
        self,
        user_id: Optional[int],
        ntype: str,
        title: str,
        body: Optional[str] = None,
        link: Optional[str] = None,
        dedup_key: Optional[str] = None,
    ) -> "Notification":
        n, _ = self._create_deduped(user_id, ntype, title, body, link, dedup_key)
        self.db.commit()
        self.db.refresh(n)
        return n

    def list_for_user(
        self, user_id: int, unread_only: bool = False, limit: int = 20
    ) -> list:
        q = self.db.query(Notification).filter(
            or_(Notification.user_id == user_id, Notification.user_id.is_(None))
        )
        if unread_only:
            q = q.filter(Notification.is_read == False)
        return q.order_by(Notification.created_at.desc()).limit(limit).all()

    def mark_read(self, notification_id: int, user_id: int) -> None:
        n = self.db.query(Notification).filter(
            Notification.id == notification_id,
            or_(Notification.user_id == user_id, Notification.user_id.is_(None)),
        ).first()
        if n:
            n.is_read = True
            self.db.commit()

    def mark_all_read(self, user_id: int) -> None:
        self.db.query(Notification).filter(
            or_(Notification.user_id == user_id, Notification.user_id.is_(None)),
            Notification.is_read == False,
        ).update({"is_read": True}, synchronize_session=False)
        self.db.commit()

    def unread_count(self, user_id: int) -> int:
        return self.db.query(Notification).filter(
            or_(Notification.user_id == user_id, Notification.user_id.is_(None)),
            Notification.is_read == False,
        ).count()

    # ── 키워드 매칭: bid_id당 사용자당 1건 ─────────────────────────────
    def create_keyword_match(self, bid, matched_keywords: list) -> list:
        kw_str = ", ".join(matched_keywords)
        title = f"[키워드 매칭] {bid.title[:60]}"
        body = f"키워드 '{kw_str}' 에 매칭된 공고입니다."
        link = f"/bids/{bid.id}"
        users = self.db.query(User).filter(User.is_active == True).all()
        results = []
        for u in users:
            dedup_key = f"keyword_match:{bid.id}"
            n, is_new = self._create_deduped(
                u.id, "keyword_match", title, body, link, dedup_key=dedup_key
            )
            if is_new:
                results.append(n)
        self.db.commit()
        for n in results:
            self.db.refresh(n)
        return results

    # ── 사정율 급변: 기관명+날짜 기준 하루 1건 ─────────────────────────
    def create_srate_spike(
        self,
        agency_name: str,
        industry_name: str,
        direction: str,
        delta_pct: float,
    ) -> list:
        from datetime import date
        arrow = "▲" if direction == "up" else "▼"
        title = f"[사정율 급변] {agency_name} {arrow}{abs(delta_pct):.1f}%"
        body = f"{industry_name} 공종 사정율이 {arrow}{abs(delta_pct):.1f}% 변동했습니다."
        today_str = date.today().strftime("%Y%m%d")
        safe_agency = agency_name.replace(":", "_")[:80]
        dedup_key = f"srate_spike:{safe_agency}:{today_str}"
        n, is_new = self._create_deduped(
            None, "srate_spike", title, body, dedup_key=dedup_key
        )
        self.db.commit()
        if is_new:
            self.db.refresh(n)
        return [n]

    # ── 투찰 마감 알림: execution_id + days_left + 날짜 기준 1건 ────────
    def create_execution_deadline(
        self, user_id: int, exec_title: str, days_left: int,
        execution_id: Optional[int] = None,
    ) -> "Notification":
        from datetime import date
        if days_left == 0:
            title = f"[오늘 개찰] {exec_title[:45]}"
            body = "오늘 개찰 마감입니다. 투찰 완료 여부를 확인하세요."
        else:
            title = f"[D-{days_left}] 내일 개찰: {exec_title[:40]}"
            body = f"{days_left}일 후 개찰 마감입니다. 투찰률을 최종 확인하세요."
        today_str = date.today().strftime("%Y%m%d")
        dedup_key = f"exec_deadline:{execution_id}:D{days_left}:{today_str}" if execution_id else None
        n, is_new = self._create_deduped(
            user_id, "execution_deadline", title, body,
            link="/executions", dedup_key=dedup_key,
        )
        self.db.commit()
        if is_new:
            self.db.refresh(n)
        return n

    # ── 결과 입력 리마인더: execution_id + 날짜 기준 하루 1건 ────────────
    def create_result_reminder(
        self, user_id: int, exec_title: str,
        execution_id: Optional[int] = None,
    ) -> "Notification":
        from datetime import date
        title = f"[결과 입력 요청] {exec_title[:45]}"
        body = "개찰이 완료된 것으로 보입니다. 낙찰/패찰 결과를 입력해주세요."
        today_str = date.today().strftime("%Y%m%d")
        dedup_key = f"result_reminder:{execution_id}:{today_str}" if execution_id else None
        n, is_new = self._create_deduped(
            user_id, "execution_result", title, body,
            link="/executions", dedup_key=dedup_key,
        )
        self.db.commit()
        if is_new:
            self.db.refresh(n)
        return n


# ==================================================
# 전참여자 조회 서비스 (inpo21c_participants)
# ==================================================

class InpoParticipantService:
    """inpo21c_participants → 공고별 전체 참여자 목록 반환."""

    def get(self, db: Session, bid_id: int) -> list:
        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid or not bid.announcement_no:
            return []
        rows = db.execute(text("""
            SELECT ip.rank, ip.company_name, ip.biz_reg_no,
                   ip.bid_rate, ip.base_ratio, ip.is_winner
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ib.announcement_no LIKE :ano
            ORDER BY ip.rank
        """), {"ano": bid.announcement_no + "%"}).fetchall()
        return [
            {
                "rank":         r[0],
                "company_name": r[1],
                "biz_reg_no":   r[2],
                "bid_rate":     float(r[3]) if r[3] is not None else None,
                "base_ratio":   float(r[4]) if r[4] is not None else None,
                "is_winner":    bool(r[5]),
            }
            for r in rows
        ]


# ==================================================
# 惜敗 분석 서비스 — my_bid_records × inpo21c_participants
# ==================================================

class SekihaiService:
    """자사 투찰이력 × inpo21c 실측 순위 교차 분석."""

    def get_rank(self, db: Session, announcement_no: str) -> dict:
        """announcement_no 기준 inpo21c 참여자 목록에서 모든 순위 반환."""
        rows = db.execute(text("""
            SELECT ip.rank, ip.company_name, ip.bid_rate, ip.is_winner,
                   ib.base_amount, ib.estimated_amount
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ib.announcement_no LIKE :ano
            ORDER BY ip.rank
        """), {"ano": announcement_no + "%"}).fetchall()
        if not rows:
            return {"found": False, "participants": []}
        winner = next((r for r in rows if r[3]), None)
        return {
            "found": True,
            "total_count": len(rows),
            "winner_rate": float(winner[2]) if winner and winner[2] else None,
            "base_amount": int(rows[0][4]) if rows[0][4] else None,
            "participants": [
                {
                    "rank":         r[0],
                    "company_name": r[1],
                    "bid_rate":     float(r[2]) if r[2] is not None else None,
                    "is_winner":    bool(r[3]),
                }
                for r in rows[:20]
            ],
        }

    def batch_ranks(self, db: Session, announcement_nos: list[str]) -> dict:
        """여러 announcement_no에 대한 순위 정보 일괄 반환."""
        result: dict = {}
        for ano in announcement_nos:
            if not ano:
                continue
            rows = db.execute(text("""
                SELECT ip.rank, ip.company_name, ip.bid_rate, ip.is_winner
                FROM inpo21c_participants ip
                JOIN inpo21c_bids ib USING (inpo21c_bid_id)
                WHERE ib.announcement_no LIKE :ano
                ORDER BY ip.rank
            """), {"ano": ano + "%"}).fetchall()
            if rows:
                winner = next((r for r in rows if r[3]), None)
                result[ano] = {
                    "found": True,
                    "total_count": len(rows),
                    "winner_rate": float(winner[2]) if winner and winner[2] else None,
                    "top5": [
                        {"rank": r[0], "company_name": r[1], "bid_rate": float(r[2]) if r[2] else None, "is_winner": bool(r[3])}
                        for r in rows[:5]
                    ],
                }
            else:
                result[ano] = {"found": False}
        return result


# ==================================================
# 경쟁 레이더 서비스 — 동반 입찰 경쟁사 분석
# ==================================================

class RivalRadarService:
    """inpo21c_participants 기반 동반 입찰 경쟁사 레이더."""

    def get(self, db: Session, bid_id: int, top_k: int = 15) -> dict:
        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid or not bid.announcement_no:
            return {"rivals": [], "bid_id": bid_id}

        # 이 공고에 참여한 업체들
        participants = db.execute(text("""
            SELECT ip.company_name, ip.biz_reg_no, ip.bid_rate, ip.rank, ip.is_winner
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ib.announcement_no LIKE :ano
            ORDER BY ip.rank
        """), {"ano": bid.announcement_no + "%"}).fetchall()

        if not participants:
            return {"rivals": [], "bid_id": bid_id, "announcement_no": bid.announcement_no}

        # 같은 업체들이 함께 참여한 다른 공고 수 (동반 빈도)
        biz_nos = [r[1] for r in participants if r[1]]
        if not biz_nos:
            company_names = [r[0] for r in participants if r[0]]
            freq_rows = db.execute(text("""
                SELECT ip2.company_name, COUNT(DISTINCT ib2.inpo21c_bid_id) as co_count,
                       AVG(ip2.bid_rate::numeric) as avg_rate,
                       SUM(CASE WHEN ip2.is_winner THEN 1 ELSE 0 END) as win_count
                FROM inpo21c_participants ip2
                JOIN inpo21c_bids ib2 USING (inpo21c_bid_id)
                WHERE ip2.company_name = ANY(:names)
                  AND ib2.announcement_no NOT LIKE :ano
                  AND ib2.agency_name = :agency
                GROUP BY ip2.company_name
                ORDER BY co_count DESC
                LIMIT :k
            """), {
                "names": company_names,
                "ano": bid.announcement_no + "%",
                "agency": bid.agency.name if bid.agency else "",
                "k": top_k,
            }).fetchall()
        else:
            freq_rows = db.execute(text("""
                SELECT ip2.company_name, COUNT(DISTINCT ib2.inpo21c_bid_id) as co_count,
                       AVG(ip2.bid_rate::numeric) as avg_rate,
                       SUM(CASE WHEN ip2.is_winner THEN 1 ELSE 0 END) as win_count
                FROM inpo21c_participants ip2
                JOIN inpo21c_bids ib2 USING (inpo21c_bid_id)
                WHERE ip2.biz_reg_no = ANY(:nos)
                  AND ib2.announcement_no NOT LIKE :ano
                GROUP BY ip2.company_name
                ORDER BY co_count DESC
                LIMIT :k
            """), {
                "nos": biz_nos,
                "ano": bid.announcement_no + "%",
                "k": top_k,
            }).fetchall()

        rivals = [
            {
                "company_name": r[0],
                "co_bid_count": int(r[1]),
                "avg_bid_rate": float(r[2]) if r[2] else None,
                "win_count":    int(r[3]),
            }
            for r in freq_rows
        ]

        winner = next((r for r in participants if r[4]), None)
        return {
            "bid_id": bid_id,
            "announcement_no": bid.announcement_no,
            "total_participants": len(participants),
            "winner_company": winner[0] if winner else None,
            "winner_rate":    float(winner[2]) if winner and winner[2] else None,
            "rivals": rivals,
            "current_participants": [
                {"rank": r[3], "company_name": r[0], "bid_rate": float(r[2]) if r[2] else None, "is_winner": bool(r[4])}
                for r in participants[:20]
            ],
        }


# ==================================================
# 실측 낙찰 구간 서비스 — inpo21c 기반 실측 분포
# ==================================================

class ActualWinZoneService:
    """inpo21c_participants의 낙찰자 투찰률 실측 분포 산출."""

    def get(self, db: Session, bid_id: int) -> dict:
        bid = db.query(Bid).filter(Bid.id == bid_id).first()
        if not bid:
            return {"zones": [], "sample_count": 0}

        # 유사 공고(같은 발주처 or 공종) 낙찰율 분포
        agency_name = bid.agency.name if bid.agency else None
        industry_id = bid.industry_id

        rows = db.execute(text("""
            SELECT ip.bid_rate::numeric
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            JOIN agencies a ON a.name = ib.agency_name
            WHERE ip.is_winner = TRUE
              AND (:agency_name IS NULL OR ib.agency_name = :agency_name)
            ORDER BY ip.bid_rate
            LIMIT 500
        """), {"agency_name": agency_name}).fetchall()

        rates = [float(r[0]) for r in rows if r[0] is not None]
        if not rates:
            return {"zones": [], "sample_count": 0, "message": "충분한 실측 데이터 없음"}

        # 0.2%p 구간으로 빈도 집계
        import math
        step = 0.002
        lo = math.floor(min(rates) / step) * step
        hi = math.ceil(max(rates) / step) * step
        buckets: dict = {}
        for rate in rates:
            key = round(math.floor(rate / step) * step, 4)
            buckets[key] = buckets.get(key, 0) + 1

        total = len(rates)
        zones = [
            {
                "range_lo":   k,
                "range_hi":   round(k + step, 4),
                "count":      v,
                "probability": round(v / total * 100, 1),
            }
            for k, v in sorted(buckets.items())
        ]

        mean_rate = round(sum(rates) / total, 5)
        peak_zone = max(zones, key=lambda z: z["count"]) if zones else None

        return {
            "sample_count": total,
            "mean_winner_rate": mean_rate,
            "peak_zone":    peak_zone,
            "zones":        zones,
            "agency_name":  agency_name,
        }


# ==================================================
# 시장 인텔리전스 서비스
# ==================================================

class MarketIntelService:
    """inpo21c_participants 기반 시장 인텔리전스 — 발주처×사정율 히트맵."""

    def agency_heatmap(self, db: Session, months: int = 12, top_n: int = 20) -> dict:
        """발주처별 낙찰율 분포 히트맵 데이터."""
        months_safe = max(1, min(int(months), 60))
        top_n_safe  = max(5, min(int(top_n), 100))
        rows = db.execute(text(f"""
            SELECT ib.agency_name,
                   COUNT(DISTINCT ib.inpo21c_bid_id)          AS bid_count,
                   AVG(ip.bid_rate::numeric)                  AS avg_winner_rate,
                   PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY ip.bid_rate::numeric) AS p25,
                   PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY ip.bid_rate::numeric) AS p75,
                   MIN(ip.bid_rate::numeric)                  AS min_rate,
                   MAX(ip.bid_rate::numeric)                  AS max_rate
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ip.is_winner = TRUE
              AND (ib.open_datetime IS NULL OR ib.open_datetime >= NOW() - INTERVAL '{months_safe} months')
            GROUP BY ib.agency_name
            HAVING COUNT(*) >= 3
            ORDER BY bid_count DESC
            LIMIT {top_n_safe}
        """)).fetchall()

        return {
            "months":  months,
            "agencies": [
                {
                    "agency_name":    r[0],
                    "bid_count":      int(r[1]),
                    "avg_rate":       float(r[2]) if r[2] else None,
                    "p25":            float(r[3]) if r[3] else None,
                    "p75":            float(r[4]) if r[4] else None,
                    "min_rate":       float(r[5]) if r[5] else None,
                    "max_rate":       float(r[6]) if r[6] else None,
                }
                for r in rows
            ],
        }

    def winner_rate_trend(self, db: Session, agency_name: Optional[str] = None) -> dict:
        """월별 낙찰율 추세 (open_datetime 우선, 없으면 created_at 사용)."""
        rows = db.execute(text("""
            SELECT EXTRACT(YEAR  FROM COALESCE(ib.open_datetime, ib.created_at))  AS yr,
                   EXTRACT(MONTH FROM COALESCE(ib.open_datetime, ib.created_at))  AS mo,
                   COUNT(DISTINCT ib.inpo21c_bid_id)    AS bid_count,
                   AVG(ip.bid_rate::numeric)             AS avg_rate
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ip.is_winner = TRUE
              AND (:agency IS NULL OR ib.agency_name = :agency)
            GROUP BY yr, mo
            ORDER BY yr, mo
            LIMIT 36
        """), {"agency": agency_name}).fetchall()
        return {
            "agency_name": agency_name,
            "trend": [
                {
                    "year":       int(r[0]),
                    "month":      int(r[1]),
                    "bid_count":  int(r[2]),
                    "avg_rate":   float(r[3]) if r[3] else None,
                }
                for r in rows
            ],
        }

    def top_winner_companies(self, db: Session, agency_name: Optional[str] = None, top_n: int = 10) -> list:
        """낙찰 다발 업체 순위."""
        rows = db.execute(text("""
            SELECT ip.company_name,
                   COUNT(*) as win_count,
                   AVG(ip.bid_rate::numeric) as avg_rate,
                   MIN(ip.bid_rate::numeric) as min_rate,
                   MAX(ip.bid_rate::numeric) as max_rate
            FROM inpo21c_participants ip
            JOIN inpo21c_bids ib USING (inpo21c_bid_id)
            WHERE ip.is_winner = TRUE
              AND (:agency IS NULL OR ib.agency_name = :agency)
            GROUP BY ip.company_name
            ORDER BY win_count DESC
            LIMIT :n
        """), {"agency": agency_name, "n": top_n}).fetchall()
        return [
            {
                "company_name": r[0],
                "win_count":    int(r[1]),
                "avg_rate":     float(r[2]) if r[2] else None,
                "min_rate":     float(r[3]) if r[3] else None,
                "max_rate":     float(r[4]) if r[4] else None,
            }
            for r in rows
        ]


# ==================================================
# 투찰 실행 관리 서비스
# ==================================================

class ExecutionService:
    """투찰 수명주기 관리 + SUCVIEW/인포 엑셀 파싱"""

    STATUS_ORDER = ["검토중", "참여결정", "투찰완료", "개찰대기", "낙찰", "패찰", "포기"]

    def __init__(self, db: Session):
        self.db = db

    # ── CRUD ─────────────────────────────────────────────────

    def list_executions(self, user_id: int, status: str = None, page: int = 1, size: int = 20) -> dict:
        from .schemas import BidExecutionOut
        q = self.db.query(BidExecution).filter(BidExecution.user_id == user_id)
        if status:
            q = q.filter(BidExecution.status == status)
        total = q.count()
        items = q.order_by(BidExecution.created_at.desc()).offset((page - 1) * size).limit(size).all()
        return {
            "total": total, "page": page, "size": size,
            "items": [BidExecutionOut.model_validate(i).model_dump() for i in items],
        }

    def get_summary(self, user_id: int) -> dict:
        from sqlalchemy import func as sqlfunc
        rows = (
            self.db.query(BidExecution.status, sqlfunc.count().label("cnt"))
            .filter(BidExecution.user_id == user_id)
            .group_by(BidExecution.status)
            .all()
        )
        summary = {s: 0 for s in self.STATUS_ORDER}
        for status, cnt in rows:
            summary[status] = cnt

        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start.replace(hour=23, minute=59, second=59)
        today_closing = (
            self.db.query(BidExecution)
            .filter(
                BidExecution.user_id == user_id,
                BidExecution.status.in_(["참여결정", "투찰완료"]),
                BidExecution.bid_open_date >= today_start,
                BidExecution.bid_open_date <= today_end,
            )
            .all()
        )
        from .schemas import BidExecutionOut
        return {
            "status_counts": summary,
            "today_closing": [BidExecutionOut.model_validate(e).model_dump() for e in today_closing],
        }

    def get(self, exec_id: int):
        return self.db.query(BidExecution).filter(BidExecution.id == exec_id).first()

    def create(self, user_id: int, data) -> BidExecution:
        obj = BidExecution(user_id=user_id, **data.model_dump(exclude_none=True))
        self.db.add(obj)
        self.db.commit()
        self.db.refresh(obj)
        return obj

    def update(self, exec_id: int, user_id: int, data) -> BidExecution:
        obj = self.db.query(BidExecution).filter(
            BidExecution.id == exec_id,
            BidExecution.user_id == user_id,
        ).first()
        if not obj:
            from fastapi import HTTPException
            raise HTTPException(404, "Not found")
        updates = data.model_dump(exclude_none=True)
        for k, v in updates.items():
            setattr(obj, k, v)
        # 패찰 → 자동 원인 분석 + 알림
        if data.status == "패찰":
            self._auto_defeat_analysis(obj)
            da = self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == obj.id).first()
            cause_str = da.cause_primary if da else "기타"
            improvement_str = da.improvement if da else ""
            nsvc = NotificationService(self.db)
            nsvc.create(
                user_id=obj.user_id,
                ntype="execution_result",
                title=f"[패찰] {obj.title[:45]}",
                body=f"원인: {cause_str}. {improvement_str}",
                link="/executions",
            )
        # 낙찰 → 축하 알림
        elif data.status == "낙찰":
            nsvc = NotificationService(self.db)
            nsvc.create(
                user_id=obj.user_id,
                ntype="execution_result",
                title=f"[낙찰] {obj.title[:45]}",
                body=f"축하합니다! 낙찰 확정되었습니다. 투찰률: {obj.submitted_rate:.3%}" if obj.submitted_rate else "낙찰 확정",
                link="/executions",
            )
        self.db.commit()
        self.db.refresh(obj)
        return obj

    def delete(self, exec_id: int, user_id: int):
        obj = self.db.query(BidExecution).filter(
            BidExecution.id == exec_id,
            BidExecution.user_id == user_id,
        ).first()
        if obj:
            # defeat_analyses has NOT NULL FK — delete child rows first
            self.db.query(DefeatAnalysis).filter(
                DefeatAnalysis.execution_id == exec_id
            ).delete(synchronize_session=False)
            self.db.delete(obj)
            self.db.commit()

    # ── 패찰 원인 분석 ────────────────────────────────────────

    def get_defeat_analysis(self, exec_id: int):
        return self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == exec_id).first()

    def _auto_defeat_analysis(self, obj: BidExecution):
        existing = self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == obj.id).first()
        if existing:
            return

        cause, detail, adj = "기타", "", 0.0
        winner_gap_pct = None

        if obj.submitted_rate and obj.winner_rate:
            gap = float(obj.submitted_rate) - float(obj.winner_rate)
            winner_gap_pct = round(gap * 100, 3)
            if gap > 0.005:
                cause = "투찰률과도"
                detail = f"투찰률이 낙찰률보다 {gap*100:.3f}%p 높았음"
                adj = round(-gap * 0.6, 4)
            elif gap > 0.001:
                cause = "투찰률과도"
                detail = f"투찰률이 낙찰률보다 {gap*100:.3f}%p 높았음 (미세)"
                adj = round(-gap * 0.5, 4)

        if cause == "기타" and obj.total_bidders and obj.total_bidders >= 15:
            cause = "경쟁사과다"
            detail = f"참여업체 {obj.total_bidders}개사 — 경쟁 과열"

        improvement = "다음 동일기관 입찰 시 " + (
            f"{abs(adj)*100:.2f}%p 낮게 조정 권장" if adj < 0 else "현 전략 유지"
        )

        da = DefeatAnalysis(
            execution_id=obj.id,
            cause_primary=cause,
            cause_detail=detail,
            winner_gap_pct=winner_gap_pct,
            competitor_cnt=obj.total_bidders,
            our_rank=obj.result_rank,
            floor_rate=obj.floor_rate,
            improvement=improvement,
            next_rate_adj=adj,
        )
        self.db.add(da)

    # ── SUCVIEW 엑셀 파싱 ─────────────────────────────────────

    def import_sucview(self, file_bytes: bytes, user_id: int):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        imported, skipped, competitors_added = 0, 0, 0
        errors, details = [], []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                result = self._parse_sucview_sheet(ws, user_id)
                imported += result["imported"]
                skipped += result["skipped"]
                competitors_added += result["competitors_added"]
                details.extend(result["details"])
            except Exception as e:
                errors.append(f"시트 {sheet_name}: {e}")

        self.db.commit()
        return {
            "imported": imported,
            "skipped": skipped,
            "competitors_added": competitors_added,
            "errors": errors,
            "details": details,
        }

    def _parse_sucview_sheet(self, ws, user_id: int) -> dict:
        data = {}
        participants = []
        in_participant = False

        for row in ws.iter_rows(values_only=True):
            if not any(c for c in row if c is not None):
                continue
            label = str(row[0] or "").strip()
            val = row[1] if len(row) > 1 else None

            if label == "공고번호":
                data["announcement_no"] = str(val or "").strip()
            elif label == "공고명":
                data["title"] = str(val or "").strip()
            elif label == "발주기관":
                data["agency_name"] = str(val or "").strip()
            elif label in ("낙찰하한율", "사정율하한"):
                try:
                    data["floor_rate"] = float(val)
                except Exception:
                    pass
            elif label in ("기초금액", "예정가격"):
                try:
                    data["base_amount"] = int(float(str(val or "0").replace(",", "")))
                except Exception:
                    pass
            elif label == "개찰일시":
                if val:
                    try:
                        data["opened_at"] = val if isinstance(val, datetime) else datetime.fromisoformat(str(val))
                    except Exception:
                        pass
            elif label in ("순위", "No", "번호") or (str(label).isdigit()):
                in_participant = True

            if in_participant and len(row) >= 5:
                try:
                    rank = int(row[0]) if row[0] else None
                    biz_no = str(row[1] or "").strip()
                    name = str(row[2] or "").strip()
                    amount_raw = str(row[3] or "0").replace(",", "")
                    rate_raw = row[4]
                    if rank and name and len(name) > 1:
                        participants.append({
                            "rank": rank,
                            "biz_no": biz_no,
                            "company_name": name,
                            "amount": int(float(amount_raw)) if amount_raw else None,
                            "rate": float(rate_raw) if rate_raw else None,
                        })
                except Exception:
                    pass

        if not data.get("title"):
            return {"imported": 0, "skipped": 1, "competitors_added": 0, "details": []}

        # 낙찰자 추출
        winner = next((p for p in participants if p["rank"] == 1), None)
        if winner:
            data["winner_name"] = winner["company_name"]
            data["winner_biz_no"] = winner["biz_no"]
            data["winner_amount"] = winner["amount"]
            data["winner_rate"] = winner["rate"]
            data["total_bidders"] = len(participants)

        # 중복 체크
        existing = self.db.query(BidExecution).filter(
            BidExecution.announcement_no == data.get("announcement_no"),
            BidExecution.user_id == user_id,
        ).first()
        if existing:
            return {"imported": 0, "skipped": 1, "competitors_added": 0, "details": [f"중복: {data.get('title', '')}"][:1]}

        data["status"] = "개찰대기" if not winner else "패찰"
        data["source"] = "sucview"
        data["sucview_raw"] = participants
        data["user_id"] = user_id

        exec_obj = BidExecution(**data)
        self.db.add(exec_obj)
        self.db.flush()

        comp_added = self._update_our_competitors_raw(participants)
        detail_msg = f"가져옴: {data.get('title', '')[:40]}"
        return {"imported": 1, "skipped": 0, "competitors_added": comp_added, "details": [detail_msg]}

    def _update_our_competitors_raw(self, participants: list) -> int:
        added = 0
        for p in participants:
            if not p.get("company_name"):
                continue
            existing = self.db.query(OurCompetitor).filter(
                OurCompetitor.company_name == p["company_name"]
            ).first()
            if existing:
                existing.co_participation_cnt = (existing.co_participation_cnt or 0) + 1
                if p.get("rank") == 1:
                    existing.co_win_cnt = (existing.co_win_cnt or 0) + 1
                if p.get("rate"):
                    existing.avg_bid_rate = p["rate"]
                existing.last_seen_at = datetime.now().date()
                existing.updated_at = datetime.now()
            else:
                new_comp = OurCompetitor(
                    company_name=p["company_name"],
                    biz_reg_no=p.get("biz_no"),
                    co_participation_cnt=1,
                    co_win_cnt=1 if p.get("rank") == 1 else 0,
                    avg_bid_rate=p.get("rate"),
                    last_seen_at=datetime.now().date(),
                    source="sucview",
                )
                self.db.add(new_comp)
                added += 1
        return added

    def import_inpo_history(self, file_bytes: bytes, user_id: int):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        imported, skipped = 0, 0
        errors, details = [], []

        headers = []
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [str(c or "").strip() for c in row]
                continue
            row_dict = dict(zip(headers, row))
            try:
                announcement_no = str(row_dict.get("공고번호") or "").strip()
                title = str(row_dict.get("공고명") or "").strip()
                if not title:
                    skipped += 1
                    continue

                existing = self.db.query(BidExecution).filter(
                    BidExecution.announcement_no == announcement_no,
                    BidExecution.user_id == user_id,
                ).first() if announcement_no else None

                if existing:
                    skipped += 1
                    continue

                bid_rate_raw = row_dict.get("투찰율") or row_dict.get("투찰률")
                bid_rate = None
                try:
                    bid_rate = float(str(bid_rate_raw).replace("%", "")) / 100 if bid_rate_raw else None
                except Exception:
                    pass

                result_raw = str(row_dict.get("결과") or row_dict.get("낙패") or "").strip()
                status = "낙찰" if "낙찰" in result_raw else ("패찰" if "패찰" in result_raw else "개찰대기")

                base_raw = str(row_dict.get("기초금액") or row_dict.get("예정가격") or "0").replace(",", "")
                try:
                    base_amount = int(float(base_raw))
                except Exception:
                    base_amount = 0

                exec_obj = BidExecution(
                    user_id=user_id,
                    announcement_no=announcement_no,
                    title=title,
                    agency_name=str(row_dict.get("발주기관") or ""),
                    base_amount=base_amount,
                    submitted_rate=bid_rate,
                    status=status,
                    source="inpo_history",
                )
                self.db.add(exec_obj)
                imported += 1
                details.append(f"가져옴: {title[:40]}")
            except Exception as e:
                errors.append(f"행 오류: {e}")
                skipped += 1

        self.db.commit()
        return {
            "imported": imported,
            "skipped": skipped,
            "competitors_added": 0,
            "errors": errors,
            "details": details,
        }


# ==================================================
# 발주기관 빈도표 서비스
# ==================================================

class FrequencyService:
    """발주기관별 낙찰률 빈도표 + 히스토그램 (bid_results 기반)"""

    PERIOD_MAP = {"6M": 6, "12M": 12, "24M": 24, "48M": 48}
    BUCKET_WIDTH = 0.005

    def __init__(self, db: Session):
        self.db = db

    def get_agency_freq(self, agency_id: int, industry_code: str = "ALL", period: str = "48M") -> dict:
        months = self.PERIOD_MAP.get(period, 48)
        rows = self.db.query(RateFrequencyTable).filter(
            RateFrequencyTable.agency_id == agency_id,
            RateFrequencyTable.industry_code == industry_code,
            RateFrequencyTable.period_type == period,
        ).order_by(RateFrequencyTable.bucket_from).all()

        if not rows:
            self._build_agency_freq(agency_id, industry_code, months, period)
            self.db.commit()
            rows = self.db.query(RateFrequencyTable).filter(
                RateFrequencyTable.agency_id == agency_id,
                RateFrequencyTable.industry_code == industry_code,
                RateFrequencyTable.period_type == period,
            ).order_by(RateFrequencyTable.bucket_from).all()

        agency = self.db.query(Agency).filter(Agency.id == agency_id).first()
        buckets = [
            {
                "from": float(r.bucket_from),
                "to": float(r.bucket_to),
                "count": r.count,
                "win_count": r.win_count,
                "win_rate": float(r.win_rate) if r.win_rate else 0,
            }
            for r in rows
        ]
        total = sum(b["count"] for b in buckets)
        return {
            "agency_id": agency_id,
            "agency_name": agency.name if agency else "",
            "industry_code": industry_code,
            "period": period,
            "total_bids": total,
            "buckets": buckets,
        }

    def rebuild_all(self) -> dict:
        agencies = self.db.execute(text(
            "SELECT DISTINCT b.agency_id FROM bid_results r JOIN bids b ON b.id = r.bid_id LIMIT 500"
        )).fetchall()
        built = 0
        for (agency_id,) in agencies:
            for period_label, months in self.PERIOD_MAP.items():
                try:
                    self._build_agency_freq(agency_id, "ALL", months, period_label)
                    built += 1
                except Exception:
                    pass
        self.db.commit()
        return {"built": built, "agencies": len(agencies)}

    def _build_agency_freq(self, agency_id: int, industry_code: str, months: int, period_label: str):
        cutoff = f"NOW() - INTERVAL '{months} months'"
        sql = text(f"""
            SELECT
                r.assessment_rate,
                CASE WHEN r.is_winner THEN 1 ELSE 0 END AS is_win
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :agency_id
              AND r.assessment_rate IS NOT NULL
              AND b.bid_open_date >= {cutoff}
        """)
        rows = self.db.execute(sql, {"agency_id": agency_id}).fetchall()
        if not rows:
            return

        # 기존 행 삭제
        self.db.query(RateFrequencyTable).filter(
            RateFrequencyTable.agency_id == agency_id,
            RateFrequencyTable.industry_code == industry_code,
            RateFrequencyTable.period_type == period_label,
        ).delete()

        buckets: dict[float, dict] = {}
        for rate, is_win in rows:
            rate_f = float(rate)
            bucket_from = round(int(rate_f / self.BUCKET_WIDTH) * self.BUCKET_WIDTH, 6)
            bucket_to = round(bucket_from + self.BUCKET_WIDTH, 6)
            if bucket_from not in buckets:
                buckets[bucket_from] = {"to": bucket_to, "count": 0, "win": 0}
            buckets[bucket_from]["count"] += 1
            buckets[bucket_from]["win"] += is_win

        for bf, bdata in buckets.items():
            win_rate = round(bdata["win"] / bdata["count"], 4) if bdata["count"] > 0 else None
            self.db.add(RateFrequencyTable(
                agency_id=agency_id,
                industry_code=industry_code,
                period_type=period_label,
                bucket_from=bf,
                bucket_to=bdata["to"],
                bucket_width=self.BUCKET_WIDTH,
                count=bdata["count"],
                win_count=bdata["win"],
                win_rate=win_rate,
            ))


# ==================================================
# 발주기관 전략 DB 서비스 (AgencyStrategy 집계)
# ==================================================

class AgencyStrategyService:
    """발주기관별 48개월 낙찰률 통계 + 빈도표 + 히스토그램 (agency_strategies 테이블)"""

    BUCKET_W = 0.005

    def __init__(self, db: Session):
        self.db = db

    def get(self, agency_id: int, industry_code: str = "ALL", period_months: int = 48) -> dict | None:
        row = self.db.query(AgencyStrategy).filter(
            AgencyStrategy.agency_id == agency_id,
            AgencyStrategy.industry_code == industry_code,
            AgencyStrategy.period_months == period_months,
        ).first()
        if not row:
            return None
        return self._serialize(row)

    def get_or_build(self, agency_id: int, industry_code: str = "ALL", period_months: int = 48) -> dict:
        row = self.db.query(AgencyStrategy).filter(
            AgencyStrategy.agency_id == agency_id,
            AgencyStrategy.industry_code == industry_code,
            AgencyStrategy.period_months == period_months,
        ).first()
        if not row:
            self._build_one(agency_id, industry_code, period_months)
            self.db.commit()
            row = self.db.query(AgencyStrategy).filter(
                AgencyStrategy.agency_id == agency_id,
                AgencyStrategy.industry_code == industry_code,
                AgencyStrategy.period_months == period_months,
            ).first()
        if not row:
            return {"agency_id": agency_id, "total_bid_count": 0, "message": "데이터 부족"}
        return self._serialize(row)

    def rebuild_all(self) -> dict:
        """전체 기관 strategy 재계산 (데이터 5건 이상 기관만)"""
        agencies = self.db.execute(text(
            "SELECT DISTINCT b.agency_id FROM bids b "
            "JOIN bid_results r ON r.bid_id = b.id "
            "WHERE b.bid_open_date >= NOW() - INTERVAL '48 months' "
            "GROUP BY b.agency_id HAVING COUNT(CASE WHEN r.is_winner THEN 1 END) >= 5"
        )).fetchall()
        built, skipped = 0, 0
        for (agency_id,) in agencies:
            try:
                self._build_one(agency_id, "ALL", 48)
                built += 1
            except Exception:
                skipped += 1
        self.db.commit()
        return {"built": built, "skipped": skipped, "agencies": len(agencies)}

    def _build_one(self, agency_id: int, industry_code: str, period_months: int):
        from datetime import datetime, timedelta

        cutoff = datetime.utcnow() - timedelta(days=30 * period_months)

        # 낙찰률 (winner rows)
        winner_rows = self.db.execute(text("""
            SELECT r.bid_rate::float
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :aid
              AND r.is_winner = true
              AND r.bid_rate IS NOT NULL
              AND b.bid_open_date >= :cutoff
        """), {"aid": agency_id, "cutoff": cutoff}).fetchall()

        rates = [row[0] for row in winner_rows]
        n = len(rates)
        if n < 5:
            return

        arr = np.array(rates)

        # 건당 평균 경쟁업체 수
        avg_comp = self.db.execute(text("""
            SELECT COALESCE(AVG(cnt), 0)::float FROM (
                SELECT COUNT(*) AS cnt
                FROM bid_results r
                JOIN bids b ON b.id = r.bid_id
                WHERE b.agency_id = :aid AND b.bid_open_date >= :cutoff
                GROUP BY r.bid_id
            ) sub
        """), {"aid": agency_id, "cutoff": cutoff}).scalar() or 0.0

        # 공격성 지수: 낙찰하한율 이하 낙찰 비율
        aggression_row = self.db.execute(text("""
            SELECT
                COUNT(CASE WHEN r.bid_rate < b.min_bid_rate THEN 1 END)::float
                / NULLIF(COUNT(*), 0)::float
            FROM bid_results r
            JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :aid
              AND r.is_winner = true
              AND b.min_bid_rate IS NOT NULL
              AND b.bid_open_date >= :cutoff
        """), {"aid": agency_id, "cutoff": cutoff}).scalar()
        aggression_index = float(aggression_row) if aggression_row else 0.0

        # 최근 30일 변동성
        recent_cutoff = datetime.utcnow() - timedelta(days=30)
        recent_rates = [r[0] for r in self.db.execute(text("""
            SELECT r.bid_rate::float
            FROM bid_results r JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :aid AND r.is_winner = true
              AND r.bid_rate IS NOT NULL AND b.bid_open_date >= :rc
        """), {"aid": agency_id, "rc": recent_cutoff}).fetchall()]
        volatility_30d = float(np.std(recent_rates)) if len(recent_rates) >= 3 else None

        # 추세: 최근 12M vs 12M~24M
        cutoff_12m = datetime.utcnow() - timedelta(days=365)
        cutoff_24m = datetime.utcnow() - timedelta(days=730)
        m_recent = self.db.execute(text("""
            SELECT AVG(r.bid_rate)::float FROM bid_results r JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :aid AND r.is_winner = true AND b.bid_open_date >= :c12
        """), {"aid": agency_id, "c12": cutoff_12m}).scalar()
        m_older = self.db.execute(text("""
            SELECT AVG(r.bid_rate)::float FROM bid_results r JOIN bids b ON b.id = r.bid_id
            WHERE b.agency_id = :aid AND r.is_winner = true
              AND b.bid_open_date BETWEEN :c24 AND :c12
        """), {"aid": agency_id, "c24": cutoff_24m, "c12": cutoff_12m}).scalar()
        trend_direction = "stable"
        if m_recent and m_older:
            diff = m_recent - m_older
            trend_direction = "up" if diff > 0.002 else ("down" if diff < -0.002 else "stable")

        # 빈도표 JSON (0.5% 구간)
        freq_table = []
        lo_b = round(int(float(arr.min()) / self.BUCKET_W) * self.BUCKET_W, 6)
        b_ptr = lo_b
        while b_ptr <= float(arr.max()) + 1e-9:
            cnt = int(np.sum((arr >= b_ptr) & (arr < b_ptr + self.BUCKET_W)))
            if cnt > 0:
                freq_table.append({"from": round(b_ptr, 4), "to": round(b_ptr + self.BUCKET_W, 4), "count": cnt})
            b_ptr = round(b_ptr + self.BUCKET_W, 6)

        # 히스토그램 데이터 (0.860~0.965 고정 범위)
        hist_bins = np.arange(0.860, 0.966, self.BUCKET_W)
        hist_counts, _ = np.histogram(arr, bins=hist_bins)
        histogram_data = [
            [round(float(hist_bins[i]), 4), int(hist_counts[i])]
            for i in range(len(hist_counts))
        ]

        # 난이도 분류
        cv = float(np.std(arr) / np.mean(arr)) if np.mean(arr) > 0 else 0
        if avg_comp >= 10 and cv > 0.01:
            qual_difficulty = "難"
        elif avg_comp >= 5 or cv > 0.005:
            qual_difficulty = "中"
        else:
            qual_difficulty = "易"

        p25 = float(np.percentile(arr, 25))
        p75 = float(np.percentile(arr, 75))

        upsert_data = dict(
            total_bid_count=n,
            avg_win_rate=float(np.mean(arr)),
            std_win_rate=float(np.std(arr)),
            min_win_rate=float(arr.min()),
            max_win_rate=float(arr.max()),
            win_rate_p10=float(np.percentile(arr, 10)),
            win_rate_p25=p25,
            win_rate_p50=float(np.percentile(arr, 50)),
            win_rate_p75=p75,
            win_rate_p90=float(np.percentile(arr, 90)),
            avg_competitor_cnt=avg_comp,
            aggression_index=aggression_index,
            qual_difficulty=qual_difficulty,
            freq_table=freq_table,
            histogram_data=histogram_data,
            volatility_30d=volatility_30d,
            trend_direction=trend_direction,
            recommended_range_lo=p25,
            recommended_range_hi=p75,
        )

        from sqlalchemy.dialects.postgresql import insert as pg_insert
        from sqlalchemy.exc import IntegrityError
        stmt = pg_insert(AgencyStrategy).values(
            agency_id=agency_id,
            industry_code=industry_code,
            period_months=period_months,
            **upsert_data,
        ).on_conflict_do_update(
            constraint="agency_strategies_agency_id_industry_code_period_months_key",
            set_=upsert_data,
        )
        try:
            self.db.execute(stmt)
        except IntegrityError:
            self.db.rollback()

    @staticmethod
    def _serialize(row: AgencyStrategy) -> dict:
        return {
            "agency_id":          row.agency_id,
            "industry_code":      row.industry_code,
            "period_months":      row.period_months,
            "total_bid_count":    row.total_bid_count,
            "avg_win_rate":       float(row.avg_win_rate)       if row.avg_win_rate else None,
            "std_win_rate":       float(row.std_win_rate)       if row.std_win_rate else None,
            "min_win_rate":       float(row.min_win_rate)       if row.min_win_rate else None,
            "max_win_rate":       float(row.max_win_rate)       if row.max_win_rate else None,
            "win_rate_p10":       float(row.win_rate_p10)       if row.win_rate_p10 else None,
            "win_rate_p25":       float(row.win_rate_p25)       if row.win_rate_p25 else None,
            "win_rate_p50":       float(row.win_rate_p50)       if row.win_rate_p50 else None,
            "win_rate_p75":       float(row.win_rate_p75)       if row.win_rate_p75 else None,
            "win_rate_p90":       float(row.win_rate_p90)       if row.win_rate_p90 else None,
            "avg_competitor_cnt": float(row.avg_competitor_cnt) if row.avg_competitor_cnt else None,
            "aggression_index":   float(row.aggression_index)   if row.aggression_index else None,
            "qual_difficulty":    row.qual_difficulty,
            "freq_table":         row.freq_table or [],
            "histogram_data":     row.histogram_data or [],
            "volatility_30d":     float(row.volatility_30d)     if row.volatility_30d else None,
            "trend_direction":    row.trend_direction,
            "recommended_range_lo": float(row.recommended_range_lo) if row.recommended_range_lo else None,
            "recommended_range_hi": float(row.recommended_range_hi) if row.recommended_range_hi else None,
            "updated_at":         row.updated_at.isoformat() if row.updated_at else None,
        }


# ==================================================
# 자사 전용 경쟁사 서비스
# ==================================================

class OurCompetitorService:
    """자사가 자주 만나는 경쟁사 목록 관리"""

    def __init__(self, db: Session):
        self.db = db

    def list_competitors(self, limit: int = 30) -> list:
        rows = (
            self.db.query(OurCompetitor)
            .order_by(OurCompetitor.co_participation_cnt.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "id": r.id,
                "company_name": r.company_name,
                "biz_reg_no": r.biz_reg_no,
                "co_participation_cnt": r.co_participation_cnt,
                "co_win_cnt": r.co_win_cnt,
                "our_win_when_meet": r.our_win_when_meet,
                "avg_bid_rate": float(r.avg_bid_rate) if r.avg_bid_rate else None,
                "aggression": float(r.aggression) if r.aggression else None,
                "last_seen_at": str(r.last_seen_at) if r.last_seen_at else None,
                "last_seen_agency": r.last_seen_agency,
                "is_primary_rival": r.is_primary_rival,
            }
            for r in rows
        ]


# ==================================================
# 백테스트 엔진
# ==================================================

class BacktestService:
    """
    과거 투찰 이력(my_bid_records / bid_executions)과
    실제 낙찰 결과(bid_results)를 비교해 수주율 개선 추정.

    핵심 질문:
      "우리 시스템의 추천율로 투찰했다면 얼마나 더 낙찰됐을까?"
    """

    def __init__(self, db: Session):
        self.db = db

    def run(self, user_id: int, months: int = 60) -> dict:
        # 1) 결과가 확정된 내 투찰 기록
        my_records = self.db.execute(text("""
            SELECT
                e.id,
                e.title,
                e.agency_name,
                e.base_amount,
                e.submitted_rate,
                e.recommended_rate,
                e.winner_rate,
                e.floor_rate,
                e.status,
                e.result_rank,
                e.total_bidders,
                e.opened_at
            FROM bid_executions e
            WHERE e.user_id = :uid
              AND e.status IN ('낙찰', '패찰')
              AND e.submitted_rate IS NOT NULL
              AND e.winner_rate IS NOT NULL
              AND (e.opened_at IS NULL OR e.opened_at >= NOW() - INTERVAL :months_str)
            ORDER BY e.created_at DESC
            LIMIT 500
        """), {"uid": user_id, "months_str": f"{months} months"}).fetchall()

        if not my_records:
            # my_bid_records 폴백
            my_records = self.db.execute(text("""
                SELECT
                    m.id,
                    m.title,
                    m.agency_name,
                    m.base_amount,
                    m.submitted_rate,
                    m.recommendation_rate AS recommended_rate,
                    m.actual_winner_rate  AS winner_rate,
                    NULL AS floor_rate,
                    m.result AS status,
                    NULL AS result_rank,
                    NULL AS total_bidders,
                    m.bid_date AS opened_at
                FROM my_bid_records m
                WHERE m.user_id = :uid
                  AND m.result IN ('won', 'lost', '낙찰', '패찰')
                  AND m.submitted_rate IS NOT NULL
                  AND m.actual_winner_rate IS NOT NULL
                ORDER BY m.bid_date DESC
                LIMIT 500
            """), {"uid": user_id}).fetchall()

        if not my_records:
            return self._empty_result()

        cols = ["id", "title", "agency_name", "base_amount", "submitted_rate",
                "recommended_rate", "winner_rate", "floor_rate", "status",
                "result_rank", "total_bidders", "opened_at"]
        records = [dict(zip(cols, r)) for r in my_records]

        total = len(records)
        actual_wins = sum(1 for r in records if r["status"] in ("낙찰", "won"))
        actual_win_rate = round(actual_wins / total * 100, 1) if total else 0

        # 2) 추천율로 투찰했을 때 낙찰 시뮬레이션
        #    기준: submitted_rate <= winner_rate 이면 낙찰 가능 (하한율 이상)
        sim_wins = 0
        sim_won_list, sim_miss_list = [], []

        for r in records:
            winner = float(r["winner_rate"])
            floor = float(r["floor_rate"]) if r["floor_rate"] else 0.87745
            rec = float(r["recommended_rate"]) if r["recommended_rate"] else None
            actual = float(r["submitted_rate"])
            actual_won = r["status"] in ("낙찰", "won")

            if rec is None:
                sim_wins += 1 if actual_won else 0
                continue

            # 추천율이 하한율 이상이고 낙찰율 이하면 낙찰로 가정
            would_win = floor <= rec <= winner + 0.002  # 2bp 여유

            if would_win:
                sim_wins += 1
                if not actual_won:
                    sim_won_list.append({
                        "title": (r["title"] or "")[:40],
                        "agency": r["agency_name"] or "",
                        "actual_rate": round(actual * 100, 4),
                        "recommended_rate": round(rec * 100, 4),
                        "winner_rate": round(winner * 100, 4),
                        "gap_improvement": round((rec - actual) * 100, 4),
                    })

            if actual_won and not would_win:
                sim_miss_list.append({
                    "title": (r["title"] or "")[:40],
                    "agency": r["agency_name"] or "",
                    "actual_rate": round(actual * 100, 4),
                    "recommended_rate": round(rec * 100, 4),
                    "winner_rate": round(winner * 100, 4),
                })

        sim_win_rate = round(sim_wins / total * 100, 1) if total else 0
        improvement = round(sim_win_rate - actual_win_rate, 1)

        # 3) 패찰 원인 분포
        cause_dist: dict = {}
        for r in records:
            if r["status"] not in ("패찰", "lost"):
                continue
            winner = float(r["winner_rate"])
            actual = float(r["submitted_rate"])
            gap = actual - winner
            if gap > 0.005:
                cause = "투찰률과도"
            elif gap > 0.001:
                cause = "투찰률과도(미세)"
            elif r["total_bidders"] and r["total_bidders"] >= 15:
                cause = "경쟁과다"
            else:
                cause = "기타"
            cause_dist[cause] = cause_dist.get(cause, 0) + 1

        # 4) 월별 추이 (최근 12개월)
        monthly: dict = {}
        for r in records:
            dt = r["opened_at"]
            if dt is None:
                continue
            if hasattr(dt, "strftime"):
                key = dt.strftime("%Y-%m")
            else:
                try:
                    key = str(dt)[:7]
                except Exception:
                    continue
            if key not in monthly:
                monthly[key] = {"total": 0, "actual_win": 0, "sim_win": 0}
            monthly[key]["total"] += 1
            if r["status"] in ("낙찰", "won"):
                monthly[key]["actual_win"] += 1

        monthly_trend = [
            {
                "month": k,
                "total": v["total"],
                "actual_win": v["actual_win"],
                "actual_rate": round(v["actual_win"] / v["total"] * 100, 1) if v["total"] else 0,
            }
            for k, v in sorted(monthly.items())[-12:]
        ]

        return {
            "period_months": months,
            "total_bids": total,
            "actual_wins": actual_wins,
            "actual_win_rate": actual_win_rate,
            "simulated_wins": sim_wins,
            "simulated_win_rate": sim_win_rate,
            "improvement_pct": improvement,
            "cause_distribution": [
                {"cause": k, "count": v} for k, v in sorted(cause_dist.items(), key=lambda x: -x[1])
            ],
            "monthly_trend": monthly_trend,
            "sample_improvements": sim_won_list[:10],
            "sample_regressions": sim_miss_list[:5],
            "data_source": "bid_executions" if my_records else "my_bid_records",
        }

    def _empty_result(self) -> dict:
        return {
            "period_months": 0,
            "total_bids": 0,
            "actual_wins": 0,
            "actual_win_rate": 0,
            "simulated_wins": 0,
            "simulated_win_rate": 0,
            "improvement_pct": 0,
            "cause_distribution": [],
            "monthly_trend": [],
            "sample_improvements": [],
            "sample_regressions": [],
            "data_source": "none",
            "message": "SUCVIEW 파일을 업로드하거나 투찰 실행 관리에 결과를 입력해주세요.",
        }


class InpoNoticesSyncService:
    """
    inpo21c_bid_notices → bids 자동 동기화.

    G2B BidPublicInfoService02가 죽어있는 동안 info21c 입찰공고 사전정보를
    bids 테이블로 자동 전환하여 공고 파이프라인을 유지한다.
    매일 09:00 KST inpo21c 입찰공고 수집 직후 호출됨.
    """

    def sync(self, db: Session) -> dict:
        rows = db.execute(text("""
            SELECT n.inpo21c_bid_id, n.announcement_no, n.agency_name,
                   n.base_amount, n.open_datetime, n.min_bid_rate,
                   n.yega_method, n.yega_draw_count, n.yega_total_count,
                   n.region, n.industry
            FROM inpo21c_bid_notices n
            WHERE n.open_datetime > NOW() - INTERVAL '1 day'
              AND n.announcement_no IS NOT NULL
        """)).fetchall()

        created = skipped = 0
        for r in rows:
            # suffix 제거: "R26BK01531315-000" → "R26BK01531315"
            ann_no = re.sub(r'-\d+$', '', (r.announcement_no or '').strip())
            if not ann_no:
                skipped += 1
                continue

            existing = db.execute(
                text("SELECT id FROM bids WHERE announcement_no = :a"),
                {"a": ann_no}
            ).fetchone()
            if existing:
                skipped += 1
                continue

            try:
                agency_row = db.execute(
                    text("SELECT id FROM agencies WHERE name = :n"),
                    {"n": r.agency_name or ''}
                ).fetchone()
                if not agency_row:
                    db.execute(
                        text("INSERT INTO agencies (name) VALUES (:n) ON CONFLICT (name) DO NOTHING"),
                        {"n": r.agency_name or '미상'}
                    )
                    db.flush()
                    agency_row = db.execute(
                        text("SELECT id FROM agencies WHERE name = :n"),
                        {"n": r.agency_name or '미상'}
                    ).fetchone()

                yega_info = ""
                if r.yega_draw_count and r.yega_total_count:
                    yega_info = f"복수예가:{r.yega_draw_count}/{r.yega_total_count}"
                elif r.yega_method:
                    yega_info = r.yega_method[:100]

                db.execute(text("""
                    INSERT INTO bids
                        (announcement_no, title, agency_id, base_amount,
                         bid_open_date, min_bid_rate, status, source, bid_method)
                    VALUES
                        (:ann, :title, :agency_id, :base_amount,
                         :open_dt, :min_rate, 'open', 'inpo21c', :bid_method)
                    ON CONFLICT (announcement_no) DO NOTHING
                """), {
                    "ann":        ann_no,
                    "title":      f"[inpo21c] {r.agency_name or ''} ({ann_no})",
                    "agency_id":  agency_row[0],
                    "base_amount": r.base_amount or 0,
                    "open_dt":    r.open_datetime,
                    "min_rate":   r.min_bid_rate,
                    "bid_method": yega_info or None,
                })
                db.commit()
                created += 1
            except Exception as exc:
                db.rollback()
                logger.warning("inpo21c_bid_notices sync 실패 [%s]: %s", ann_no, exc)
                skipped += 1

        logger.info("inpo21c → bids 동기화: 신규=%d, 스킵=%d", created, skipped)
        return {"created": created, "skipped": skipped}


# DecisionService and JournalService extracted to dedicated modules.
# Re-exported here for backward compatibility with all existing router imports.
from .decision_service import DecisionService  # noqa: E402, F401
from .journal_service  import JournalService   # noqa: E402, F401

# ── kept below for reference during transition ──────────────────
class _DecisionService_REMOVED:
    """투찰 결정 전용 서비스 — TenderDecisionPage 백엔드."""

    def get_bid_context(self, db: Session, bid_id: int) -> dict:
        from .ml.a_value import calc_floor_rate
        from .ml.yega import load_inpo21c_yega_stats
        from .ml.competitor_predict import predict_bid_zone

        b = db.query(Bid).filter(Bid.id == bid_id).first()
        if not b:
            return {}

        industry_name = b.industry.name if b.industry else ""
        agency_id = b.agency_id or 0
        industry_id = b.industry_id or 0
        floor_rate = calc_floor_rate(industry_name)

        features = load_srate_stats(db, agency_id, industry_id, 0, b.base_amount)
        ep = predict_srate(features, b.base_amount)
        srate_center = ep["srate_range"]["center"]
        srate_std = (
            features.get("agency_srate_std")
            or features.get("global_srate_std")
            or 0.012
        )
        expected_n = int(
            features.get("expected_competitor_count")
            or features.get("global_comp_count")
            or 8
        )

        yega_stats = load_inpo21c_yega_stats(db, agency_id) if agency_id else {}
        pos_weights = yega_stats.get("pos_weights")

        competitor_zones: list = []
        try:
            czones = predict_bid_zone(db, agency_id, industry_id, b.base_amount)
            competitor_zones = czones if isinstance(czones, list) else []
        except Exception:
            pass

        from datetime import datetime as _dt, date as _date
        def _to_date(v):
            if v is None:
                return None
            if isinstance(v, _date) and not isinstance(v, _dt):
                return v
            if isinstance(v, _dt):
                return v.date()
            return v

        return {
            "bid_id":               b.id,
            "announcement_no":      b.announcement_no,
            "title":                b.title,
            "base_amount":          b.base_amount,
            "agency_id":            b.agency_id,
            "agency_name":          b.agency.name if b.agency else "",
            "industry_id":          b.industry_id,
            "industry_name":        industry_name,
            "floor_rate":           floor_rate,
            "a_value":              b.a_value,
            "srate_center":         round(srate_center, 4),
            "srate_std":            round(srate_std, 4),
            "expected_competitors": expected_n,
            "pos_weights":          pos_weights,
            "competitor_zones":     competitor_zones,
            "notice_date":          _to_date(b.notice_date),
            "bid_open_date":        _to_date(b.bid_open_date),
            "status":               b.status,
        }

    def simulate_bid(self, db: Session, bid_id: int, req) -> dict:
        from itertools import combinations as _comb
        import numpy as _np
        from .ml.a_value import calc_floor_rate
        from .ml.yega import load_inpo21c_yega_stats, calc_yega_frequency
        from .ml.rank_model import get_inpo_raw_rates
        from .ml.simulation import monte_carlo_win_prob_empirical, monte_carlo_win_prob

        b = db.query(Bid).filter(Bid.id == bid_id).first()
        if not b:
            return {}

        industry_name = b.industry.name if b.industry else ""
        agency_id = b.agency_id or 0
        industry_id = b.industry_id or 0
        base_amount = b.base_amount
        # base_amount=0인 공고는 시뮬레이션 불가 (수집 미완료)
        if not base_amount or base_amount <= 0:
            return {"error": "base_amount_missing", "bid_id": bid_id}
        floor_rate = calc_floor_rate(industry_name)

        features = load_srate_stats(db, agency_id, industry_id, 0, base_amount)
        ep = predict_srate(features, base_amount)
        srate_center = ep["srate_range"]["center"]
        srate_std = (
            features.get("agency_srate_std")
            or features.get("global_srate_std")
            or 0.012
        )

        yega_stats = load_inpo21c_yega_stats(db, agency_id) if agency_id else {}
        pos_weights = yega_stats.get("pos_weights")

        n_sim = min(max(req.n_sim, 5_000), 50_000)
        yega_values = req.yega_values

        # 실측 모드: 15개 모두 양수여야 함
        if yega_values and len(yega_values) == 15 and all(v > 0 for v in yega_values):
            mode = "real"
            srate_dist = simulate_yejung_from_real(yega_values, base_amount)
            candidates = [
                {"idx": i + 1, "amount": int(yega_values[i]), "rate": round(yega_values[i] / base_amount, 4)}
                for i in range(15)
            ]
            # C(15,4)=1365 전수 열거
            vals = _np.array(yega_values, dtype=_np.float64)
            combos = [(list(c), float(vals[list(c)].mean())) for c in _comb(range(15), 4)]
            # 예정가격에 가까운 순 상위 20개
            top_combos = sorted(combos, key=lambda x: abs(x[1] - base_amount * srate_center))[:20]
            top_combinations = [
                {
                    "combo": [i + 1 for i in c[0]],
                    "amount": int(round(c[1])),
                    "rate": round(c[1] / base_amount, 4),
                    "prob": round(1 / 1365, 5),
                }
                for c in top_combos
            ]
        else:
            mode = "estimated"
            rng = _np.random.default_rng(42)
            srate_dist = simulate_yejung(base_amount, srate_center, srate_std, n_sim, rng, pos_weights)
            yega_res = calc_yega_frequency(base_amount, b.a_value, srate_center)
            candidates = [
                {"idx": i + 1, "amount": int(c["amount"]), "rate": round(c["rate"], 4)}
                for i, c in enumerate(yega_res.get("candidates", []))
            ]
            top_combinations = [
                {
                    "combo": [],
                    "amount": int(t["amount"]),
                    "rate": round(t["rate"], 4),
                    "prob": round(t["probability"], 5),
                }
                for t in yega_res.get("top10", [])
            ]

        expected_n = int(
            features.get("expected_competitor_count")
            or features.get("global_comp_count")
            or 8
        )
        # 경쟁사 투찰률: 수동 입력 우선, 없으면 DB 조회
        inpo_rates = None
        manual_rates = req.competitor_rates if req.competitor_rates else None
        if manual_rates and len(manual_rates) >= 2:
            valid = [r for r in manual_rates if 0.80 <= r <= 1.00]
            if len(valid) >= 2:
                inpo_rates = valid
                expected_n = max(expected_n, len(valid))
        if inpo_rates is None:
            try:
                inpo_rates = get_inpo_raw_rates(db, expected_n)
            except Exception:
                pass
        # journal 낙찰자 투찰률로 보강 (기관별 실전 데이터)
        if inpo_rates is not None:
            try:
                from .ml.rank_model import get_journal_winner_rates
                _journal_rates = get_journal_winner_rates(db, agency_id)
                if _journal_rates is not None and len(_journal_rates) >= 5:
                    import numpy as _np2
                    _base = _np2.array(inpo_rates) if not isinstance(inpo_rates, _np2.ndarray) else inpo_rates
                    # journal 낙찰자율 30% 혼합 (실전 데이터 우선)
                    n_journal = min(len(_journal_rates), int(len(_base) * 0.3))
                    inpo_rates = _np2.concatenate([_base, _journal_rates[:n_journal]])
            except Exception:
                pass

        all_zones, top_zones = scan_zones_from_dist(
            srate_dist, floor_rate, base_amount,
            inpo_rates, expected_n, n_sim=min(n_sim, 20_000),
        )

        eff_floor = floor_rate * float(_np.median(srate_dist))
        rate_agg  = round(eff_floor + 0.0003, 4)
        rate_bal  = round(max(eff_floor + 0.0015, rate_agg + 0.0005), 4)
        rate_con  = round(max(eff_floor + 0.003,  rate_bal + 0.0005), 4)
        rng2 = _np.random.default_rng(42)

        def _wp(r):
            if inpo_rates is not None and len(inpo_rates) >= 5 and expected_n > 0:
                return monte_carlo_win_prob_empirical(r, floor_rate, srate_dist, inpo_rates, expected_n, n_sim, rng2)
            return monte_carlo_win_prob(r, floor_rate, srate_dist, [], [], n_sim, rng2)

        strategies = {
            "aggressive":   {**_wp(rate_agg), "rate": rate_agg, "amount": int(round(rate_agg * base_amount)), "label": "공격형"},
            "balanced":     {**_wp(rate_bal), "rate": rate_bal, "amount": int(round(rate_bal * base_amount)), "label": "균형형"},
            "conservative": {**_wp(rate_con), "rate": rate_con, "amount": int(round(rate_con * base_amount)), "label": "안정형"},
        }

        best = max(top_zones, key=lambda z: z["win_prob"]) if top_zones else None
        if best:
            optimal = {
                "rate":     best["rate"],
                "amount":   best["amount"],
                "win_prob": best["win_prob"],
                "srate":    round(float(_np.median(srate_dist)), 4),
                "floor_ok": best.get("floor_ok", True),
            }
        else:
            optimal = {}

        hist_counts, bin_edges = _np.histogram(srate_dist, bins=30)
        histogram = [
            {
                "bin_center": round(float((bin_edges[i] + bin_edges[i + 1]) / 2), 4),
                "count":      int(hist_counts[i]),
                "prob":       round(float(hist_counts[i]) / len(srate_dist), 4),
            }
            for i in range(len(hist_counts))
        ]

        # prediction_logs_v2에 이번 추천 저장 (bid_id 반드시 기록)
        pred_log_id = None
        try:
            from .models import PredictionLogV2
            from .ml.engine import get_model_meta
            meta = get_model_meta()
            plog = PredictionLogV2(
                bid_id=bid_id,
                model_version=meta.get("version"),
                srate_pred_center=round(srate_center, 6),
                rate_aggressive=strategies["aggressive"]["rate"],
                rate_balanced=strategies["balanced"]["rate"],
                rate_conservative=strategies["conservative"]["rate"],
                win_prob_center=strategies["balanced"].get("win_prob"),
            )
            db.add(plog)
            db.commit()
            db.refresh(plog)
            pred_log_id = plog.id
        except Exception as _e:
            logger.warning(f"prediction_logs_v2 저장 실패: {_e}")

        return {
            "bid_id":           bid_id,
            "base_amount":      base_amount,
            "floor_rate":       floor_rate,
            "srate_center":     round(srate_center, 4),
            "srate_std":        round(srate_std, 4),
            "mode":             mode,
            "pred_log_id":      pred_log_id,
            "yega_candidates":  candidates,
            "top_combinations": top_combinations,
            "all_zones":        [dict(z) for z in all_zones],
            "top_zones":        [dict(z) for z in top_zones],
            "strategies":       strategies,
            "optimal":          optimal,
            "histogram":        histogram,
        }


# ─────────────────────────────────────────────────────────────
# JournalService extracted → app/journal_service.py (re-exported above)
# ─────────────────────────────────────────────────────────────

class _JournalService_REMOVED:
    """
    투찰 의사결정 → 실제투찰 → 개찰결과 를 bid_journal에 기록.
    prediction_logs_v2와 연결해 모델 성능을 실측으로 검증한다.
    """

    def create(self, db: Session, user_id: int, req) -> "BidJournal":
        from .models import BidJournal, Bid
        from .schemas import JournalCreateRequest
        from datetime import datetime as _dt

        b = db.query(Bid).filter(Bid.id == req.bid_id).first()
        ann_no = b.announcement_no if b else None

        rate_delta = None
        if req.submitted_rate and req.recommended_rate:
            rate_delta = round(float(req.submitted_rate) - float(req.recommended_rate), 6)

        submitted_amount = req.submitted_amount
        if not submitted_amount and req.submitted_rate and b and b.base_amount:
            submitted_amount = int(round(float(req.submitted_rate) * b.base_amount))

        obj = BidJournal(
            bid_id=req.bid_id,
            user_id=user_id,
            announcement_no=ann_no,
            predicted_at=_dt.now() if req.pred_log_id else None,
            pred_log_id=req.pred_log_id,
            recommended_rate=req.recommended_rate,
            recommended_amount=req.recommended_amount,
            pred_win_prob=req.pred_win_prob,
            pred_srate_center=req.pred_srate_center,
            strategy_chosen=req.strategy_chosen,
            submitted_at=_dt.now(),
            submitted_rate=req.submitted_rate,
            submitted_amount=submitted_amount,
            floor_rate=req.floor_rate,
            rate_delta=rate_delta,
            note=req.note,
        )
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    def record_result(self, db: Session, journal_id: int, user_id: int, req) -> "BidJournal":
        from .models import BidJournal
        from datetime import datetime as _dt

        VALID_RESULTS = {"낙찰", "패찰", "무효", "취소"}
        if req.result not in VALID_RESULTS:
            from fastapi import HTTPException
            raise HTTPException(400, f"result는 {VALID_RESULTS} 중 하나여야 합니다.")

        obj = db.query(BidJournal).filter(
            BidJournal.id == journal_id,
            BidJournal.user_id == user_id,
        ).first()
        if not obj:
            from fastapi import HTTPException
            raise HTTPException(404, "저널 레코드를 찾을 수 없습니다.")

        obj.result = req.result
        obj.opened_at = _dt.now()
        if req.actual_srate is not None:
            obj.actual_srate = req.actual_srate
        if req.our_rank is not None:
            obj.our_rank = req.our_rank
        if req.total_bidders is not None:
            obj.total_bidders = req.total_bidders
        if req.winner_rate is not None:
            obj.winner_rate = req.winner_rate
        if req.winner_amount is not None:
            obj.winner_amount = req.winner_amount
        if req.winner_biz_no is not None:
            obj.winner_biz_no = req.winner_biz_no
        if req.winner_name is not None:
            obj.winner_name = req.winner_name
        if req.note is not None:
            obj.note = req.note

        # 파생 필드 계산
        if obj.winner_rate and obj.submitted_rate:
            obj.rate_gap = round(float(obj.winner_rate) - float(obj.submitted_rate), 6)
        if obj.actual_srate and obj.pred_srate_center:
            obj.srate_error = round(float(obj.actual_srate) - float(obj.pred_srate_center), 6)

        db.commit()
        db.refresh(obj)
        return obj

    def list_journals(self, db: Session, user_id: int, result_filter: str = None,
                      page: int = 1, size: int = 20) -> dict:
        from .models import BidJournal
        from .schemas import JournalOut

        q = db.query(BidJournal).filter(BidJournal.user_id == user_id)
        if result_filter:
            if result_filter == "pending":
                q = q.filter(BidJournal.result == None)  # noqa: E711
            else:
                q = q.filter(BidJournal.result == result_filter)

        total = q.count()
        items = q.order_by(BidJournal.created_at.desc()).offset((page - 1) * size).limit(size).all()
        return {
            "total": total,
            "page": page,
            "size": size,
            "items": [JournalOut.model_validate(i).model_dump() for i in items],
        }

    def get_stats(self, db: Session, user_id: int) -> dict:
        """피드백 루프 현황 + 모델 성능 지표."""
        from sqlalchemy import text as _text

        rows = db.execute(_text("""
            SELECT
              COUNT(*)                                              AS total,
              COUNT(CASE WHEN result IS NOT NULL  THEN 1 END)      AS with_result,
              COUNT(CASE WHEN result = '낙찰'    THEN 1 END)      AS wins,
              COUNT(CASE WHEN result = '패찰'    THEN 1 END)      AS losses,
              COUNT(CASE WHEN result IS NULL
                          AND submitted_rate IS NOT NULL THEN 1 END) AS pending_result,
              ROUND(AVG(CASE WHEN result IS NOT NULL
                             THEN ABS(srate_error) END)::numeric, 4) AS avg_srate_mae,
              ROUND(AVG(CASE WHEN result = '패찰'
                             THEN rate_gap END)::numeric, 4)      AS avg_rate_gap_loss,
              ROUND(AVG(CASE WHEN result IS NOT NULL
                             THEN rate_delta END)::numeric, 4)    AS avg_rate_delta
            FROM bid_journal
            WHERE user_id = :uid
        """), {"uid": user_id}).fetchone()

        total      = rows[0] or 0
        with_result= rows[1] or 0
        wins       = rows[2] or 0
        losses     = rows[3] or 0
        pending    = rows[4] or 0
        mae        = float(rows[5]) if rows[5] else None
        avg_gap    = float(rows[6]) if rows[6] else None
        avg_delta  = float(rows[7]) if rows[7] else None

        win_rate = round(wins / with_result, 4) if with_result > 0 else None

        return {
            "total":            total,
            "with_result":      with_result,
            "pending_result":   pending,
            "wins":             wins,
            "losses":           losses,
            "win_rate":         win_rate,
            "avg_srate_mae":    mae,
            "avg_rate_gap_loss": avg_gap,   # 패찰 시 낙찰자와 우리 투찰률 차이
            "avg_rate_delta":   avg_delta,   # AI추천 vs 실제 투찰 차이
            "feedback_completeness": round(with_result / total, 4) if total > 0 else 0,
        }


