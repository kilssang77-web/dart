"""
투찰 실행 서비스
"""
import io
import re
import math
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from sqlalchemy.orm import Session
from sqlalchemy import text, func, and_, or_, desc

from ..models import (
    Bid, BidResult, Competitor, Agency, Industry, Region,
    FeatureStore, PredictionLog, PredictionLogV2, CompetitorStat, User, AuditLog,
    IndustryFilter, BidBookmark, CollectionLog, MyBidRecord, Notification,
    BidExecution, DefeatAnalysis, AgencyStrategy, RateFrequencyTable, OurCompetitor,
    ActualBidOutcome, ModelPerformanceLog, BidJournal,
)
from ..schemas import (
    BidCreate, BidResultCreate, RecommendRequest, RecommendResponse,
    RateRange, WinProbabilities, Explanation, ExplanationFactor, RiskInfo,
    SimilarCase, BidSummary, BidDetail, BidResultOut, CompetitorDetail
)
from ..ml.engine import build_features, get_engine, FEATURE_LABELS
from ..ml.assessment  import load_srate_stats, predict_srate, compute_market_trend
from ..ml.competition import compute_competition_features, get_competitor_profiles, get_market_competitor_distributions
from ..ml.simulation  import recommend_with_simulation, simulate_yejung_from_real, simulate_yejung, scan_zones_from_dist
from ..ml.rank_model  import get_inpo_raw_rates
from ..ml.personal    import PersonalBiasAnalyzer
from ..ml.prism       import scan_prism_zones
from ..ml.yega        import calc_yega_frequency, get_inpo21c_pattern_direct, load_inpo21c_yega_stats
from ..ml.a_value     import calc_floor_rate

from ._common import get_active_industry_ids, _build_ind_sql, _compute_yega_ml_features

logger = logging.getLogger(__name__)

class ExecutionService:
    """투찰 수명주기 관리 + SUCVIEW/인포 엑셀 파싱"""

    STATUS_ORDER = ["검토중", "참여결정", "투찰완료", "개찰대기", "낙찰", "패찰", "포기"]

    def __init__(self, db: Session):
        self.db = db

    # ── CRUD ─────────────────────────────────────────────────

    def list_executions(self, user_id: int, status: str = None, page: int = 1, size: int = 20) -> dict:
        from ..schemas import BidExecutionOut
        q = self.db.query(BidExecution).filter(BidExecution.user_id == user_id)
        if status:
            q = q.filter(BidExecution.status == status)
        total = q.count()
        items = q.order_by(BidExecution.created_at.desc()).offset((page - 1) * size).limit(size).all()
        return {
            "total": total, "page": page, "size": size,
            "items": [BidExecutionOut.model_validate(i).model_dump() for i in items],
        }

    def get_summary(self, user_id: int) -> dict:
        from sqlalchemy import func as sqlfunc
        rows = (
            self.db.query(BidExecution.status, sqlfunc.count().label("cnt"))
            .filter(BidExecution.user_id == user_id)
            .group_by(BidExecution.status)
            .all()
        )
        summary = {s: 0 for s in self.STATUS_ORDER}
        for status, cnt in rows:
            summary[status] = cnt

        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start.replace(hour=23, minute=59, second=59)
        today_closing = (
            self.db.query(BidExecution)
            .filter(
                BidExecution.user_id == user_id,
                BidExecution.status.in_(["참여결정", "투찰완료"]),
                BidExecution.bid_open_date >= today_start,
                BidExecution.bid_open_date <= today_end,
            )
            .all()
        )
        from ..schemas import BidExecutionOut
        return {
            "status_counts": summary,
            "today_closing": [BidExecutionOut.model_validate(e).model_dump() for e in today_closing],
        }

    def get(self, exec_id: int):
        return self.db.query(BidExecution).filter(BidExecution.id == exec_id).first()

    def create(self, user_id: int, data) -> BidExecution:
        obj = BidExecution(user_id=user_id, **data.model_dump(exclude_none=True))
        self.db.add(obj)
        self.db.commit()
        self.db.refresh(obj)
        return obj

    def update(self, exec_id: int, user_id: int, data) -> BidExecution:
        obj = self.db.query(BidExecution).filter(
            BidExecution.id == exec_id,
            BidExecution.user_id == user_id,
        ).first()
        if not obj:
            from fastapi import HTTPException
            raise HTTPException(404, "Not found")
        updates = data.model_dump(exclude_none=True)
        for k, v in updates.items():
            setattr(obj, k, v)
        # 패찰 → 자동 원인 분석 + 알림
        if data.status == "패찰":
            self._auto_defeat_analysis(obj)
            da = self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == obj.id).first()
            cause_str = da.cause_primary if da else "기타"
            improvement_str = da.improvement if da else ""
            from .notifications import NotificationService
            nsvc = NotificationService(self.db)
            nsvc.create(
                user_id=obj.user_id,
                ntype="execution_result",
                title=f"[패찰] {obj.title[:45]}",
                body=f"원인: {cause_str}. {improvement_str}",
                link="/executions",
            )
        # 낙찰 → 축하 알림
        elif data.status == "낙찰":
            from .notifications import NotificationService
            nsvc = NotificationService(self.db)
            nsvc.create(
                user_id=obj.user_id,
                ntype="execution_result",
                title=f"[낙찰] {obj.title[:45]}",
                body=f"축하합니다! 낙찰 확정되었습니다. 투찰률: {obj.submitted_rate:.3%}" if obj.submitted_rate else "낙찰 확정",
                link="/executions",
            )
        # 낙찰/패찰 확정 시 actual_bid_outcomes + bid_journal 자동 동기화
        if data.status in ("낙찰", "패찰") and obj.submitted_rate:
            self._sync_actual_bid_outcome(obj)
            self._create_journal_from_execution(obj, obj.user_id)
        self.db.commit()
        self.db.refresh(obj)
        return obj

    def _sync_actual_bid_outcome(self, obj: "BidExecution"):
        """낙찰/패찰 확정 시 actual_bid_outcomes에 자동 upsert — 피드백 루프 핵심."""
        result_code = "WON" if obj.status == "낙찰" else "LOST"

        # bid_journal에서 예측값 조회 (있으면 정확도 추적)
        journal = None
        if obj.announcement_no:
            journal = (
                self.db.query(BidJournal)
                .filter(BidJournal.announcement_no == obj.announcement_no)
                .order_by(BidJournal.id.desc())
                .first()
            )
        if journal is None and obj.bid_id:
            journal = (
                self.db.query(BidJournal)
                .filter(BidJournal.bid_id == obj.bid_id)
                .order_by(BidJournal.id.desc())
                .first()
            )

        predicted_wp     = float(journal.pred_win_prob)    if journal and journal.pred_win_prob    else None
        predicted_srate  = float(journal.pred_srate_center) if journal and journal.pred_srate_center else None
        actual_srate     = float(obj.winner_rate) if obj.winner_rate else None  # 복수예가: 낙찰자 투찰률 ≈ 사정율
        srate_err        = abs(predicted_srate - actual_srate) if (predicted_srate and actual_srate) else None

        existing = None
        if obj.announcement_no:
            existing = (
                self.db.query(ActualBidOutcome)
                .filter(ActualBidOutcome.announcement_no == obj.announcement_no)
                .first()
            )
        if existing is None and obj.bid_id:
            existing = (
                self.db.query(ActualBidOutcome)
                .filter(
                    ActualBidOutcome.bid_id == obj.bid_id,
                    ActualBidOutcome.user_id == obj.user_id,
                )
                .first()
            )

        if existing:
            existing.result             = result_code
            existing.submitted_rate     = obj.submitted_rate
            existing.actual_srate       = actual_srate
            existing.winner_rate        = float(obj.winner_rate) if obj.winner_rate else None
            existing.winner_biz_no      = obj.winner_biz_no
            existing.our_rank           = obj.result_rank
            existing.total_bidders      = obj.total_bidders
            existing.predicted_win_prob = predicted_wp
            existing.predicted_srate    = predicted_srate
            existing.srate_error        = srate_err
            existing.collected_at       = obj.opened_at
        else:
            outcome = ActualBidOutcome(
                bid_id              = obj.bid_id,
                user_id             = obj.user_id,
                announcement_no     = obj.announcement_no,
                submitted_rate      = obj.submitted_rate,
                result              = result_code,
                actual_srate        = actual_srate,
                winner_rate         = float(obj.winner_rate) if obj.winner_rate else None,
                winner_biz_no       = obj.winner_biz_no,
                our_rank            = obj.result_rank,
                total_bidders       = obj.total_bidders,
                predicted_win_prob  = predicted_wp,
                predicted_srate     = predicted_srate,
                srate_error         = srate_err,
                collected_at        = obj.opened_at,
            )
            self.db.add(outcome)

    def delete(self, exec_id: int, user_id: int):
        obj = self.db.query(BidExecution).filter(
            BidExecution.id == exec_id,
            BidExecution.user_id == user_id,
        ).first()
        if obj:
            # defeat_analyses has NOT NULL FK — delete child rows first
            self.db.query(DefeatAnalysis).filter(
                DefeatAnalysis.execution_id == exec_id
            ).delete(synchronize_session=False)
            self.db.delete(obj)
            self.db.commit()

    # ── 패찰 원인 분석 ────────────────────────────────────────

    def get_defeat_analysis(self, exec_id: int):
        return self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == exec_id).first()

    def _auto_defeat_analysis(self, obj: BidExecution):
        existing = self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == obj.id).first()
        if existing:
            return

        cause, detail, adj = "기타", "", 0.0
        winner_gap_pct = None

        if obj.submitted_rate and obj.winner_rate:
            gap = float(obj.submitted_rate) - float(obj.winner_rate)
            winner_gap_pct = round(gap * 100, 3)
            if gap > 0.005:
                cause = "투찰률과도"
                detail = f"투찰률이 낙찰률보다 {gap*100:.3f}%p 높았음"
                adj = round(-gap * 0.6, 4)
            elif gap > 0.001:
                cause = "투찰률과도"
                detail = f"투찰률이 낙찰률보다 {gap*100:.3f}%p 높았음 (미세)"
                adj = round(-gap * 0.5, 4)

        if cause == "기타" and obj.total_bidders and obj.total_bidders >= 15:
            cause = "경쟁사과다"
            detail = f"참여업체 {obj.total_bidders}개사 — 경쟁 과열"

        improvement = "다음 동일기관 입찰 시 " + (
            f"{abs(adj)*100:.2f}%p 낮게 조정 권장" if adj < 0 else "현 전략 유지"
        )

        da = DefeatAnalysis(
            execution_id=obj.id,
            cause_primary=cause,
            cause_detail=detail,
            winner_gap_pct=winner_gap_pct,
            competitor_cnt=obj.total_bidders,
            our_rank=obj.result_rank,
            floor_rate=obj.floor_rate,
            improvement=improvement,
            next_rate_adj=adj,
        )
        self.db.add(da)

    # ── SUCVIEW 엑셀 파싱 ─────────────────────────────────────

    def import_sucview(self, file_bytes: bytes, user_id: int):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        imported, skipped, competitors_added, journal_created = 0, 0, 0, 0
        errors, details = [], []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                result = self._parse_sucview_sheet(ws, user_id)
                imported += result["imported"]
                skipped += result["skipped"]
                competitors_added += result["competitors_added"]
                journal_created += result.get("journal_created", 0)
                details.extend(result["details"])
            except Exception as e:
                errors.append(f"시트 {sheet_name}: {e}")

        self.db.commit()
        return {
            "imported": imported,
            "skipped": skipped,
            "competitors_added": competitors_added,
            "journal_created": journal_created,
            "errors": errors,
            "details": details,
        }

    def _parse_sucview_sheet(self, ws, user_id: int) -> dict:
        data = {}
        participants = []
        in_participant = False

        for row in ws.iter_rows(values_only=True):
            if not any(c for c in row if c is not None):
                continue
            label = str(row[0] or "").strip()
            val = row[1] if len(row) > 1 else None

            if label == "공고번호":
                data["announcement_no"] = str(val or "").strip()
            elif label == "공고명":
                data["title"] = str(val or "").strip()
            elif label == "발주기관":
                data["agency_name"] = str(val or "").strip()
            elif label in ("낙찰하한율", "사정율하한"):
                try:
                    data["floor_rate"] = float(val)
                except Exception:
                    pass
            elif label in ("기초금액", "예정가격"):
                try:
                    data["base_amount"] = int(float(str(val or "0").replace(",", "")))
                except Exception:
                    pass
            elif label == "개찰일시":
                if val:
                    try:
                        data["opened_at"] = val if isinstance(val, datetime) else datetime.fromisoformat(str(val))
                    except Exception:
                        pass
            elif label in ("순위", "No", "번호") or (str(label).isdigit()):
                in_participant = True

            if in_participant and len(row) >= 5:
                try:
                    rank = int(row[0]) if row[0] else None
                    biz_no = str(row[1] or "").strip()
                    name = str(row[2] or "").strip()
                    amount_raw = str(row[3] or "0").replace(",", "")
                    rate_raw = row[4]
                    if rank and name and len(name) > 1:
                        participants.append({
                            "rank": rank,
                            "biz_no": biz_no,
                            "company_name": name,
                            "amount": int(float(amount_raw)) if amount_raw else None,
                            "rate": float(rate_raw) if rate_raw else None,
                        })
                except Exception:
                    pass

        if not data.get("title"):
            return {"imported": 0, "skipped": 1, "competitors_added": 0, "details": []}

        # 낙찰자 추출
        winner = next((p for p in participants if p["rank"] == 1), None)
        if winner:
            data["winner_name"] = winner["company_name"]
            data["winner_biz_no"] = winner["biz_no"]
            data["winner_amount"] = winner["amount"]
            data["winner_rate"] = winner["rate"]
            data["total_bidders"] = len(participants)

        # 중복 체크
        existing = self.db.query(BidExecution).filter(
            BidExecution.announcement_no == data.get("announcement_no"),
            BidExecution.user_id == user_id,
        ).first()
        if existing:
            return {"imported": 0, "skipped": 1, "competitors_added": 0, "details": [f"중복: {data.get('title', '')}"][:1]}

        data["status"] = "개찰대기" if not winner else "패찰"
        data["source"] = "sucview"
        data["sucview_raw"] = participants
        data["user_id"] = user_id

        exec_obj = BidExecution(**data)
        self.db.add(exec_obj)
        self.db.flush()

        comp_added = self._update_our_competitors_raw(participants)
        j_created = 1 if self._create_journal_from_execution(exec_obj, user_id) else 0
        detail_msg = f"가져옴: {data.get('title', '')[:40]}"
        if j_created:
            detail_msg += " [저널생성]"
        return {"imported": 1, "skipped": 0, "competitors_added": comp_added, "journal_created": j_created, "details": [detail_msg]}

    def _create_journal_from_execution(self, exec_obj: "BidExecution", user_id: int) -> bool:
        """SUCVIEW/인포 import 후 bid_journal 레코드 자동 생성 (피드백 루프용)."""
        if not exec_obj.announcement_no:
            return False

        bid = self.db.query(Bid).filter(Bid.announcement_no == exec_obj.announcement_no).first()
        if not bid:
            return False

        existing = self.db.query(BidJournal).filter(
            BidJournal.bid_id == bid.id,
            BidJournal.user_id == user_id,
        ).first()
        if existing:
            return False

        result = "낙찰" if exec_obj.status == "낙찰" else "패찰"
        winner_rate_val = float(exec_obj.winner_rate) if exec_obj.winner_rate else None
        submitted_val = float(exec_obj.submitted_rate) if exec_obj.submitted_rate else None
        rate_gap = round(winner_rate_val - submitted_val, 6) if (winner_rate_val and submitted_val) else None

        journal = BidJournal(
            bid_id=bid.id,
            user_id=user_id,
            announcement_no=exec_obj.announcement_no,
            opened_at=exec_obj.opened_at,
            result=result,
            submitted_rate=exec_obj.submitted_rate,
            floor_rate=exec_obj.floor_rate,
            actual_srate=exec_obj.floor_rate,
            winner_rate=exec_obj.winner_rate,
            winner_amount=exec_obj.winner_amount,
            winner_biz_no=exec_obj.winner_biz_no,
            winner_name=exec_obj.winner_name,
            total_bidders=exec_obj.total_bidders,
            rate_gap=rate_gap,
        )
        self.db.add(journal)
        return True

    def _update_our_competitors_raw(self, participants: list) -> int:
        added = 0
        for p in participants:
            if not p.get("company_name"):
                continue
            existing = self.db.query(OurCompetitor).filter(
                OurCompetitor.company_name == p["company_name"]
            ).first()
            if existing:
                existing.co_participation_cnt = (existing.co_participation_cnt or 0) + 1
                if p.get("rank") == 1:
                    existing.co_win_cnt = (existing.co_win_cnt or 0) + 1
                if p.get("rate"):
                    existing.avg_bid_rate = p["rate"]
                existing.last_seen_at = datetime.now().date()
                existing.updated_at = datetime.now()
            else:
                new_comp = OurCompetitor(
                    company_name=p["company_name"],
                    biz_reg_no=p.get("biz_no"),
                    co_participation_cnt=1,
                    co_win_cnt=1 if p.get("rank") == 1 else 0,
                    avg_bid_rate=p.get("rate"),
                    last_seen_at=datetime.now().date(),
                    source="sucview",
                )
                self.db.add(new_comp)
                added += 1
        return added

    def import_inpo_history(self, file_bytes: bytes, user_id: int):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        imported, skipped, journal_created = 0, 0, 0
        errors, details = [], []

        headers = []
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [str(c or "").strip() for c in row]
                continue
            row_dict = dict(zip(headers, row))
            try:
                announcement_no = str(row_dict.get("공고번호") or "").strip()
                title = str(row_dict.get("공고명") or "").strip()
                if not title:
                    skipped += 1
                    continue

                existing = self.db.query(BidExecution).filter(
                    BidExecution.announcement_no == announcement_no,
                    BidExecution.user_id == user_id,
                ).first() if announcement_no else None

                if existing:
                    skipped += 1
                    continue

                bid_rate_raw = row_dict.get("투찰율") or row_dict.get("투찰률")
                bid_rate = None
                try:
                    bid_rate = float(str(bid_rate_raw).replace("%", "")) / 100 if bid_rate_raw else None
                except Exception:
                    pass

                result_raw = str(row_dict.get("결과") or row_dict.get("낙패") or "").strip()
                status = "낙찰" if "낙찰" in result_raw else ("패찰" if "패찰" in result_raw else "개찰대기")

                base_raw = str(row_dict.get("기초금액") or row_dict.get("예정가격") or "0").replace(",", "")
                try:
                    base_amount = int(float(base_raw))
                except Exception:
                    base_amount = 0

                exec_obj = BidExecution(
                    user_id=user_id,
                    announcement_no=announcement_no,
                    title=title,
                    agency_name=str(row_dict.get("발주기관") or ""),
                    base_amount=base_amount,
                    submitted_rate=bid_rate,
                    status=status,
                    source="inpo_history",
                )
                self.db.add(exec_obj)
                self.db.flush()
                if status in ("낙찰", "패찰"):
                    if self._create_journal_from_execution(exec_obj, user_id):
                        journal_created += 1
                imported += 1
                details.append(f"가져옴: {title[:40]}")
            except Exception as e:
                errors.append(f"행 오류: {e}")
                skipped += 1

        self.db.commit()
        return {
            "imported": imported,
            "skipped": skipped,
            "competitors_added": 0,
            "journal_created": journal_created,
            "errors": errors,
            "details": details,
        }

    def defeat_summary(self, user_id: int) -> dict:
        """패찰 이력 집계 분석 — 원인별 건수, 평균 gap, 권장 조정량."""
        defeats = (
            self.db.query(BidExecution)
            .filter(BidExecution.user_id == user_id, BidExecution.status == "패찰")
            .all()
        )
        total_defeats = len(defeats)
        if total_defeats == 0:
            return {"total_defeats": 0, "by_cause": [], "avg_winner_gap_pct": None, "avg_rate_adj": None, "recent": []}

        cause_map: dict = {}
        gaps, adjs = [], []

        for ex in defeats:
            da = self.db.query(DefeatAnalysis).filter(DefeatAnalysis.execution_id == ex.id).first()
            cause = da.cause_primary if da else "미분석"
            cause_map[cause] = cause_map.get(cause, {"count": 0, "gap_sum": 0.0, "adj_sum": 0.0})
            cause_map[cause]["count"] += 1
            if da and da.winner_gap_pct is not None:
                cause_map[cause]["gap_sum"] += float(da.winner_gap_pct)
                gaps.append(float(da.winner_gap_pct))
            if da and da.next_rate_adj is not None:
                cause_map[cause]["adj_sum"] += float(da.next_rate_adj)
                adjs.append(float(da.next_rate_adj))

        by_cause = [
            {
                "cause":        cause,
                "count":        v["count"],
                "pct":          round(v["count"] / total_defeats * 100, 1),
                "avg_gap_pct":  round(v["gap_sum"] / v["count"], 3) if v["count"] else None,
                "avg_rate_adj": round(v["adj_sum"] / v["count"], 4) if v["count"] else None,
            }
            for cause, v in sorted(cause_map.items(), key=lambda x: -x[1]["count"])
        ]

        recent_rows = (
            self.db.query(BidExecution)
            .filter(BidExecution.user_id == user_id, BidExecution.status == "패찰")
            .order_by(BidExecution.bid_open_date.desc().nullslast())
            .limit(10)
            .all()
        )
        recent = [
            {
                "id":             ex.id,
                "title":          ex.title,
                "bid_open_date":  ex.bid_open_date.isoformat() if ex.bid_open_date else None,
                "submitted_rate": float(ex.submitted_rate) if ex.submitted_rate else None,
                "winner_rate":    float(ex.winner_rate) if ex.winner_rate else None,
                "gap_pct":        round((float(ex.submitted_rate) - float(ex.winner_rate)) * 100, 3)
                                  if ex.submitted_rate and ex.winner_rate else None,
            }
            for ex in recent_rows
        ]

        return {
            "total_defeats":     total_defeats,
            "by_cause":          by_cause,
            "avg_winner_gap_pct": round(sum(gaps) / len(gaps), 3) if gaps else None,
            "avg_rate_adj":       round(sum(adjs) / len(adjs), 4) if adjs else None,
            "recent":             recent,
        }


# ==================================================
# 발주기관 빈도표 서비스
# ==================================================
